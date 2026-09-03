import os, json, csv, time, threading, subprocess, sys, signal, logging
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
COMMON_DIR = os.path.join(BASE_DIR, "common")
for p in [BASE_DIR, COMMON_DIR]:
    if p not in sys.path:
        sys.path.insert(0, p)
import paths
from datetime import datetime as dt, time as datetime_time, timedelta
from flask import Flask, render_template_string, jsonify, request, Response, session, redirect
from kiteconnect import KiteConnect
import trade_db
import dashboard_auth
from dashboard_sl_overrides import write_sl_overrides
from trading_core import (
    lookup_scan_sl_target,
    derive_sl_targets_for_contract,
    load_kite_session,
    close_position as shared_close_position,
    close_stock_position as shared_close_stock_position,
    clear_executed_exit,
    is_contract_exit_executed,
    log_to_journal,
    fetch_and_resample_candles,
    calc_rr,
    get_ist_now
)
from ema_engine import (
    start_ema_engine, stop_ema_engine, get_ema_engine_status, get_ema_scan_data,
    EMA_DISPLAY_FILE_STOCK
)

def resolve_underlying(contract_or_symbol, engine="nifty50"):
    """Return the real underlying registry symbol for a contract string.

    Fixes the stale-ACTIVE anomaly where the DB `symbol` was stored as the full
    contract string instead of the underlying. Falls back to raw input.
    """
    from trading_core import INDEX_REGISTRY, STOCK_REGISTRY
    raw = str(contract_or_symbol or "").replace(" ", "").upper()
    if not raw:
        return contract_or_symbol or ""
    reg = INDEX_REGISTRY if engine == "index" else STOCK_REGISTRY
    for sym in sorted(reg.keys(), key=len, reverse=True):
        if sym.replace(" ", "").upper() in raw:
            return sym
    for sym in sorted(STOCK_REGISTRY.keys(), key=len, reverse=True):
        if sym.replace(" ", "").upper() in raw:
            return sym
    return contract_or_symbol or ""

app = Flask(__name__)
app.secret_key = dashboard_auth.get_secret_key()

from blueprints.routes_token import token_bp
app.register_blueprint(token_bp)

# ──────────────────────────────────────────────
#  FILE PATHS & DASHBOARD CONFIG
# ──────────────────────────────────────────────


def get_kite_credentials():
    """Read Kite API key/secret from environment, config, or token file."""
    api_key = os.environ.get("KITE_API_KEY", "")
    api_secret = os.environ.get("KITE_API_SECRET", "")
    cfg = load_config()
    if not api_key: api_key = cfg.get("api_key", "")
    if not api_secret: api_secret = cfg.get("api_secret", "")
    if (not api_key or not api_secret) and os.path.exists(TOKEN_FILE):
        try:
            with open(TOKEN_FILE) as f:
                td = json.load(f)
            if not api_key: api_key = td.get("api_key", "")
            if not api_secret: api_secret = td.get("api_secret", "")
        except Exception:
            pass
    if not api_secret:
        opt_cfg = os.path.join(BASE_DIR, "input", "program_config.json")
        if os.path.exists(opt_cfg):
            try:
                with open(opt_cfg) as f:
                    cdata = json.load(f)
                if not api_key: api_key = cdata.get("api_key", "")
                if not api_secret: api_secret = cdata.get("api_secret", "")
            except Exception:
                pass
    if not api_key or not api_secret:
        logging.warning("api_key/api_secret missing in program_config.json")
    return api_key, api_secret


BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
TOKEN_FILE = paths.TOKEN_FILE
CONFIG_FILE = paths.PROGRAM_CONFIG_FILE
SCAN_DISPLAY_FILE = paths.SCAN_DISPLAY_STOCK_FILE
SCAN_DISPLAY_INDEX_FILE = paths.SCAN_DISPLAY_INDEX_FILE
SCAN_DISPLAY_BEAR_FILE = paths.SCAN_DISPLAY_BEAR_FILE
POSITIONS_FILE = os.path.join(BASE_DIR, "output", "monitor", "positions_stock.json")
JOURNAL_FILE = os.path.join(BASE_DIR, "output", "journal_stock.json")
DAILY_LOG_FILE = paths.BULL_DAILY_SCAN_LOG
BEAR_LOG_FILE = paths.BEAR_DAILY_SCAN_LOG
EMA_LOG_FILE = paths.EMA_LOG_FILE

LIVE_EXECUTION_FLAG = paths.NIFTY50_LIVE_FLAG
LIVE_EXECUTION_FLAG_INDEX = paths.INDEX_LIVE_FLAG
DASHBOARD_PORT = 5051

REFRESH_SECONDS = 1
ACTIVE_EDIT_LOCKS = set()

PROGRAMS = {
    "daily": {
        "name": "Stock_Bullish_Reversal_Scanner",
        "file": "Trade_Stock/stock_bullish_reversal_scanner.py",
        "desc": "Scans Nifty 50 on selected timeframe (Default: Day) for Bullish setups, exports to Excel",
        "color": "#d29922",
        "log_file": DAILY_LOG_FILE,
        "config_fields": {
            "product_type": {
                "label": "Execution Mode",
                "type": "select",
                "options": ["CNC", "MIS"],
                "default": "CNC"
            },
            "target_index": {
                "label": "Target Index Universe",
                "type": "select",
                "options": ["NIFTY50", "NIFTY_NEXT_100", "NIFTY_MIDCAP_100", "NIFTY_SMALLCAP_250"],
                "default": "NIFTY50"
            },
            "timeframe": {
                "label": "Timeframe",
                "type": "select",
                "options": ["day", "week", "4hr", "1hr", "75min", "30min", "15min"],
                "default": "day"
            },
            "max_daily_loss_pct": {"label": "Daily Loss Limit (%)", "type": "number", "default": 3.0},
            "enable_swing_filter": {"label": "Swing Filter", "type": "select", "options": ["true", "false"], "default": "true"},
            "swing_min_waves": {"label": "Min Swings", "type": "number", "default": 2},
            "swing_min_r2": {"label": "Swing R² (0.0-1.0)", "type": "number", "default": 0.55}
        }
    },
    "bear_trade": {
        "name": "Stock_Bearish_Reversal_Scanner",
        "file": "Trade_Stock/stock_bearish_reversal_scanner.py",
        "desc": "Scans predefined NSE indices on selected timeframe for Bearish setups & Negation targets",
        "color": "#f85149",
        "log_file": BEAR_LOG_FILE,
        "config_fields": {
            "product_type": {
                "label": "Execution Mode",
                "type": "select",
                "options": ["CNC", "MIS"],
                "default": "CNC"
            },
            "target_index": {
                "label": "Target Index Universe",
                "type": "select",
                "options": ["NIFTY50", "NIFTY_NEXT_100", "NIFTY_MIDCAP_100", "NIFTY_SMALLCAP_250"],
                "default": "NIFTY50"
            },
            "timeframe": {
                "label": "Timeframe",
                "type": "select",
                "options": ["day", "week", "4hr", "1hr", "75min", "30min", "15min"],
                "default": "day"
            },
            "max_daily_loss_pct": {"label": "Daily Loss Limit (%)", "type": "number", "default": 3.0},
            "enable_swing_filter": {"label": "Swing Filter", "type": "select", "options": ["true", "false"], "default": "true"},
            "swing_min_waves": {"label": "Min Swings", "type": "number", "default": 2},
            "swing_min_r2": {"label": "Swing R² (0.0-1.0)", "type": "number", "default": 0.55}
        }
    },
    "weekly": {
        "name": "Stock_Weekly_Bull_Scanner",
        "file": "Trade_Stock/stock_weekly_bull_scanner.py",
        "desc": "Scans NSE stocks on Weekly timeframe for Bullish A-B-C-D setups (Investment Grade)",
        "color": "#3fb950",
        "log_file": paths.WEEKLY_BULL_SCAN_LOG,
        "config_fields": {
            "target_index": {
                "label": "Target Index Universe",
                "type": "select",
                "options": ["NIFTY50", "NIFTY_NEXT_100", "NIFTY_MIDCAP_100", "NIFTY_SMALLCAP_250"],
                "default": "NIFTY_MIDCAP_100"
            },
            "enable_swing_filter": {"label": "Swing Filter", "type": "select", "options": ["true", "false"], "default": "true"},
            "swing_min_waves": {"label": "Min Swings", "type": "number", "default": 2},
            "swing_min_r2": {"label": "Swing R² (0.0-1.0)", "type": "number", "default": 0.45}
        }
    },
    "weekly_bear": {
        "name": "Stock_Weekly_Bear_Scanner",
        "file": "Trade_Stock/stock_weekly_bear_scanner.py",
        "desc": "Scans NSE stocks on Weekly timeframe for Bearish A-B-C-D setups (Investment Grade Short)",
        "color": "#ff7b72",
        "log_file": paths.WEEKLY_BEAR_SCAN_LOG,
        "config_fields": {
            "target_index": {
                "label": "Target Index Universe",
                "type": "select",
                "options": ["NIFTY50", "NIFTY_NEXT_100", "NIFTY_MIDCAP_100", "NIFTY_SMALLCAP_250"],
                "default": "NIFTY_MIDCAP_100"
            },
            "enable_swing_filter": {"label": "Swing Filter", "type": "select", "options": ["true", "false"], "default": "true"},
            "swing_min_waves": {"label": "Min Swings", "type": "number", "default": 2},
            "swing_min_r2": {"label": "Swing R² (0.0-1.0)", "type": "number", "default": 0.45}
        }
    },
    "ema_engine": {
        "name": "Stock EMA Engine",
        "file": "common/ema_engine.py",
        "desc": "Scans 13 EMA & 44 EMA crossovers on stock spot symbols",
        "color": "#a371f7",
        "log_file": EMA_LOG_FILE,
        "config_fields": {
            "timeframe": {"label": "Timeframe", "type": "select", "options": ["1d", "4hr", "60minute", "30minute", "15minute", "5minute"], "default": "1d"},
            "target_universe": {"label": "Target Universe", "type": "select", "options": ["ALL", "NIFTY50", "NIFTY_NEXT_100", "NIFTY_MIDCAP_100", "NIFTY_SMALLCAP_250", "INDEX_OPTIONS"], "default": "ALL"}
        }
    }
}

processes = {}
process_lock = threading.Lock()

data_lock = threading.Lock()
cached_data = {
    "positions": {},
    "journal": [],
    "log_tail": {pid: [] for pid in PROGRAMS},
    "stats": {"total_trades": 0, "win_rate": 0, "active_positions": 0, "pnl": 0},
    "scans": {"daily": [], "bear_trade": [], "weekly": [], "weekly_bear": []},
    "scan_summary": {"daily": {"anchors": {}, "abc_matches": {}}, "bear_trade": {"anchors": {}, "abc_matches": {}}, "weekly": {"anchors": {}, "abc_matches": {}}, "weekly_bear": {"anchors": {}, "abc_matches": {}}},
    "all_trades": [],
    "kite_positions": [],
    "ltp": {},
    "anchor_status": {"running": False, "engine": None, "requested_at": None, "completed_at": None},
    "scan_display": {"date": "", "timestamp": "", "staged_trades": [], "active_positions": []},
    "live_execution": False,
    "live_execution_index": False,
    "executed_exits": {},
    "expired_contracts": []
}
_expired_cache_day = ""
_expired_cache_set = set()
_ltp_last_fetch = 0
_kite_positions_last_fetch = 0
_kite_session = None
_last_scan_reset = ""
_ltp_memory = {}
_pnl_memory = {}

# ──────────────────────────────────────────────
#  PROCESS MANAGEMENT (Start/Stop Programs)
# ──────────────────────────────────────────────

def get_pid_for_program(prog_id):
    with process_lock:
        p = processes.get(prog_id)
        if p and p.poll() is None:
            return p.pid
    return None

def start_program(prog_id):
    token = check_token_valid()
    if not token["valid"]:
        print(f"Cannot start {prog_id}: {token['reason']}")
        return False
    with process_lock:
        if prog_id in processes and processes[prog_id].poll() is None:
            return False
        rel_file = PROGRAMS[prog_id]["file"]
        script_path = os.path.join(BASE_DIR, rel_file)
        if not os.path.exists(script_path):
            script_path = os.path.join(os.path.dirname(__file__), os.path.basename(rel_file))
        try:
            p = subprocess.Popen(
                [sys.executable, "-u", script_path],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                cwd=BASE_DIR, creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
            )
            processes[prog_id] = p
            return True
        except Exception as e:
            print(f"Failed to start {prog_id}: {e}")
            return False

def stop_program(prog_id):
    with process_lock:
        p = processes.get(prog_id)
        if p and p.poll() is None:
            if os.name == "nt":
                subprocess.run(["taskkill", "/F", "/T", "/PID", str(p.pid)], capture_output=True)
            else:
                os.kill(p.pid, signal.SIGTERM)
            processes.pop(prog_id, None)
            return True
    return False

# ──────────────────────────────────────────────
#  CONFIGURATION (program_config.json)
# ──────────────────────────────────────────────

def load_config():
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE) as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def save_config(prog_id, data):
    cfg = load_config()
    cleaned = {}
    for k, v in data.items():
        if isinstance(v, str):
            if v.lower() == "true":
                cleaned[k] = True
            elif v.lower() == "false":
                cleaned[k] = False
            else:
                try:
                    if "." in v:
                        cleaned[k] = float(v)
                    else:
                        cleaned[k] = int(v)
                except (ValueError, TypeError):
                    cleaned[k] = v
        else:
            cleaned[k] = v
    cfg[prog_id] = cleaned
    if "max_daily_loss_pct" in cleaned:
        if "portfolio_risk" not in cfg or not isinstance(cfg["portfolio_risk"], dict):
            cfg["portfolio_risk"] = {}
        cfg["portfolio_risk"]["max_daily_loss_pct"] = float(cleaned["max_daily_loss_pct"])
    os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)
    return True

def get_backtest_mode():
    return load_config().get("_backtest", False)

def set_backtest_mode(enabled):
    cfg = load_config()
    cfg["_backtest"] = enabled
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

# ──────────────────────────────────────────────
#  KITE TOKEN MANAGEMENT
# ──────────────────────────────────────────────

def check_token_valid():
    if not os.path.exists(TOKEN_FILE):
        return {"valid": False, "reason": "Token file not found"}
    try:
        with open(TOKEN_FILE) as f:
            data = json.load(f)
        if not data.get("api_key") or not data.get("access_token"):
            return {"valid": False, "reason": "Invalid token file"}
        date_str = data.get("generated_at", "")
        if date_str:
            try:
                from datetime import datetime as dt2
                gen_dt = dt2.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                now_dt = dt.now()
                gen_date = gen_dt.date()
                today = now_dt.date()
                if gen_date < today:
                    return {"valid": False, "reason": f"Token expired (generated {date_str})"}
                reset_cutoff = gen_dt.replace(hour=6, minute=0, second=0, microsecond=0)
                if gen_dt < reset_cutoff and now_dt >= reset_cutoff:
                    return {"valid": False, "reason": f"Token expired (generated {date_str} before 06:00 AM Zerodha reset)"}
            except Exception:
                try:
                    gen_date = dt.strptime(date_str.split()[0], "%Y-%m-%d").date()
                    if gen_date < dt.now().date():
                        return {"valid": False, "reason": f"Token expired (generated {date_str})"}
                except Exception:
                    pass
        return {"valid": True, "reason": "Token valid"}
    except Exception as e:
        return {"valid": False, "reason": f"Token read error: {e}"}

def get_login_url():
    api_key, _ = get_kite_credentials()
    if not api_key:
        return ""
    return f"https://kite.zerodha.com/connect/login?api_key={api_key}"

def exchange_request_token(request_token):
    """Exchange Kite request_token for access_token and save to file."""
    try:
        api_key, api_secret = get_kite_credentials()
        kite = KiteConnect(api_key=api_key)
        session = kite.generate_session(request_token, api_secret=api_secret)
        access_token = session["access_token"]
        token_data = {
            "api_key": api_key,
            "access_token": access_token,
            "generated_at": dt.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        os.makedirs(os.path.dirname(TOKEN_FILE), exist_ok=True)
        with open(TOKEN_FILE, "w") as f:
            json.dump(token_data, f, indent=4)
        return {"ok": True, "access_token": access_token[:8] + "..."}
    except Exception as e:
        return {"ok": False, "error": str(e)}

# ──────────────────────────────────────────────
#  DATA LOADING (trade_db, journal, logs)
# ──────────────────────────────────────────────

def load_positions():
    try:
        active = trade_db.get_active_trades()
        return {t["symbol"]: t for t in active}
    except Exception:
        return {}

def load_journal():
    try:
        from daily_trade_journal import load_journal_entries
        entries = load_journal_entries()
        if entries:
            return entries
    except Exception:
        pass
    rows = []
    if os.path.exists(JOURNAL_FILE):
        try:
            with open(JOURNAL_FILE, newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f, delimiter="\t")
                seen = set()
                for row in reader:
                    key = (row.get("Symbol", ""), row.get("Timestamp", ""))
                    if key not in seen:
                        seen.add(key)
                        rows.append(row)
        except Exception:
            pass
    return rows[-200:]


def tail_log(filepath, n=200):
    if not os.path.exists(filepath):
        return []
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            lines = [l.rstrip("\r\n") for l in f.readlines()]
        return lines[-n:]
    except Exception as e:
        logging.warning(f"tail_log error reading {filepath}: {e}")
        return []

def compute_stats(positions, journal):
    active = len(positions)
    total = len(journal)
    wins = sum(1 for j in (journal or []) if str(j.get("P&L %") or "").replace("%", "").replace("-", "").strip()
               and str(j.get("Action") or "").startswith("EXIT_"))
    win_rate = round((wins / total) * 100, 1) if total > 0 else 0
    pnl = 0.0
    for j in (journal or []):
        try:
            pnl_str = str(j.get("P&L %") or "").replace("%", "")
            if pnl_str and pnl_str != "-":
                pnl += float(pnl_str)
        except Exception:
            pass
    return {"total_trades": total, "win_rate": win_rate, "active_positions": active, "pnl": round(pnl, 2)}

SCAN_SYMBOLS = [
    "RELIANCE","TCS","HDFCBANK","ICICIBANK","INFY","ITC","SBIN","BHARTIARTL","LT","WIPRO",
    "ADANIENT","ADANIPORTS","APOLLOHOSP","ASIANPAINT","AXISBANK","BAJAJ-AUTO","BAJAJFINSV",
    "BAJFINANCE","BEL","CIPLA","COALINDIA","DRREDDY","EICHERMOT","ETERNAL","GRASIM","HCLTECH",
    "HDFCLIFE","HINDALCO","HINDUNILVR","INDIGO","JIOFIN","JSWSTEEL",
    "KOTAKBANK","M&M","MARUTI","MAXHEALTH","NESTLEIND","NTPC","ONGC","POWERGRID","SBILIFE",
    "SHRIRAMFIN","SUNPHARMA","TATACONSUM","TMPV","TATASTEEL","TECHM","TITAN","TRENT","ULTRACEMCO",
    "NIFTY","BANKNIFTY"
]

# ──────────────────────────────────────────────
#  SCAN PARSING — Split anchor vs ABC per symbol
# ──────────────────────────────────────────────

def _extract_symbol(line):
    for sym in SCAN_SYMBOLS:
        if sym in line:
            return sym
    return None

def parse_scans_for_program(log_lines, prog_id):
    matches = []
    anchors = {}
    abc_matches = {}
    for line in log_lines:
        clean = line.strip()
        if "ANCHOR" in clean:
            matches.append(clean)
            sym = _extract_symbol(clean)
            if sym:
                anchors[sym] = clean
        elif "MATCH" in clean or "BEST TRADE" in clean or "Match" in clean:
            matches.append(clean)
            sym = _extract_symbol(clean)
            if sym:
                abc_matches[sym] = clean
    return matches, anchors, abc_matches

_file_mtime_cache = {}
_parsed_json_cache = {}

def refresh_data(single_run=False):
    global cached_data, _ltp_last_fetch, _kite_positions_last_fetch, _kite_session, _last_scan_reset
    try:
        trade_db.run_db_housekeeping()
    except Exception:
        pass
    while True:
        with data_lock:
            pos = load_positions()
            journal = load_journal()
            cached_data["positions"] = pos
            cached_data["journal"] = journal
            cached_data["stats"] = compute_stats(pos, journal)
            try:
                all_t = trade_db.get_all_trades()
                cached_data["all_trades"] = all_t
            except Exception:
                cached_data["all_trades"] = []
            for pid in PROGRAMS:
                log_file = PROGRAMS[pid].get("log_file")
                log_lines = tail_log(log_file) if log_file else []
                cached_data["log_tail"][pid] = log_lines
                scan_lines, anchors, abc_matches = parse_scans_for_program(log_lines, pid)
                cached_data["scans"][pid] = scan_lines
                cached_data["scan_summary"][pid] = {"anchors": anchors, "abc_matches": abc_matches}
            now_ist = get_ist_now(naive=True)
            today_str = now_ist.strftime("%Y-%m-%d")
            market_open = now_ist.replace(hour=9, minute=0, second=0, microsecond=0)
            if _last_scan_reset != today_str and now_ist >= market_open:
                _last_scan_reset = today_str
                for f in [SCAN_DISPLAY_FILE, SCAN_DISPLAY_INDEX_FILE]:
                    try:
                        if os.path.exists(f):
                            with open(f, "r") as fh:
                                existing = json.load(fh)
                            if existing.get("date") == today_str:
                                continue  # Keep today's scan display data intact
                        empty_scan = {"date": today_str, "timestamp": now_ist.strftime("%Y-%m-%d %H:%M:%S"), "staged_trades": [], "carry_forward": [], "active_live": []}
                        with open(f, "w") as fh:
                            json.dump(empty_scan, fh)
                    except Exception:
                        pass
            scan_display = {}
            try:
                if os.path.exists(SCAN_DISPLAY_FILE):
                    mtime = os.path.getmtime(SCAN_DISPLAY_FILE)
                    if _file_mtime_cache.get(SCAN_DISPLAY_FILE) == mtime and SCAN_DISPLAY_FILE in _parsed_json_cache:
                        scan_display["nifty50"] = _parsed_json_cache[SCAN_DISPLAY_FILE]
                    else:
                        with open(SCAN_DISPLAY_FILE, "r") as f:
                            data_obj = json.load(f)
                        _file_mtime_cache[SCAN_DISPLAY_FILE] = mtime
                        _parsed_json_cache[SCAN_DISPLAY_FILE] = data_obj
                        scan_display["nifty50"] = data_obj
            except Exception:
                pass
            try:
                if os.path.exists(SCAN_DISPLAY_INDEX_FILE):
                    with open(SCAN_DISPLAY_INDEX_FILE, "r") as f:
                        scan_display["index"] = json.load(f)
            except Exception:
                pass
            cached_data["scan_display"] = scan_display
            cached_data["live_execution"] = os.path.exists(LIVE_EXECUTION_FLAG)
            cached_data["live_execution_index"] = os.path.exists(LIVE_EXECUTION_FLAG_INDEX)
            try:
                from trading_core import load_executed_exits, EXECUTED_EXITS
                load_executed_exits()
                cached_data["executed_exits"] = dict(EXECUTED_EXITS)
            except Exception:
                pass
            day_key = get_ist_now().strftime("%Y-%m-%d")
            global _expired_cache_day, _expired_cache_set
            if _expired_cache_day != day_key:
                try:
                    from trading_core import contract_is_expired
                    sym_set = set()
                    for _cat in ("staged_trades", "all_staged_today", "carry_forward", "active_live"):
                        for item in (cached_data["scan_display"].get("nifty50") or {}).get(_cat) or []:
                            c = str(item.get("contract") or item.get("symbol") or "").replace(" ", "").upper()
                            if c:
                                sym_set.add(c)
                    for _cat in ("staged_trades", "all_staged_today", "carry_forward", "active_live"):
                        for item in (cached_data["scan_display"].get("index") or {}).get(_cat) or []:
                            c = str(item.get("contract") or item.get("symbol") or "").replace(" ", "").upper()
                            if c:
                                sym_set.add(c)
                    for t in cached_data.get("all_trades", []):
                        c = str(t.get("contract") or t.get("symbol") or "").replace(" ", "").upper()
                        if c:
                            sym_set.add(c)
                    _expired_cache_set = {c for c in sym_set if contract_is_expired(c)}
                    _expired_cache_day = day_key
                except Exception:
                    pass
            cached_data["expired_contracts"] = sorted(_expired_cache_set)
            now = time.time()
            if now - _ltp_last_fetch > 3 and cached_data["all_trades"]:
                _ltp_last_fetch = now
                try:
                    active = trade_db.get_active_trades()
                    if active:
                        if not _kite_session:
                            tk = check_token_valid()
                            if tk["valid"]:
                                td = json.load(open(TOKEN_FILE))
                                api_key, _ = get_kite_credentials()
                                ks = KiteConnect(api_key=api_key)
                                ks.set_access_token(td["access_token"])
                                _kite_session = ks
                        if _kite_session:
                            try:
                                trade_db.reconcile_broker_live_positions(_kite_session)
                            except Exception:
                                pass
                            syms = []
                            for t in active:
                                tok = t.get("option_token") or t.get("index_token")
                                if tok: syms.append(int(tok))
                            if syms:
                                quotes = _kite_session.quote(syms)
                                ltp = {}
                                for key, q in quotes.items():
                                    ltp[key.split(":")[-1]] = q.get("last_price", 0)
                                cached_data["ltp"] = ltp
                except Exception:
                    _kite_session = None
        if now - _kite_positions_last_fetch > 3:
            _kite_positions_last_fetch = now
            try:
                if not _kite_session:
                    tk = check_token_valid()
                    if tk["valid"]:
                        td = json.load(open(TOKEN_FILE))
                        api_key, _ = get_kite_credentials()
                        ks = KiteConnect(api_key=api_key)
                        ks.set_access_token(td["access_token"])
                        _kite_session = ks
                if _kite_session:
                    kite_positions = _kite_session.positions()
                    merged = []
                    net_pos = [p for p in kite_positions.get("net", []) if p.get("tradingsymbol") and int(p.get("quantity", 0)) > 0]
                    q_keys = [f"{p.get('exchange', 'NSE')}:{p.get('tradingsymbol')}" for p in net_pos]
                    quotes_bulk = {}
                    if q_keys:
                        try:
                            quotes_bulk = _kite_session.quote(q_keys)
                        except Exception:
                            pass
                    for p in net_pos:
                        sym = p.get("tradingsymbol", "")
                        qty = int(p.get("quantity", 0))
                        entry_pr = float(p.get("average_price", 0))
                        exch = p.get("exchange", "NSE")
                        q_key = f"{exch}:{sym}"
                        live_ltp = 0
                        if q_key in quotes_bulk:
                            live_ltp = float(quotes_bulk[q_key].get("last_price", 0))

                        tok_id = str(p.get("instrument_token", ""))
                        sym_str = str(sym)

                        if live_ltp > 0:
                            _ltp_memory[sym_str] = live_ltp
                            if tok_id:
                                _ltp_memory[tok_id] = live_ltp
                            cached_data["ltp"][sym_str] = live_ltp
                            if tok_id:
                                cached_data["ltp"][tok_id] = live_ltp
                        else:
                            live_ltp = _ltp_memory.get(sym_str) or _ltp_memory.get(tok_id) or 0
                            if live_ltp > 0:
                                cached_data["ltp"][sym_str] = live_ltp
                                if tok_id:
                                    cached_data["ltp"][tok_id] = live_ltp

                        if live_ltp > 0 and entry_pr > 0:
                            live_pnl = round((live_ltp - entry_pr) * qty, 2)
                            _pnl_memory[sym_str] = live_pnl
                        else:
                            live_pnl = _pnl_memory.get(sym_str, float(p.get("pnl", 0)))

                        if exch not in ("NSE", "BSE") or sym.endswith("CE") or sym.endswith("PE"):
                            continue

                        # Fail-Safe Active Position Risk Monitor
                        try:
                            scan_sl = lookup_scan_sl_target(sym, sym, "daily", _kite_session, entry_pr, is_stock=True)
                            
                            pos_item = {
                                "contract": sym,
                                "symbol": sym,
                                "quantity": qty,
                                "entry_price": entry_pr,
                                "entry_spot": entry_pr,
                                "ltp": live_ltp,
                                "pnl": live_pnl,
                                "exchange": exch,
                                "source": "kite"
                            }
                            if scan_sl:
                                pos_item["current_sl"] = scan_sl.get("current_sl", 0)
                                pos_item["t1"] = scan_sl.get("t1", 0)
                                pos_item["t2"] = scan_sl.get("t2", 0)
                                pos_item["t3"] = scan_sl.get("t3", 0)
                                pos_item["pattern"] = scan_sl.get("pattern", "SCAN_LINKED")
                            merged.append(pos_item)

                            if scan_sl:
                                ltp_val = live_ltp
                                sl_val = float(scan_sl.get("current_sl", 0))
                                t1_val = float(scan_sl.get("t1", 0) or 0)
                                t2_val = float(scan_sl.get("t2", 0) or 0)
                                t3_val = float(scan_sl.get("t3", 0) or 0)

                                clean_sym = str(sym).replace(" ", "").upper()
                                now_t = get_ist_now().time()
                                cfg_f = load_config()
                                fs_start_str = cfg_f.get("failsafe_start_time", "09:45")
                                try:
                                    f_h, f_m = map(int, fs_start_str.split(":"))
                                    fs_start_t = datetime_time(f_h, f_m)
                                except Exception:
                                    fs_start_t = datetime_time(9, 45)

                                sl_buffered = round(sl_val * 0.995, 2)
                                is_below_buffer = ltp_val <= sl_buffered
                                is_deep_break = ltp_val <= round(sl_val * 0.985, 2)

                                prev_closed_below = False
                                token_id = scan_sl.get("option_token") or scan_sl.get("index_token") or scan_sl.get("token")
                                if token_id and _kite_session:
                                    try:
                                        df_hist = fetch_and_resample_candles(_kite_session, token_id, (dt.now() - timedelta(days=2)).strftime("%Y-%m-%d"), dt.now().strftime("%Y-%m-%d"), "15minute")
                                        if len(df_hist) >= 2:
                                            prev_close_val = float(df_hist.iloc[-2]["close"])
                                            if prev_close_val > 0 and prev_close_val <= sl_val:
                                                prev_closed_below = True
                                    except Exception:
                                        pass

                                def _t_early_buf(v):
                                    if not v or v <= 0: return 0.0
                                    if v <= 50: return max(0.50, round(v * 0.015, 2))
                                    elif v <= 200: return max(1.00, round(v * 0.015, 2))
                                    else: return max(2.00, round(v * 0.010, 2))

                                # TASK 1: Pause automated exit execution if user is actively editing this symbol on the UI
                                if clean_sym in ACTIVE_EDIT_LOCKS:
                                    logging.info(f"[FAILSAFE PAUSED] {sym} is currently being edited on UI. Automated exit execution paused.")
                                # TASK 1b: Skip if exit order has already been executed/submitted
                                elif is_contract_exit_executed(sym):
                                    pass
                                # 2. Check T3 Target Hit Exit (Active from 09:15 AM)
                                elif ltp_val > 0 and t3_val > 0 and ltp_val >= t3_val:
                                    logging.info(f"[FAILSAFE MONITOR EXIT T3] {sym} LTP={ltp_val} >= T3={t3_val}")
                                    pos_obj = {"contract": sym, "position_size": qty, "quantity": qty, "symbol": sym}
                                    shared_close_stock_position(_kite_session, pos_obj, True, p.get("product"))
                                # 2b. Check T2 Target Exit (No T3 -> Full exit on T2 touch, Active from 09:15 AM)
                                elif ltp_val > 0 and t2_val > 0 and (t3_val <= 0 or t3_val is None) and ltp_val >= (t2_val - _t_early_buf(t2_val)):
                                    logging.warning(f"[FAILSAFE MONITOR EXIT T2 (no T3)] {sym} LTP={ltp_val} >= T2-buffer={t2_val - _t_early_buf(t2_val):.2f} (Target: {t2_val:.2f})")
                                    pos_obj = {"contract": sym, "position_size": qty, "quantity": qty, "symbol": sym}
                                    shared_close_stock_position(_kite_session, pos_obj, True, p.get("product"))
                                # 2c. Check T1 Target Exit (No T2/T3 -> Full exit on T1 touch, Active from 09:15 AM)
                                elif ltp_val > 0 and t1_val > 0 and t2_val <= 0 and (t3_val <= 0 or t3_val is None) and ltp_val >= (t1_val - _t_early_buf(t1_val)):
                                    logging.warning(f"[FAILSAFE MONITOR EXIT T1 (no T2/T3)] {sym} LTP={ltp_val} >= T1-buffer={t1_val - _t_early_buf(t1_val):.2f} (Target: {t1_val:.2f})")
                                    pos_obj = {"contract": sym, "position_size": qty, "quantity": qty, "symbol": sym}
                                    shared_close_stock_position(_kite_session, pos_obj, True, p.get("product"))
                                effective_entry = entry_pr if entry_pr > 0 else float(scan_sl.get("entry_spot") or 0)
                                hard_max_8pct_break = (ltp_val <= round(effective_entry * 0.92, 2)) if effective_entry > 0 else False

                                # TASK 2: Execute SL exit ONLY IF after 09:45 AM AND (candle closed below SL OR emergency deep break OR hard max 8% loss cap hit)
                                if now_t >= fs_start_t and ltp_val > 0 and (sl_val > 0 or hard_max_8pct_break) and (is_below_buffer or hard_max_8pct_break) and (prev_closed_below or is_deep_break or hard_max_8pct_break):
                                    logging.warning(f"[FAILSAFE MONITOR EXIT SL CONFIRMED] {sym} LTP={ltp_val} (Reason: {'HARD_MAX_8PCT_SL' if hard_max_8pct_break else ('CANDLE_CLOSE_SL' if prev_closed_below else 'EMERGENCY_HARD_SL')}, Entry={effective_entry}, SL={sl_val})")
                                    pos_obj = {"contract": sym, "position_size": qty, "quantity": qty, "symbol": sym}
                                    shared_close_stock_position(_kite_session, pos_obj, True, p.get("product"))
                                elif now_t < fs_start_t and ltp_val > 0 and sl_val > 0 and is_below_buffer:
                                    logging.info(f"[FAILSAFE SL PAUSED BEFORE {fs_start_str} AM] {sym} SL check paused until {fs_start_str} AM (Current time: {now_t.strftime('%H:%M:%S')}).")
                        except Exception as fs_err:
                            logging.debug(f"Failsafe monitor error for {sym}: {fs_err}")

                    cached_data["kite_positions"] = merged
            except Exception:
                pass
        if single_run:
            break
        if int(time.time()) % 3600 < REFRESH_SECONDS:
            auto_export_if_new_month()
        time.sleep(REFRESH_SECONDS)

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates/index.html')

AUTH_PAGE_STYLE = """
    body { background:#0d1117; color:#c9d1d9; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Helvetica, Arial, sans-serif; margin:0; }
    .auth-wrap { max-width:420px; margin:80px auto; padding:28px 24px; background:#161b22; border:1px solid #30363d; border-radius:10px; box-shadow:0 8px 24px rgba(0,0,0,0.5); }
    .auth-wrap h2 { margin:0 0 16px; color:#f0f6fc; font-size:20px; font-weight:600; text-align:center; }
    .auth-wrap label { display:block; font-size:12px; color:#8b949e; margin:14px 0 6px; font-weight:500; }
    .auth-wrap input { width:100%; box-sizing:border-box; padding:10px 12px; background:#0d1117; color:#c9d1d9; border:1px solid #30363d; border-radius:6px; font-size:14px; outline:none; transition:border-color 0.2s; }
    .auth-wrap input:focus { border-color:#58a6ff; }
    .auth-wrap button { width:100%; margin-top:20px; padding:10px; background:#238636; color:#fff; border:none; border-radius:6px; font-size:14px; font-weight:600; cursor:pointer; transition:background 0.2s; }
    .auth-wrap button:hover { background:#2ea043; }
    .auth-error { color:#f85149; font-size:13px; margin-top:12px; background:#f8514915; padding:8px 12px; border-radius:6px; border:1px solid #f8514933; text-align:center; }
    .auth-notice { color:#3fb950; font-size:13px; margin-top:12px; background:#3fb95015; padding:8px 12px; border-radius:6px; border:1px solid #3fb95033; text-align:center; }
    .auth-link { display:block; text-align:center; margin-top:16px; font-size:13px; color:#58a6ff; text-decoration:none; }
    .auth-link:hover { text-decoration:underline; }
    table.admin-table { width:100%; border-collapse:collapse; margin-top:16px; }
    table.admin-table th, table.admin-table td { text-align:left; padding:10px; border-bottom:1px solid #30363d; font-size:13px; }
    table.admin-table th { color:#8b949e; font-weight:600; background:#21262d; }
    .badge-approved { color:#3fb950; font-weight:600; }
    .badge-pending { color:#d29922; font-weight:600; }
    .btn-sm { padding:4px 10px; border:none; border-radius:4px; font-size:12px; font-weight:600; cursor:pointer; margin-right:4px; }
    .btn-approve { background:#238636; color:#fff; }
    .btn-approve:hover { background:#2ea043; }
    .btn-reject { background:#da3633; color:#fff; }
    .btn-reject:hover { background:#f85149; }
    .btn-del { background:#6e7681; color:#fff; }
    .btn-del:hover { background:#8b949e; }
"""

LOGIN_TEMPLATE = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Login - Stock Trading Control Center</title><style>""" + AUTH_PAGE_STYLE + """</style></head><body>
<div class="auth-wrap">
    <h2>🎯 Stock Control Center</h2>
    <form method="POST" action="/login">
        <label>Username</label>
        <input type="text" name="username" autocomplete="username" placeholder="Enter username" required autofocus>
        <label>Password</label>
        <input type="password" name="password" autocomplete="current-password" placeholder="Enter password" required>
        <button type="submit">Sign In</button>
    </form>
    {% if error %}<div class="auth-error">{{ error }}</div>{% endif %}
    {% if notice %}<div class="auth-notice">{{ notice }}</div>{% endif %}
    <a class="auth-link" href="/register">Request access / Register</a>
</div>
</body></html>"""

REGISTER_TEMPLATE = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Register - Stock Trading Control Center</title><style>""" + AUTH_PAGE_STYLE + """</style></head><body>
<div class="auth-wrap">
    <h2>🎯 Request Access</h2>
    <form method="POST" action="/register">
        <label>Username (min 3 chars)</label>
        <input type="text" name="username" autocomplete="username" placeholder="Choose a username" required autofocus>
        <label>Password (min 4 chars)</label>
        <input type="password" name="password" autocomplete="new-password" placeholder="Create a password" required>
        <button type="submit">Create Account</button>
    </form>
    {% if error %}<div class="auth-error">{{ error }}</div>{% endif %}
    {% if notice %}<div class="auth-notice">{{ notice }}</div>{% endif %}
    <a class="auth-link" href="/login">&larr; Back to login</a>
</div>
</body></html>"""

ADMIN_TEMPLATE = """<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Admin - User Management</title><style>""" + AUTH_PAGE_STYLE + """</style></head><body>
<div class="auth-wrap" style="max-width:680px;">
    <h2>👥 Registered Users (Admin: {{ user }})</h2>
    <table class="admin-table">
        <thead><tr><th>Username</th><th>Role</th><th>Status</th><th>Registered</th><th>Actions</th></tr></thead>
        <tbody>
        {% for u in users %}
        <tr>
            <td><strong>{{ u.username }}</strong></td>
            <td>{{ u.role }}</td>
            <td class="{% if u.approved %}badge-approved{% else %}badge-pending{% endif %}">{% if u.approved %}Approved{% else %}Pending Approval{% endif %}</td>
            <td style="font-size:11px;color:#8b949e;">{{ u.created_at[:16].replace('T', ' ') if u.created_at else '-' }}</td>
            <td>
                {% if not u.approved %}
                <form method="POST" action="/api/admin/approve" style="display:inline;">
                    <input type="hidden" name="username" value="{{ u.username }}">
                    <button class="btn-sm btn-approve" type="submit">Approve</button>
                </form>
                {% else %}
                <form method="POST" action="/api/admin/reject" style="display:inline;">
                    <input type="hidden" name="username" value="{{ u.username }}">
                    <button class="btn-sm btn-reject" type="submit">Revoke</button>
                </form>
                {% endif %}
                <form method="POST" action="/api/admin/delete" style="display:inline;">
                    <input type="hidden" name="username" value="{{ u.username }}">
                    <button class="btn-sm btn-del" type="submit" onclick="return confirm('Delete user {{ u.username }}?')">Delete</button>
                </form>
            </td>
        </tr>
        {% endfor %}
        </tbody>
    </table>
    <div style="display:flex;justify-content:space-between;margin-top:20px;">
        <a class="auth-link" href="/">&larr; Back to Dashboard</a>
        <a class="auth-link" href="/logout" style="color:#f85149;">Logout</a>
    </div>
</div>
</body></html>"""

# ──────────────────────────────────────────────
#  FLASK ROUTES — API Endpoints & Auth Gate
# ──────────────────────────────────────────────

@app.before_request
def auth_gate():
    if request.path in ("/login", "/register", "/logout", "/favicon.ico") or request.path.startswith("/api/token/callback") or request.path.startswith("/api/postback"):
        return None
    if request.path == "/admin":
        if not session.get("user"):
            return redirect("/login")
        if session.get("role") != "admin":
            return jsonify({"ok": False, "error": "Admin only"}), 403
        return None
    if request.path.startswith("/api/admin/"):
        if not session.get("user"):
            return jsonify({"ok": False, "error": "Login required"}), 401
        if session.get("role") != "admin":
            return jsonify({"ok": False, "error": "Admin only"}), 403
        return None
    if not session.get("user"):
        if request.path.startswith("/api/"):
            return jsonify({"ok": False, "error": "Login required"}), 401
        return redirect("/login")
    return None

@app.route("/favicon.ico")
def favicon():
    return "", 204

@app.route("/")
def dashboard():
    with open(TEMPLATE_PATH, encoding="utf-8") as _template_f:
        tpl = _template_f.read()
    return render_template_string(tpl, refresh=REFRESH_SECONDS, programs=PROGRAMS, user=session.get("user", ""), role=session.get("role", ""))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = dashboard_auth.verify_user(username, password)
        if not user:
            return render_template_string(LOGIN_TEMPLATE, error="Invalid username or password", notice="")
        if not user["approved"]:
            return render_template_string(LOGIN_TEMPLATE, error="Your account is awaiting administrator approval", notice="")
        session["user"] = user["username"]
        session["role"] = user["role"]
        return redirect("/")
    return render_template_string(LOGIN_TEMPLATE, error="", notice="")

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        created, err = dashboard_auth.register_user(username, password)
        if err:
            return render_template_string(REGISTER_TEMPLATE, error=err, notice="")
        notice = "First account created as Administrator." if len(dashboard_auth.list_users()) == 1 else "Registration submitted! Awaiting admin approval."
        return render_template_string(LOGIN_TEMPLATE, error="", notice=notice + " You can now sign in.")
    return render_template_string(REGISTER_TEMPLATE, error="", notice="")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/admin")
def admin_page():
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "Admin only"}), 403
    return render_template_string(ADMIN_TEMPLATE, user=session.get("user", ""), users=dashboard_auth.list_users())

@app.route("/api/admin/approve", methods=["POST"])
def admin_approve():
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "Admin only"}), 403
    dashboard_auth.approve_user(request.form.get("username", ""))
    return redirect("/admin")

@app.route("/api/admin/reject", methods=["POST"])
def admin_reject():
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "Admin only"}), 403
    dashboard_auth.reject_user(request.form.get("username", ""))
    return redirect("/admin")

@app.route("/api/admin/delete", methods=["POST"])
def admin_delete():
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "Admin only"}), 403
    dashboard_auth.delete_user(request.form.get("username", ""))
    return redirect("/admin")

@app.route("/api/admin/users", methods=["GET"])
def admin_users():
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "Admin only"}), 403
    return jsonify({"ok": True, "users": dashboard_auth.list_users()})

@app.route("/api/status")
def api_status():
    with data_lock:
        prog_status = {}
        for pid in PROGRAMS:
            if pid == "ema_engine":
                pid_running = get_ema_engine_status(is_options_mode=False)
            else:
                pid_running = get_pid_for_program(pid) is not None
            log_lines = cached_data["log_tail"].get(pid, [])
            if not log_lines and PROGRAMS[pid].get("log_file"):
                log_lines = tail_log(PROGRAMS[pid].get("log_file"))
            prog_status[pid] = {
                "running": pid_running,
                "scans": cached_data["scans"].get(pid, []),
                "log_tail": log_lines,
                "scan_summary": cached_data["scan_summary"].get(pid, {"anchors": {}, "abc_matches": {}})
            }
        cfg = load_config()
        p_risk_dl = float(cfg.get("portfolio_risk", {}).get("max_daily_loss_pct", 3.0))
        for p_id in ["daily", "bear_trade"]:
            if p_id in cfg and isinstance(cfg[p_id], dict):
                if "max_daily_loss_pct" not in cfg[p_id]:
                    cfg[p_id]["max_daily_loss_pct"] = p_risk_dl
        return jsonify({
            "programs": prog_status,
            "positions": cached_data["positions"],
            "all_trades": cached_data["all_trades"],
            "kite_positions": cached_data["kite_positions"],
            "ltp": {str(k): v for k, v in cached_data["ltp"].items()},
            "journal": cached_data["journal"],
            "stats": cached_data["stats"],
            "config": cfg,
            "scan_display": cached_data["scan_display"],
            "ema_scan": get_ema_scan_data(is_options_mode=False),
            "live_execution": cached_data["live_execution"],
            "live_execution_index": cached_data["live_execution_index"],
            "executed_exits": cached_data.get("executed_exits", {}),
            "expired_contracts": cached_data.get("expired_contracts", [])
        })

@app.route("/api/token/check")
def api_token_check():
    return jsonify(check_token_valid())

@app.route("/api/token/url")
def api_token_url():
    url = get_login_url()
    if not url:
        return jsonify({"url": "", "error": "Kite API Key is missing. Please save your API Key & Secret below or configure input/program_config.json."})
    return jsonify({"url": url})

@app.route("/api/token/save-credentials", methods=["POST"])
def api_token_save_credentials():
    data = request.get_json(force=True, silent=True) or {}
    api_key = str(data.get("api_key", "")).strip()
    api_secret = str(data.get("api_secret", "")).strip()
    if not api_key or not api_secret:
        return jsonify({"ok": False, "error": "Both API Key and API Secret are required."})
    cfg = load_config()
    cfg["api_key"] = api_key
    cfg["api_secret"] = api_secret
    os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)
    return jsonify({"ok": True, "url": get_login_url()})

@app.route("/api/token/callback", methods=["GET", "POST"])
def api_token_callback():
    req_token = request.args.get("request_token") or request.form.get("request_token")
    if not req_token:
        data = request.get_json(force=True, silent=True) or {}
        req_token = data.get("request_token")
    if not req_token:
        return "<h3>Error: No request_token received from Zerodha</h3><p><a href='/'>Return to Dashboard</a></p>", 400
    res = exchange_request_token(req_token.strip())
    if res.get("ok"):
        return """
        <!DOCTYPE html>
        <html>
        <head><title>Kite Token Generated</title></head>
        <body style="font-family:sans-serif;background:#0d1117;color:#c9d1d9;text-align:center;padding:50px;">
            <div style="background:#161b22;border:1px solid #238636;border-radius:8px;padding:30px;max-width:500px;margin:auto;box-shadow:0 4px 12px rgba(0,0,0,0.5);">
                <h2 style="color:#3fb950;margin-top:0;">&#x2705; Token Generated Successfully!</h2>
                <p style="color:#8b949e;">Zerodha access token has been generated and saved for today.</p>
                <a href="/" style="display:inline-block;background:#238636;color:#ffffff;padding:10px 20px;border-radius:6px;text-decoration:none;font-weight:bold;margin-top:15px;">Return to Dashboard</a>
            </div>
            <script>setTimeout(function(){ window.location.href = '/'; }, 1500);</script>
        </body>
        </html>
        """
    else:
        err = res.get("error", "Unknown error")
        return f"<h3>Token Exchange Failed</h3><p>Error: {err}</p><p><a href='/'>Return to Dashboard</a></p>", 500

@app.route("/api/postback", methods=["POST"])
def api_postback():
    try:
        data = request.get_json(force=True, silent=True) or request.form.to_dict()
        logging.info(f"[POSTBACK] Received Zerodha order update: {data}")
    except Exception as e:
        logging.warning(f"[POSTBACK] Error parsing postback: {e}")
    return jsonify({"status": "success"}), 200

@app.route("/api/token/exchange", methods=["POST"])
def api_token_exchange():
    data = request.get_json(force=True, silent=True)
    if not data or not data.get("request_token"):
        return jsonify({"ok": False, "error": "No request_token provided"})
    result = exchange_request_token(data["request_token"].strip())
    return jsonify(result)

@app.route("/api/backtest/mode", methods=["GET", "POST"])
def api_backtest_mode():
    if request.method == "POST":
        data = request.get_json(force=True, silent=True)
        enabled = bool(data.get("enabled", False))
        set_backtest_mode(enabled)
    return jsonify({"enabled": get_backtest_mode()})

@app.route("/api/config/<prog_id>", methods=["POST"])
def api_save_config(prog_id):
    if prog_id not in PROGRAMS:
        return jsonify({"ok": False, "error": "Unknown program"})
    data = request.get_json(force=True, silent=True)
    if not data:
        return jsonify({"ok": False, "error": "Invalid JSON"})
    save_config(prog_id, data)
    return jsonify({"ok": True})

@app.route("/api/scan/clear", methods=["POST"])
def api_scan_clear():
    now_str = dt.now().strftime("%Y-%m-%d %H:%M:%S")
    today_str = dt.now().strftime("%Y-%m-%d")
    for f in [paths.SCAN_DISPLAY_FILE, paths.SCAN_DISPLAY_INDEX_FILE, paths.SCAN_DISPLAY_STOCK_FILE, paths.SCAN_DISPLAY_BEAR_FILE, paths.SCAN_DISPLAY_WEEKLY_FILE, paths.SCAN_DISPLAY_WEEKLY_BEAR_FILE]:
        try:
            empty_scan = {
                "date": today_str,
                "timestamp": now_str,
                "cleared_at": now_str,
                "staged_trades": [],
                "all_staged_today": [],
                "carry_forward": [],
                "active_live": []
            }
            os.makedirs(os.path.dirname(f), exist_ok=True)
            with open(f, "w", encoding="utf-8") as fh:
                json.dump(empty_scan, fh, indent=2)
        except Exception:
            pass
    try:
        if os.path.exists(paths.CYCLE_STORE_FILE):
            with open(paths.CYCLE_STORE_FILE, "w", encoding="utf-8") as fh:
                json.dump({}, fh)
    except Exception:
        pass
    try:
        import trade_db
        for eng in ["daily", "bear_trade", "weekly", "weekly_bear", "nifty50", "index"]:
            trade_db.clear_cycle_trades(eng)
    except Exception:
        pass
    _file_mtime_cache.clear()
    _parsed_json_cache.clear()
    with data_lock:
        for k in ["daily", "bear_trade", "weekly", "weekly_bear", "nifty50", "index"]:
            cached_data["scan_display"][k] = {"staged_trades": [], "all_staged_today": [], "carry_forward": [], "active_live": [], "cleared_at": now_str}
    return jsonify({"ok": True})

@app.route("/api/scan/ema/clear", methods=["POST"])
def api_scan_ema_clear():
    now_str = dt.now().strftime("%Y-%m-%d %H:%M:%S")
    today_str = dt.now().strftime("%Y-%m-%d")
    for ema_file in [paths.SCAN_DISPLAY_EMA_FILE, paths.SCAN_DISPLAY_EMA_STOCK_FILE]:
        try:
            empty_scan = {
                "ema_engine": {"staged_trades": [], "all_staged_today": [], "carry_forward": [], "active_live": []},
                "last_updated": now_str,
                "cleared_at": now_str
            }
            os.makedirs(os.path.dirname(ema_file), exist_ok=True)
            with open(ema_file, "w", encoding="utf-8") as fh:
                json.dump(empty_scan, fh, indent=2)
        except Exception:
            pass
    with data_lock:
        cached_data["ema_scan"] = {
            "ema_engine": {"staged_trades": [], "all_staged_today": [], "carry_forward": [], "active_live": []},
            "last_updated": now_str,
            "cleared_at": now_str
        }
    return jsonify({"ok": True})

def _format_pattern_result(p):
    if not p: return '-'
    p_str = str(p)
    if 'Engulf' in p_str:
        return 'BEAR_ENG' if 'Bear' in p_str or 'BEAR' in p_str else 'BULL_ENG'
    elif 'Two_Higher' in p_str or 'Higher_Highs' in p_str:
        return 'BULL_2HH'
    elif 'Two_Lower' in p_str or 'Lower_Lows' in p_str:
        return 'BEAR_2LL'
    elif 'HH_Sweep' in p_str or 'HH_sweep' in p_str:
        return 'BEAR_HH'
    elif 'Sweep' in p_str or 'LL' in p_str:
        return 'BULL_LL'
    elif 'Star' in p_str or 'Shooting' in p_str:
        return 'BEAR_STAR'
    elif 'Baby' in p_str or 'Hammer' in p_str:
        return 'BULL_HAM'
    elif 'Harami' in p_str:
        return 'BEAR_HAR' if 'Bear' in p_str or 'BEAR' in p_str else 'BULL_HAR'
    elif 'Base' in p_str:
        return 'BULL_BASE'
    elif p_str == 'SCAN_READY':
        return 'BULL_ENG'
    return p_str

def _format_timestamp(ts):
    if not ts: return '-'
    try:
        s = str(ts).split('+')[0].replace('T', ' ')
        p = s.split(' ')
        dp = p[0].split('-') if p[0] else []
        tp = p[1].split(':') if len(p) > 1 and p[1] else []
        if len(dp) == 3 and len(tp) >= 2:
            return f"{dp[2]}-{dp[1]}-{dp[0][-2:]} {tp[0]}:{tp[1]}"
        return s
    except Exception:
        return str(ts)

def _format_float(val, dec=2):
    if val is None or val == '' or val == '-':
        return '-'
    try:
        return f"{float(val):.{dec}f}"
    except Exception:
        return str(val)

@app.route("/api/scan/export", methods=["POST"])
def api_scan_export():
    try:
        import io
        from spot_enricher import extract_underlying_symbol, evaluate_spot_trend_and_t1
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Symbol", "Contract", "Side", "Tier", "Entry", "SL", "T1", "T2", "T3",
                         "AncherT", "EntryTime", "Result", "CF", "RR", "Engine", "Status",
                         "Spot_Trend", "Spot_T1_Target"])
        files = [("Daily", SCAN_DISPLAY_FILE), ("Bear", SCAN_DISPLAY_BEAR_FILE), ("Weekly Bull", paths.SCAN_DISPLAY_WEEKLY_FILE), ("Weekly Bear", paths.SCAN_DISPLAY_WEEKLY_BEAR_FILE), ("Stock EMA", EMA_DISPLAY_FILE_STOCK)]
        spot_eval_cache = {}
        for label, path in files:
            full = os.path.join(BASE_DIR, path)
            if not os.path.exists(full):
                continue
            with open(full) as f:
                data = json.load(f)
            if isinstance(data, dict) and "ema_engine" in data:
                data = data["ema_engine"]
            for section_name, status_tag in [("staged_trades", "Staged"), ("active_live", "Active"), ("carry_forward", "CarryFwd")]:
                for t in data.get(section_name, []):
                    raw_sym = t.get("contract") or t.get("symbol") or ""
                    underlying = extract_underlying_symbol(raw_sym)
                    if underlying and underlying not in spot_eval_cache:
                        spot_eval_cache[underlying] = evaluate_spot_trend_and_t1(None, underlying)
                    spot_trend, spot_t1 = spot_eval_cache.get(underlying, ("N/A", "N/A"))
                    formatted_spot_t1 = _format_float(spot_t1) if isinstance(spot_t1, (int, float)) else str(spot_t1)

                    tb_raw = t.get("tier_badge") or t.get("tier_label")
                    if not tb_raw:
                        t_num = int(t.get("tier", 2))
                        tb_raw = "🥇 T1" if t_num == 1 else ("🥈 T2" if t_num == 2 else "🥉 T3")

                    writer.writerow([
                        t.get("symbol", ""),
                        t.get("contract", ""),
                        t.get("side", ""),
                        tb_raw,
                        _format_float(t.get("entry") or t.get("entry_spot")),
                        _format_float(t.get("sl") or t.get("current_sl")),
                        _format_float(t.get("t1")),
                        _format_float(t.get("t2")),
                        _format_float(t.get("t3")),
                        _format_timestamp(t.get("candle_a_time")),
                        _format_timestamp(t.get("entry_time")),
                        _format_pattern_result(t.get("pattern") or t.get("result")),
                        "Yes" if t.get("carry_forward") else "No",
                        _format_float(t.get("rr")),
                        label,
                        status_tag,
                        spot_trend,
                        formatted_spot_t1
                    ])
        csv_bytes = output.getvalue().encode("utf-8-sig")
        return Response(csv_bytes, mimetype="text/csv",
                        headers={"Content-Disposition": f"attachment; filename=scan_export_{dt.now().strftime('%d_%m_%y_%H%M')}.csv"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/export/ema", methods=["POST"])
def api_export_ema():
    try:
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Symbol", "Contract", "Side", "Entry", "SL", "T1", "T2", "T3",
                         "Spot", "RR", "Timeframe", "Pattern", "AncherT", "EntryTime", "Status"])
        full = EMA_DISPLAY_FILE_STOCK if os.path.isabs(EMA_DISPLAY_FILE_STOCK) else os.path.join(BASE_DIR, EMA_DISPLAY_FILE_STOCK)
        if os.path.exists(full):
            with open(full, encoding="utf-8") as f:
                data = json.load(f)
            ema_payload = data.get("ema_engine", data) if isinstance(data, dict) else {}
            for section_name, status_tag in [("staged_trades", "Staged"), ("active_live", "Active"), ("carry_forward", "CarryFwd")]:
                for t in ema_payload.get(section_name, []):
                    side_val = t.get("side", "")
                    if not side_val:
                        cnt_str = str(t.get("contract") or t.get("symbol") or "").upper()
                        if "CE" in cnt_str:
                            side_val = "CE"
                        elif "PE" in cnt_str:
                            side_val = "PE"
                    writer.writerow([
                        t.get("symbol", ""),
                        t.get("contract", ""),
                        side_val,
                        _format_float(t.get("entry")),
                        _format_float(t.get("sl") or t.get("current_sl")),
                        _format_float(t.get("t1")),
                        _format_float(t.get("t2")),
                        _format_float(t.get("t3")),
                        _format_float(t.get("entry_spot")),
                        _format_float(t.get("rr")),
                        t.get("timeframe", ""),
                        t.get("pattern", ""),
                        _format_timestamp(t.get("candle_a_time")),
                        _format_timestamp(t.get("entry_time")),
                        status_tag
                    ])
        csv_bytes = output.getvalue().encode("utf-8-sig")
        return Response(csv_bytes, mimetype="text/csv",
                        headers={"Content-Disposition": f"attachment; filename=ema_export_{dt.now().strftime('%d_%m_%y_%H%M')}.csv"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/journal/clear", methods=["POST"])
def api_journal_clear():
    try:
        if os.path.exists(JOURNAL_FILE):
            open(JOURNAL_FILE, "w").close()
    except Exception:
        pass
    with data_lock:
        cached_data["journal"] = []
        cached_data["stats"] = compute_stats(cached_data.get("positions", {}), [])
    return jsonify({"ok": True})

@app.route("/api/programs/<prog_id>/start", methods=["POST"])
def api_start(prog_id):
    if prog_id not in PROGRAMS:
        return jsonify({"ok": False, "error": "Unknown program"})
    token = check_token_valid()
    if not token["valid"]:
        return jsonify({"ok": False, "error": token["reason"]})
    if prog_id == "ema_engine":
        cfg = load_config()
        tf = cfg.get("ema_engine", {}).get("timeframe", "1d")
        tu = cfg.get("ema_engine", {}).get("target_universe", "ALL")
        interval = int(cfg.get("ema_engine", {}).get("scan_interval", 300))
        ok, msg = start_ema_engine(timeframe=tf, is_options_mode=False, scan_interval=interval, target_universe=tu)
        return jsonify({"ok": ok, "error": None if ok else msg})
    ok = start_program(prog_id)
    return jsonify({"ok": ok, "error": None if ok else "Start failed"})

@app.route("/api/programs/<prog_id>/stop", methods=["POST"])
def api_stop(prog_id):
    if prog_id not in PROGRAMS:
        return jsonify({"ok": False, "error": "Unknown program"})
    if prog_id == "ema_engine":
        ok, msg = stop_ema_engine(is_options_mode=False)
        return jsonify({"ok": ok, "error": None if ok else msg})
    ok = stop_program(prog_id)
    return jsonify({"ok": ok})

ANCHOR_SCAN_REQUEST_FILE = os.path.join("output", "monitor", "anchor_scan_request.txt")
ANCHOR_SCAN_STOP_FILE = os.path.join("output", "monitor", "anchor_scan_stop.txt")

@app.route("/api/anchor/scan", methods=["POST"])
def api_anchor_scan():
    data = request.get_json(silent=True) or {}
    engine = data.get("engine", "index")
    try:
        with data_lock:
            cached_data["anchor_status"]["running"] = True
            cached_data["anchor_status"]["engine"] = engine
            cached_data["anchor_status"]["requested_at"] = time.time()
        if os.path.exists(ANCHOR_SCAN_STOP_FILE):
            os.remove(ANCHOR_SCAN_STOP_FILE)
        # Launch a dedicated --anchor-only subprocess. We intentionally do NOT
        # write ANCHOR_SCAN_REQUEST_FILE here: a running engine also polls that
        # file in its main loop, which would cause the anchor scan to run twice.
        script = PROGRAMS.get(engine, {}).get("file")
        if script:
            script_path = os.path.join(BASE_DIR, script)
            subprocess.Popen([sys.executable, script_path, "--anchor-only"],
                             cwd=BASE_DIR, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/anchor/stop", methods=["POST"])
def api_anchor_stop():
    try:
        os.makedirs(os.path.dirname(ANCHOR_SCAN_STOP_FILE), exist_ok=True)
        with open(ANCHOR_SCAN_STOP_FILE, "w") as f:
            f.write("stop")
        with data_lock:
            cached_data["anchor_status"]["running"] = False
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/anchor/status")
def api_anchor_status():
    with data_lock:
        st = dict(cached_data["anchor_status"])
    if not st.get("running"):
        still_running = os.path.exists(ANCHOR_SCAN_REQUEST_FILE) and not os.path.exists(ANCHOR_SCAN_STOP_FILE)
        if still_running:
            st["running"] = True
            if not st.get("engine"):
                try:
                    with open(ANCHOR_SCAN_REQUEST_FILE) as f:
                        st["engine"] = f.read().strip()
                except Exception:
                    pass
    return jsonify(st)

@app.route("/api/logs/clear", methods=["POST"])
def api_logs_clear():
    log_files = [DAILY_LOG_FILE, BEAR_LOG_FILE, EMA_LOG_FILE]
    for lf in log_files:
        try:
            if os.path.exists(lf):
                open(lf, "w").close()
        except Exception:
            pass
    with data_lock:
        for pid in PROGRAMS:
            cached_data["log_tail"][pid] = []
            cached_data["scans"][pid] = []
            cached_data["scan_summary"][pid] = {"anchors": {}, "abc_matches": {}}
    return jsonify({"ok": True})

@app.route("/api/trades")
def api_trades():
    engine = request.args.get("engine")
    active_only = request.args.get("active", "false").lower() == "true"
    if active_only:
        return jsonify(trade_db.get_active_trades(engine))
    return jsonify(trade_db.get_all_trades(engine))

@app.route("/api/export/monthly", methods=["POST"])
def api_export_monthly():
    result = run_monthly_export()
    return jsonify({"ok": True, **result})

@app.route("/api/live-execution/nifty50", methods=["GET", "POST"])
def api_live_execution():
    if request.method == "POST":
        enabled = request.get_json(force=True, silent=True).get("enabled", False)
        flag_path = LIVE_EXECUTION_FLAG
        if enabled:
            with open(flag_path, "w") as f:
                f.write("1")
        else:
            if os.path.exists(flag_path):
                os.remove(flag_path)
        with data_lock:
            cached_data["live_execution"] = enabled
        return jsonify({"ok": True, "enabled": enabled})
    return jsonify({"enabled": os.path.exists(LIVE_EXECUTION_FLAG)})

@app.route("/api/live-execution/index", methods=["GET", "POST"])
def api_live_execution_index():
    if request.method == "POST":
        enabled = request.get_json(force=True, silent=True).get("enabled", False)
        flag_path = LIVE_EXECUTION_FLAG_INDEX
        if enabled:
            with open(flag_path, "w") as f:
                f.write("1")
        else:
            if os.path.exists(flag_path):
                os.remove(flag_path)
        with data_lock:
            cached_data["live_execution_index"] = enabled
        return jsonify({"ok": True, "enabled": enabled})
    return jsonify({"enabled": os.path.exists(LIVE_EXECUTION_FLAG_INDEX)})

@app.route("/api/edit-lock", methods=["POST"])
def api_edit_lock():
    data = request.json or {}
    sym = data.get("symbol")
    active = data.get("active", False)
    if sym:
        clean_s = str(sym).replace(" ", "").upper()
        if active:
            ACTIVE_EDIT_LOCKS.add(clean_s)
            logging.info(f"[EDIT LOCK ON] Automated exit execution paused for {clean_s}")
        else:
            ACTIVE_EDIT_LOCKS.discard(clean_s)
            logging.info(f"[EDIT LOCK OFF] Automated exit execution resumed for {clean_s}")
    return jsonify({"ok": True})

@app.route("/api/update-position", methods=["POST"])
def api_update_position():
    data = request.get_json(force=True, silent=True) or {}
    engine = data.get("engine", "nifty50")
    symbol = data.get("symbol", "")
    current_sl = data.get("current_sl")
    t1 = data.get("t1")
    t2 = data.get("t2")
    t3 = data.get("t3")
    entry_price = data.get("entry_price")
    if not symbol or (current_sl is None and t1 is None and t2 is None and t3 is None):
        return jsonify({"ok": False, "error": "symbol and at least one level required"}), 400
    vals = {}
    if current_sl is not None and str(current_sl).strip() != "": vals["current_sl"] = float(current_sl)
    if t1 is not None and str(t1).strip() != "": vals["t1"] = float(t1)
    if t2 is not None and str(t2).strip() != "": vals["t2"] = float(t2)
    if t3 is not None and str(t3).strip() != "": vals["t3"] = float(t3)
    if entry_price is not None and str(entry_price).strip() != "": vals["entry_spot"] = float(entry_price)
    vals["user_edited"] = True

    clean_target = str(symbol).replace(" ", "").upper()

    write_sl_overrides(engine, symbol, vals, (engine, "nifty50", "index", "daily"))

    clear_executed_exit(symbol)
    clear_executed_exit(clean_target)
    ACTIVE_EDIT_LOCKS.discard(clean_target)

    matched = False
    def _is_match(item_sym, item_cnt):
        c_sym = str(item_sym or "").replace(" ", "").upper()
        c_cnt = str(item_cnt or "").replace(" ", "").upper()
        if not clean_target: return False
        if c_cnt and clean_target == c_cnt: return True
        if c_sym and clean_target == c_sym and (not c_cnt or c_cnt == c_sym): return True
        is_opt_tgt = ("CE" in clean_target or "PE" in clean_target) and any(c.isdigit() for c in clean_target)
        if not is_opt_tgt and c_sym == clean_target: return True
        return False

    with data_lock:
        update_keys = list(vals.keys())

        # 1. Update in-memory all_trades
        for t in cached_data.get("all_trades", []):
            if _is_match(t.get("symbol"), t.get("contract")):
                matched = True
                for k in update_keys: t[k] = vals[k]
                tid = t.get("id")
                if tid:
                    trade_db.update_trade(tid, vals)

        # 2. Update in-memory positions
        for pos_key, pos in (cached_data.get("positions", {}).items() if isinstance(cached_data.get("positions"), dict) else enumerate(cached_data.get("positions", []))):
            if isinstance(pos, dict):
                if _is_match(pos.get("symbol"), pos.get("contract")):
                    matched = True
                    for k in update_keys: pos[k] = vals[k]
                    tid = pos.get("id")
                    if tid:
                        trade_db.update_trade(tid, vals)

        # 3. Update in-memory kite_positions so UI refreshes immediately
        for kp in cached_data.get("kite_positions", []):
            if _is_match(kp.get("symbol"), kp.get("contract")):
                for k in update_keys: kp[k] = vals[k]

        if not matched:
            contract = symbol
            exchange = "NSE"
            for kp in cached_data.get("kite_positions", []):
                if _is_match(kp.get("symbol"), kp.get("contract")):
                    contract = kp.get("contract", symbol)
                    exchange = kp.get("exchange", "NSE")
                    break
            is_stock = exchange == "NSE"
            trade_data = {"contract": contract, "entry_spot": vals.get("entry_spot", 0), "position_type": "stock" if is_stock else "option"}
            trade_data.update(vals)
            db_symbol = resolve_underlying(symbol or contract, engine)
            tid, _created = trade_db.create_trade(engine, db_symbol, trade_data)
            entry = {"symbol": db_symbol, "contract": contract, "id": tid, "engine": engine, "status": "ACTIVE", "position_type": "stock" if is_stock else "option"}
            entry.update(vals)
            cached_data["all_trades"].append(entry)
            cached_data["positions"][symbol] = entry
            logging.info(f"[OVERRIDE] Created new DB trade for {engine}/{symbol}")
    logging.info(f"Position override queued: {engine}/{symbol} {vals}")
    return jsonify({"ok": True})

# ──────────────────────────────────────────────
#  1-CLICK BUY SCANNED TRADE API
# ──────────────────────────────────────────────
@app.route("/api/buy-scanned-trade", methods=["POST"])
def api_buy_scanned_trade():
    try:
        data = request.json or {}
        symbol = data.get("symbol")
        contract = data.get("contract") or symbol
        side = data.get("side", "CE")
        entry_spot = float(data.get("entry_spot") or 0)
        current_sl = float(data.get("current_sl") or data.get("sl") or 0)
        t1 = float(data.get("t1") or 0)
        t2 = float(data.get("t2") or 0)
        t3 = float(data.get("t3") or 0)
        engine = data.get("engine", "daily")

        if not symbol:
            return jsonify({"ok": False, "error": "symbol is required"}), 400

        c_str = str(contract).upper()
        if "SENSEX" in c_str or "BSE" in c_str:
            exch = "BFO"
        elif "CE" in c_str or "PE" in c_str or "NIFTY" in c_str or "BANK" in c_str:
            exch = "NFO"
        else:
            exch = "NSE"

        global _kite_session
        order_id = None
        ltp = 0
        if not _kite_session:
            try:
                from common.trading_core import load_kite_session
                api_k, acc_t = load_kite_session()
                if api_k and acc_t:
                    from kiteconnect import KiteConnect
                    _kite_session = KiteConnect(api_key=api_k)
                    _kite_session.set_access_token(acc_t)
            except Exception as init_err:
                logging.warning(f"1-Click Buy auto-init kite session failed: {init_err}")

        if _kite_session:
            # ── Portfolio Risk & Sector Caps Enforcement ──
            try:
                from common.portfolio_risk import check_portfolio_risk_caps
                cfg_all = load_config()
                cap_amount = float(cfg_all.get(engine, {}).get("capital") or 100000.0)
                cand_tier = int(data.get("tier") or 2)
                p_allowed, p_reason, _ = check_portfolio_risk_caps(
                    engine=engine,
                    symbol=symbol,
                    candidate_tier=cand_tier,
                    capital=cap_amount,
                    include_db_trades=True
                )
                if not p_allowed:
                    logging.warning(f"[1-CLICK BUY REJECTED] {symbol} ({contract}): {p_reason}")
                    return jsonify({"ok": False, "error": f"Portfolio Risk Guard: {p_reason}"}), 400
            except Exception as p_err:
                logging.warning(f"Portfolio risk check error in 1-Click Buy: {p_err}")

            try:
                q_key = f"{exch}:{contract}"
                q = _kite_session.quote([q_key])
                ltp = float(q.get(q_key, {}).get("last_price", 0))
                depth = q.get(q_key, {}).get("depth", {}).get("sell", [])
                bm = float(data.get("benchmark") or 0)
                if bm > 0:
                    price = round(bm * 1.005, 1)
                else:
                    ask = float(depth[0].get("price", 0)) if (depth and len(depth) > 0) else 0
                    price = round((ask if ask > 0 else ltp) * 1.005, 1)
                    if price <= 0:
                        price = round(entry_spot * 1.005, 1)

                from common.trading_core import STOCK_REGISTRY
                lot_size = STOCK_REGISTRY.get(symbol, {}).get("lot_size", 1) if exch != "NSE" else 1
                prod = _kite_session.PRODUCT_CNC if exch == "NSE" else _kite_session.PRODUCT_NRML

                from common.trading_core import is_market_open
                market_open = is_market_open()
                order_variety = _kite_session.VARIETY_REGULAR if market_open else _kite_session.VARIETY_AMO

                try:
                    order_id = _kite_session.place_order(
                        variety=order_variety,
                        tradingsymbol=contract,
                        exchange=exch,
                        transaction_type=_kite_session.TRANSACTION_TYPE_BUY,
                        quantity=lot_size,
                        order_type=_kite_session.ORDER_TYPE_LIMIT,
                        price=price,
                        product=prod
                    )
                    v_label = "regular order" if order_variety == _kite_session.VARIETY_REGULAR else "After Market Order (AMO)"
                    logging.info(f"[1-CLICK BUY] Placed {v_label} for {contract} on {exch} (Order ID: {order_id})")
                except Exception as first_err:
                    if order_variety == _kite_session.VARIETY_REGULAR and "After Market Order" in str(first_err):
                        logging.info(f"[1-CLICK BUY] Regular order rejected; retrying with VARIETY_AMO for {contract}...")
                        order_id = _kite_session.place_order(
                            variety=_kite_session.VARIETY_AMO,
                            tradingsymbol=contract,
                            exchange=exch,
                            transaction_type=_kite_session.TRANSACTION_TYPE_BUY,
                            quantity=lot_size,
                            order_type=_kite_session.ORDER_TYPE_LIMIT,
                            price=price,
                            product=prod
                        )
                        logging.info(f"[1-CLICK BUY] Placed After Market Order (AMO) for {contract} on {exch} (Order ID: {order_id})")
                    else:
                        raise first_err
            except Exception as k_err:
                logging.warning(f"[1-CLICK BUY KITE ORDER WARNING] {contract}: {k_err}")
                err_msg = str(k_err)
                from common.trading_core import is_market_open
                if not is_market_open():
                    err_msg = f"Markets are closed (trading hours: 09:15 - 15:30 IST). Broker response: {err_msg}"
                return jsonify({"ok": False, "error": f"Kite Order Placement Failed: {err_msg}"}), 400

        candle_a_time = data.get("candle_a_time") or data.get("CandleATime")
        benchmark = data.get("benchmark")
        anchor_floor = data.get("anchor_floor")
        direction = data.get("direction")

        if not candle_a_time:
            try:
                disp_backfill = SCAN_DISPLAY_FILE
                if os.path.exists(disp_backfill):
                    with open(disp_backfill, "r", encoding="utf-8") as fh:
                        _sd = json.load(fh)
                    c_target = str(contract).replace(" ", "").upper()
                    for _cat in ["staged_trades", "all_staged_today", "carry_forward", "active_live"]:
                        for item in _sd.get(_cat) or []:
                            i_cnt = str(item.get("contract") or item.get("symbol") or "").replace(" ", "").upper()
                            if i_cnt == c_target:
                                candle_a_time = candle_a_time or item.get("candle_a_time") or item.get("CandleATime")
                                if benchmark is None:
                                    benchmark = item.get("benchmark")
                                if anchor_floor is None:
                                    anchor_floor = item.get("anchor_floor")
                                if not direction:
                                    direction = item.get("direction")
                                break
                        if candle_a_time:
                            break
            except Exception as bf_err:
                logging.warning(f"1-Click Buy display backfill skipped: {bf_err}")

        # Align option entry price with real execution/LTP price if spot price was passed or stale
        if exch != "NSE" and ltp > 0:
            if entry_spot <= 0 or (abs(entry_spot - ltp) / max(entry_spot, ltp) > 0.50):
                logging.info(f"[PRICE ALIGN] Overriding divergent entry_spot {entry_spot} with live option LTP {ltp} for {contract}")
                entry_spot = ltp

        trade_data = {
            "contract": contract,
            "entry_spot": entry_spot,
            "current_sl": current_sl,
            "t1": t1,
            "t2": t2,
            "t3": t3,
            "side": side,
            "pattern": "1CLICK_BUY",
            "position_type": "stock" if exch == "NSE" else "option",
            "user_edited": True,
            "entry_time": dt.now().isoformat()
        }
        if candle_a_time:
            trade_data["candle_a_time"] = candle_a_time
            trade_data["CandleATime"] = candle_a_time
        if benchmark is not None:
            trade_data["benchmark"] = benchmark
        if anchor_floor is not None:
            trade_data["anchor_floor"] = anchor_floor
        if direction:
            trade_data["direction"] = direction
        try:
            from trading_core import contract_is_expired
            if contract_is_expired(contract):
                return jsonify({"ok": False, "error": f"Contract {contract} is expired. Cannot place 1-Click Buy."}), 400
        except Exception as exp_check_err:
            logging.warning(f"1-Click Buy expiry check skipped: {exp_check_err}")
        symbol = resolve_underlying(symbol or contract, engine)
        tid, _created = trade_db.create_trade(engine, symbol, trade_data)
        clear_executed_exit(contract)

        return jsonify({
            "ok": True,
            "message": f"Successfully placed 1-Click BUY for {contract}" + (f" (Order ID: {order_id})" if order_id else ""),
            "trade_id": tid
        })
    except Exception as e:
        logging.error(f"1-Click Buy API failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

# ──────────────────────────────────────────────
#  MANUAL EXIT POSITION API (SINGLE & ALL)
# ──────────────────────────────────────────────
@app.route("/api/exit-position", methods=["POST"])
def api_exit_position():
    try:
        data = request.json or {}
        symbol = data.get("symbol", "")
        contract = data.get("contract") or symbol
        engine = data.get("engine", "nifty50")

        if not symbol and not contract:
            return jsonify({"ok": False, "error": "Symbol or contract name required"}), 400

        target_str = str(contract or symbol).replace(" ", "").upper()
        now_str = dt.now().strftime("%Y-%m-%d %H:%M:%S")

        all_t = trade_db.get_all_trades()
        exited_ids = []
        for t in all_t:
            t_sym = str(t.get("symbol") or "").replace(" ", "").upper()
            t_cnt = str(t.get("contract") or "").replace(" ", "").upper()
            if t.get("status") == "ACTIVE" and (target_str in (t_sym, t_cnt) or t_sym in target_str or t_cnt in target_str):
                trade_db.update_trade(t["id"], {
                    "status": "USER_EXIT",
                    "exit_time": now_str,
                    "result": "USER_EXIT",
                    "updated_at": now_str
                })
                exited_ids.append(t["id"])

        global _kite_session
        if _kite_session:
            try:
                c_str = target_str
                if "SENSEX" in c_str or "BSE" in c_str:
                    exch = "BFO"
                elif "CE" in c_str or "PE" in c_str or "NIFTY" in c_str or "BANK" in c_str:
                    exch = "NFO"
                else:
                    exch = "NSE"

                pos_obj = {
                    "contract": contract,
                    "symbol": symbol,
                    "exchange": exch,
                    "quantity": data.get("quantity", 0)
                }
                from common.trading_core import close_position as shared_close
                shared_close(_kite_session, pos_obj, True)
            except Exception as k_err:
                logging.warning(f"Live exit execution warning for {contract}: {k_err}")

        for disp_path in [SCAN_DISPLAY_FILE, SCAN_DISPLAY_INDEX_FILE]:
            if os.path.exists(disp_path):
                try:
                    with open(disp_path, "r", encoding="utf-8") as f:
                        sd = json.load(f)
                    sd["active_positions"] = [p for p in sd.get("active_positions", []) if str(p.get("contract") or p.get("symbol")).replace(" ", "").upper() != target_str]
                    sd["active_live"] = [p for p in sd.get("active_live", []) if str(p.get("contract") or p.get("symbol")).replace(" ", "").upper() != target_str]
                    with open(disp_path, "w", encoding="utf-8") as f:
                        json.dump(sd, f, indent=2)
                except Exception as e:
                    pass

        return jsonify({"ok": True, "message": f"Manual EXIT executed for {contract or symbol}"})
    except Exception as e:
        logging.error(f"Manual Exit API failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/exit-all-positions", methods=["POST"])
def api_exit_all_positions():
    try:
        now_str = dt.now().strftime("%Y-%m-%d %H:%M:%S")
        all_t = trade_db.get_all_trades()
        exited_count = 0
        for t in all_t:
            if t.get("status") == "ACTIVE":
                trade_db.update_trade(t["id"], {
                    "status": "USER_EXIT",
                    "exit_time": now_str,
                    "result": "USER_EXIT",
                    "updated_at": now_str
                })
                exited_count += 1

        for disp_path in [SCAN_DISPLAY_FILE, SCAN_DISPLAY_INDEX_FILE]:
            if os.path.exists(disp_path):
                try:
                    with open(disp_path, "r", encoding="utf-8") as f:
                        sd = json.load(f)
                    sd["active_positions"] = []
                    sd["active_live"] = []
                    with open(disp_path, "w", encoding="utf-8") as f:
                        json.dump(sd, f, indent=2)
                except Exception as e:
                    pass

        return jsonify({"ok": True, "message": f"Successfully EXITED all ({exited_count}) active positions"})
    except Exception as e:
        logging.error(f"Exit All API failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

EXPORT_STATE_FILE = paths.EXPORT_STATE_FILE

# ──────────────────────────────────────────────
#  DAILY SELF-LEARNING TRADE JOURNAL API
# ──────────────────────────────────────────────
@app.route("/api/journal/get", methods=["GET"])
def api_journal_get():
    try:
        from daily_trade_journal import load_journal_entries
        return jsonify(load_journal_entries())
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/journal/analytics", methods=["GET"])
def api_journal_analytics():
    try:
        from daily_trade_journal import get_trade_journal_analytics
        return jsonify({"ok": True, "data": get_trade_journal_analytics()})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/journal/sync", methods=["POST"])
def api_journal_sync():
    try:
        from daily_trade_journal import generate_daily_journal
        req = request.json or {}
        dt_str = req.get("date")
        entries = generate_daily_journal(dt_str, kite=_kite_session)
        return jsonify({"ok": True, "count": len(entries), "entries": entries})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/journal/update", methods=["POST"])
def api_journal_update():
    try:
        from daily_trade_journal import load_journal_entries, save_journal_entries
        data = request.json or {}
        symbol = data.get("symbol")
        date_str = data.get("date")
        trade_id = data.get("trade_id")
        remarks = data.get("remarks")
        lesson = data.get("lesson")
        if trade_id and lesson is not None:
            trade_db.update_self_learning_lesson(trade_id, lesson)
        if not symbol or not date_str:
            return jsonify({"ok": True, "message": "Updated trade_db lesson"})
        entries = load_journal_entries()
        updated = False
        for e in entries:
            if e.get("Date") == date_str and (e.get("Symbol") == symbol or symbol in e.get("Symbol", "")):
                if remarks is not None: e["Analysis_Remarks"] = remarks
                if lesson is not None: e["Self_Learning_Lesson"] = lesson
                updated = True
        if updated:
            save_journal_entries(entries)
            return jsonify({"ok": True})
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/journal/export", methods=["GET", "POST"])
def api_journal_export():
    try:
        import io
        from daily_trade_journal import load_journal_entries
        entries = load_journal_entries()
        headers = [
            "Date", "Engine", "Symbol", "Side", "Timeframe", "Pattern", "Tier", "Swing_Waves",
            "Entry_Time", "Entry_Price", "Exit_Time", "Exit_Price",
            "SL", "T1", "T2", "T3", "Quantity", "Lot_Size",
            "PnL_Rs", "PnL_Pct", "Outcome", "Analysis_Remarks", "Self_Learning_Lesson"
        ]
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for e in entries:
            writer.writerow(e)
        csv_data = output.getvalue()
        fname = f"trade_journal_export_{dt.now().strftime('%Y_%m_%d_%H%M')}.csv"
        return Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={fname}"}
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/journal/export/excel", methods=["GET", "POST"])
def api_journal_export_excel():
    try:
        import openpyxl, io
        from openpyxl.styles import Font, PatternFill, Alignment
        from daily_trade_journal import load_journal_entries

        entries = load_journal_entries()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trade Journal"

        headers = [
            "Date", "Engine", "Symbol", "Side", "Timeframe", "Pattern", "Tier", "Swing Waves",
            "Entry Time", "Entry Price", "Exit Time", "Exit Price",
            "SL", "T1", "T2", "T3", "Quantity", "Lot Size",
            "PnL (₹)", "PnL (%)", "Outcome", "Analysis Remarks", "Self-Learning Lesson"
        ]

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for e in entries:
            ws.append([
                e.get("Date", ""),
                e.get("Engine", ""),
                e.get("Symbol", ""),
                e.get("Side", ""),
                e.get("Timeframe", ""),
                e.get("Pattern", ""),
                e.get("Entry_Time", ""),
                e.get("Entry_Price", ""),
                e.get("Exit_Time", ""),
                e.get("Exit_Price", ""),
                e.get("SL", ""),
                e.get("T1", ""),
                e.get("T2", ""),
                e.get("T3", ""),
                e.get("Quantity", ""),
                e.get("Lot_Size", ""),
                e.get("PnL_Rs", 0),
                e.get("PnL_Pct", ""),
                e.get("Outcome", ""),
                e.get("Analysis_Remarks", ""),
                e.get("Self_Learning_Lesson", "")
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        ws2 = wb.create_sheet(title="Performance Summary")
        ws2.append(["Metric", "Value"])
        ws2.cell(row=1, column=1).font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=2).font = header_font
        ws2.cell(row=1, column=2).fill = header_fill

        total_trades = len(entries)
        wins = sum(1 for e in entries if float(e.get("PnL_Rs") or 0) > 0)
        losses = sum(1 for e in entries if float(e.get("PnL_Rs") or 0) < 0)
        win_rate = (wins / total_trades * 100) if total_trades > 0 else 0
        total_pnl = sum(float(e.get("PnL_Rs") or 0) for e in entries)

        ws2.append(["Total Trades", total_trades])
        ws2.append(["Winning Trades", wins])
        ws2.append(["Losing Trades", losses])
        ws2.append(["Win Rate (%)", f"{win_rate:.1f}%"])
        ws2.append(["Total Net PnL (₹)", f"₹{total_pnl:.2f}"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        fname = f"trade_journal_export_{dt.now().strftime('%Y_%m_%d_%H%M')}.xlsx"
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname}"}
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ──────────────────────────────────────────────
#  INTERACTIVE NEGATION ANALYZER API
# ──────────────────────────────────────────────
@app.route("/api/analyze-trade", methods=["POST"])
def api_analyze_trade():
    try:
        data = request.json or {}
        symbol = str(data.get("symbol", "")).strip().upper()
        entry_price = float(data.get("entry_price", 0)) if data.get("entry_price") else 0.0
        timeframe = str(data.get("timeframe", "75min")).strip()
        engine = str(data.get("engine", "nifty50")).strip()

        if not symbol:
            return jsonify({"ok": False, "error": "Valid Symbol or Contract Name required"}), 400

        kite = None
        try:
            api_k, acc_t = load_kite_session()
            kite = KiteConnect(api_key=api_k, access_token=acc_t)
        except Exception:
            kite = None

        if timeframe == "30minute":
            timeframe_entry = "30minute"
            timeframe_anchor = "30minute"
        else:
            timeframe_entry = "15minute" if timeframe in ["15minute", "75min", "60minute"] else timeframe
            timeframe_anchor = "75min" if timeframe in ["15minute", "75min"] else ("60minute" if timeframe == "60minute" else timeframe)

        analysis = derive_sl_targets_for_contract(kite, symbol, entry_price, timeframe_entry, timeframe_anchor)
        if not analysis:
            sl_val = round(entry_price * 0.90, 2) if entry_price > 0 else 0.0
            analysis = {
                "entry_price": entry_price,
                "current_sl": sl_val,
                "t1": None, "t2": None, "t3": None,
                "pattern": "NEGATION_ESTIMATED"
            }

        resolved_entry = float(analysis.get("entry_price") or entry_price or 0.0)
        sl_val = analysis.get("current_sl", round(resolved_entry * 0.90, 2) if resolved_entry > 0 else 0.0)
        t1_val = analysis.get("t1")
        t2_val = analysis.get("t2")
        t3_val = analysis.get("t3")

        risk = (resolved_entry - sl_val) if (resolved_entry > 0 and sl_val < resolved_entry) else 0
        rr = round((t1_val - resolved_entry) / risk, 2) if (t1_val and risk > 0) else 0.0

        return jsonify({
            "ok": True,
            "symbol": symbol,
            "contract": symbol,
            "entry_price": resolved_entry,
            "current_sl": sl_val,
            "t1": t1_val if t1_val else "N/A",
            "t2": t2_val if t2_val else "N/A",
            "t3": t3_val if t3_val else "N/A",
            "rr": rr,
            "pattern": analysis.get("pattern", "NEGATION_DERIVED"),
            "engine": engine
        })
    except Exception as e:
        logging.error(f"Analyze Trade API failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/get-chart-data", methods=["GET"])
def api_get_chart_data():
    try:
        contract = str(request.args.get("symbol", "")).strip().upper()
        chart_type = str(request.args.get("type", "spot")).strip().lower()
        tf = str(request.args.get("timeframe", "30minute")).strip()

        if not contract:
            return jsonify({"ok": False, "error": "Symbol is required"}), 400

        api_k, acc_t = load_kite_session()
        kite = KiteConnect(api_key=api_k, access_token=acc_t)

        token = None
        exchange = "NSE"
        spot_symbol = contract
        spot_token = None

        from trading_core import STOCK_REGISTRY
        if spot_symbol in STOCK_REGISTRY:
            spot_token = STOCK_REGISTRY[spot_symbol]["token"]
        elif spot_symbol in ["NIFTY", "NIFTY 50", "NIFTY50"]:
            spot_symbol = "NIFTY"
            spot_token = 256265
        elif spot_symbol in ["BANKNIFTY", "NIFTY BANK"]:
            spot_symbol = "BANKNIFTY"
            spot_token = 260105
        elif spot_symbol in ["SENSEX", "BSESN"]:
            spot_symbol = "SENSEX"
            spot_token = 265
        elif contract in STOCK_REGISTRY:
            spot_token = STOCK_REGISTRY[contract]["token"]
            spot_symbol = contract

        try:
            ltp_res = kite.ltp([f"NSE:{spot_symbol}"])
            if ltp_res and f"NSE:{spot_symbol}" in ltp_res:
                spot_token = ltp_res[f"NSE:{spot_symbol}"]["instrument_token"]
        except Exception:
            pass

        target_token = spot_token
        target_symbol = spot_symbol
        target_exchange = "NSE" if (spot_symbol in STOCK_REGISTRY or spot_symbol in ["NIFTY", "BANKNIFTY"]) else "BSE"

        if not target_token:
            return jsonify({"ok": False, "error": f"Instrument token not found for {contract}"}), 400

        from datetime import datetime as dt, timedelta
        from trading_core import fetch_and_resample_candles

        ref_now = dt.now()
        from_date = (ref_now - timedelta(days=60)).strftime("%Y-%m-%d")
        to_date = ref_now.strftime("%Y-%m-%d")

        df_candles = fetch_and_resample_candles(kite, target_token, from_date, to_date, tf)
        if df_candles is None or df_candles.empty:
            return jsonify({"ok": False, "error": f"No candle data available for {target_symbol}"}), 400

        import pandas as pd
        candles = []
        for _, r in df_candles.iterrows():
            c_dt = pd.to_datetime(r['date'])
            ts = int(c_dt.timestamp())
            candles.append({
                "time": ts,
                "open": round(float(r['open']), 2),
                "high": round(float(r['high']), 2),
                "low": round(float(r['low']), 2),
                "close": round(float(r['close']), 2),
                "volume": int(r['volume']) if 'volume' in r else 0
            })

        kite_url = f"https://kite.zerodha.com/chart/ext/tvc/{target_exchange}/{target_symbol}/{target_token}"

        return jsonify({
            "ok": True,
            "symbol": target_symbol,
            "contract": contract,
            "spot_symbol": spot_symbol,
            "chart_type": chart_type,
            "exchange": target_exchange,
            "token": target_token,
            "timeframe": tf,
            "candles": candles,
            "kite_url": kite_url
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500




def run_monthly_export():
    import openpyxl
    from collections import defaultdict
    completed = trade_db.get_completed_trades()
    if not completed:
        return {"exported": 0, "sheets": []}
    groups = defaultdict(list)
    for t in completed:
        ts = t.get("exit_time") or t.get("updated_at") or t.get("created_at") or ""
        parts = ts.split(" ")[0].split("-") if " " in ts else ts.split("-")
        key = (parts[0], parts[1]) if len(parts) >= 2 else ("unknown", "00")
        groups[key].append(t)
    out_dir = paths.EXPORTS_DIR
    os.makedirs(out_dir, exist_ok=True)
    xl_path = paths.TRADE_ARCHIVE_XLSX
    sheet_names = []
    if os.path.exists(xl_path):
        wb = openpyxl.load_workbook(xl_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    headers = ["ID", "Engine", "Symbol", "Pattern", "Status", "Entry Spot", "SL", "T1", "T2", "T3",
               "Trailing Stage", "Lot Size", "Position Size", "Entry Date", "Exit Date", "P&L %", "Side", "Contract"]
    for (year, month), trades in sorted(groups.items()):
        month_names = ["", "January","February","March","April","May","June","July","August","September","October","November","December"]
        sheet_name = f"{month_names[int(month)]} {year}" if month.isdigit() and 1 <= int(month) <= 12 else f"{month} {year}"
        sheet_names.append(sheet_name)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for t in trades:
            entry_dt = (t.get("created_at") or "").split(" ")[0]
            exit_dt = (t.get("exit_time") or t.get("updated_at") or "").split(" ")[0]
            ws.append([
                t.get("id", ""), t.get("engine", ""), t.get("symbol", ""), t.get("pattern", ""),
                t.get("status", ""), t.get("entry_spot", ""), t.get("current_sl", ""),
                t.get("t1", ""), t.get("t2", ""), t.get("t3", ""),
                t.get("trailing_stage", ""), t.get("lot_size", ""), t.get("position_size", ""),
                entry_dt, exit_dt,
                t.get("pnl_percent", ""), t.get("side", ""), t.get("contract", "")
            ])
    wb.save(xl_path)
    exported_ids = [t["id"] for t in completed]
    trade_db.remove_trades(exported_ids)
    return {"exported": len(completed), "sheets": sheet_names}

def auto_export_if_new_month():
    now = dt.now()
    current_month = now.strftime("%Y-%m")
    try:
        if os.path.exists(EXPORT_STATE_FILE):
            with open(EXPORT_STATE_FILE) as f:
                state = json.load(f)
                last = state.get("last_export_month", "")
        else:
            last = ""
        if current_month > last:
            result = run_monthly_export()
            if result["exported"] > 0:
                print(f"Auto-export: {result['exported']} trades to {', '.join(result['sheets'])}")
            with open(EXPORT_STATE_FILE, "w") as f:
                json.dump({"last_export_month": current_month}, f)
    except Exception as e:
        print(f"Auto-export error: {e}")

def standalone_position_monitor_daemon():
    """Standalone 24/7 background position monitor for Port 5051.
    Starts immediately upon application launch and continuously guards ALL active
    trades (SL, Trailing SL, T1/T2/T3, Emergency Hard SL) as soon as Kite session is valid,
    without requiring any scanner or engine to be manually started.
    """
    logging.info("[STANDALONE_POSITION_MONITOR_STOCK] Background position guardian initialized.")
    while True:
        try:
            from session import load_kite_session, ensure_kite_session
            from position_monitor import monitor_all_active_positions
            global _kite_session
            if not _kite_session:
                try:
                    api_k, acc_t = load_kite_session(TOKEN_FILE)
                    ks = KiteConnect(api_key=api_k)
                    ks.set_access_token(acc_t)
                    _kite_session = ks
                except Exception:
                    _kite_session = None
            else:
                try:
                    ensure_kite_session(_kite_session, TOKEN_FILE)
                except Exception:
                    pass

            if _kite_session:
                live_stock = os.path.exists(LIVE_EXECUTION_FLAG)
                monitor_all_active_positions(_kite_session, live=live_stock)
        except Exception as e:
            logging.debug(f"[STANDALONE_POSITION_MONITOR_STOCK] Iteration error: {e}")
        time.sleep(2)

def main():
    os.makedirs(paths.INPUT_DIR, exist_ok=True)
    os.makedirs(paths.LOGS_DIR, exist_ok=True)
    os.makedirs(paths.MONITOR_DIR, exist_ok=True)
    os.makedirs(paths.EXPORTS_DIR, exist_ok=True)
    threading.Thread(target=auto_export_if_new_month, daemon=True).start()
    worker = threading.Thread(target=refresh_data, daemon=True)
    worker.start()
    monitor_worker = threading.Thread(target=standalone_position_monitor_daemon, daemon=True)
    monitor_worker.start()
    print(f"Trading Control Center starting on http://localhost:{DASHBOARD_PORT}")
    print(f"Refresh interval: {REFRESH_SECONDS}s")
    print("Available programs:")
    for pid, p in PROGRAMS.items():
        print(f"  [{pid}] {p['name']}")
    app.run(host="0.0.0.0", port=DASHBOARD_PORT, debug=False, use_reloader=False)

if __name__ == "__main__":
    main()

