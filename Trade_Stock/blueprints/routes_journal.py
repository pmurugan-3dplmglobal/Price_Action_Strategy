"""
routes_journal.py — Journal and trade DB routes:
  /api/journal/clear
  /api/journal/get
  /api/journal/analytics
  /api/journal/sync
  /api/journal/update
  /api/journal/export
  /api/journal/export/excel
  /api/trades
"""
import csv, logging
from flask import Blueprint, request, jsonify, Response
from datetime import datetime as dt

journal_bp = Blueprint("journal_bp", __name__)


def _get_app():
    import app_Stock_Trade as _app
    return _app


# ── journal clear ─────────────────────────────────────────────────────────────

@journal_bp.route("/api/journal/clear", methods=["POST"])
def api_journal_clear():
    import os
    _app = _get_app()
    try:
        if os.path.exists(_app.JOURNAL_FILE):
            open(_app.JOURNAL_FILE, "w").close()
    except Exception:
        pass
    with _app.data_lock:
        _app.cached_data["journal"] = []
        _app.cached_data["stats"] = _app.compute_stats(_app.cached_data.get("positions", {}), [])
    return jsonify({"ok": True})


# ── journal CRUD ──────────────────────────────────────────────────────────────

@journal_bp.route("/api/journal/get", methods=["GET"])
def api_journal_get():
    try:
        from daily_trade_journal import load_journal_entries
        return jsonify(load_journal_entries())
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@journal_bp.route("/api/journal/analytics", methods=["GET"])
def api_journal_analytics():
    try:
        from daily_trade_journal import get_trade_journal_analytics
        return jsonify({"ok": True, "data": get_trade_journal_analytics()})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@journal_bp.route("/api/journal/sync", methods=["POST"])
def api_journal_sync():
    _app = _get_app()
    try:
        from daily_trade_journal import generate_daily_journal
        req = request.json or {}
        dt_str = req.get("date")
        entries = generate_daily_journal(dt_str, kite=_app._kite_session)
        return jsonify({"ok": True, "count": len(entries), "entries": entries})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@journal_bp.route("/api/journal/update", methods=["POST"])
def api_journal_update():
    import trade_db
    try:
        from daily_trade_journal import load_journal_entries, save_journal_entries
        data = request.json or {}
        symbol = data.get("symbol")
        date_str = data.get("date")
        trade_id = data.get("trade_id")
        remarks = data.get("remarks")
        lesson = data.get("lesson")
        if trade_id and lesson is not None:
            trade_db.update_self_learning_lesson(trade_id, lesson)
        if not symbol or not date_str:
            return jsonify({"ok": True, "message": "Updated trade_db lesson"})
        entries = load_journal_entries()
        updated = False
        for e in entries:
            if e.get("Date") == date_str and (e.get("Symbol") == symbol or symbol in e.get("Symbol", "")):
                if remarks is not None: e["Analysis_Remarks"] = remarks
                if lesson is not None: e["Self_Learning_Lesson"] = lesson
                updated = True
        if updated:
            save_journal_entries(entries)
            return jsonify({"ok": True})
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── journal export (CSV) ──────────────────────────────────────────────────────

@journal_bp.route("/api/journal/export", methods=["GET", "POST"])
def api_journal_export():
    try:
        import io
        from daily_trade_journal import load_journal_entries
        entries = load_journal_entries()
        headers = [
            "Date", "Engine", "Symbol", "Side", "Timeframe", "Pattern", "Tier", "Swing_Waves",
            "Entry_Time", "Entry_Price", "Exit_Time", "Exit_Price",
            "SL", "T1", "T2", "T3", "Quantity", "Lot_Size",
            "PnL_Rs", "PnL_Pct", "Outcome", "Analysis_Remarks", "Self_Learning_Lesson"
        ]
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for e in entries:
            writer.writerow(e)
        csv_data = output.getvalue()
        fname = f"trade_journal_export_{dt.now().strftime('%Y_%m_%d_%H%M')}.csv"
        return Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={fname}"}
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── journal export (Excel) ────────────────────────────────────────────────────

@journal_bp.route("/api/journal/export/excel", methods=["GET", "POST"])
def api_journal_export_excel():
    try:
        import openpyxl, io
        from openpyxl.styles import Font, PatternFill, Alignment
        from daily_trade_journal import load_journal_entries

        entries = load_journal_entries()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trade Journal"

        headers = [
            "Date", "Engine", "Symbol", "Side", "Timeframe", "Pattern", "Tier", "Swing Waves",
            "Entry Time", "Entry Price", "Exit Time", "Exit Price",
            "SL", "T1", "T2", "T3", "Quantity", "Lot Size",
            "PnL (₹)", "PnL (%)", "Outcome", "Analysis Remarks", "Self-Learning Lesson"
        ]

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for e in entries:
            ws.append([
                e.get("Date", ""),
                e.get("Engine", ""),
                e.get("Symbol", ""),
                e.get("Side", ""),
                e.get("Timeframe", ""),
                e.get("Pattern", ""),
                e.get("Entry_Time", ""),
                e.get("Entry_Price", ""),
                e.get("Exit_Time", ""),
                e.get("Exit_Price", ""),
                e.get("SL", ""),
                e.get("T1", ""),
                e.get("T2", ""),
                e.get("T3", ""),
                e.get("Quantity", ""),
                e.get("Lot_Size", ""),
                e.get("PnL_Rs", 0),
                e.get("PnL_Pct", ""),
                e.get("Outcome", ""),
                e.get("Analysis_Remarks", ""),
                e.get("Self_Learning_Lesson", "")
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        ws2 = wb.create_sheet(title="Performance Summary")
        ws2.append(["Metric", "Value"])
        ws2.cell(row=1, column=1).font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=2).font = header_font
        ws2.cell(row=1, column=2).fill = header_fill

        total_trades = len(entries)
        wins = sum(1 for e in entries if float(e.get("PnL_Rs") or 0) > 0)
        losses = sum(1 for e in entries if float(e.get("PnL_Rs") or 0) < 0)
        win_rate = (wins / total_trades * 100) if total_trades > 0 else 0
        total_pnl = sum(float(e.get("PnL_Rs") or 0) for e in entries)

        ws2.append(["Total Trades", total_trades])
        ws2.append(["Winning Trades", wins])
        ws2.append(["Losing Trades", losses])
        ws2.append(["Win Rate (%)", f"{win_rate:.1f}%"])
        ws2.append(["Total Net PnL (₹)", f"₹{total_pnl:.2f}"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        fname = f"trade_journal_export_{dt.now().strftime('%Y_%m_%d_%H%M')}.xlsx"
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname}"}
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── trade DB read ─────────────────────────────────────────────────────────────

@journal_bp.route("/api/trades")
def api_trades():
    import trade_db
    engine = request.args.get("engine")
    active_only = request.args.get("active", "false").lower() == "true"
    if active_only:
        return jsonify(trade_db.get_active_trades(engine))
    return jsonify(trade_db.get_all_trades(engine))
