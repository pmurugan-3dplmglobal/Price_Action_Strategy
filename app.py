import os, json, csv, time, threading, subprocess, sys, signal, logging, glob, re
from datetime import datetime as dt
from flask import Flask, render_template_string, jsonify, request
from kiteconnect import KiteConnect
import shared.trade_db as trade_db
from shared.pid_util import is_pid_alive, remove_pid_file, get_running_engines

app = Flask(__name__)

# ──────────────────────────────────────────────
#  FILE PATHS & DASHBOARD CONFIG
# ──────────────────────────────────────────────


def get_kite_credentials():
    """Read Kite API key/secret from config (moved out of source)."""
    cfg = load_config()
    api_key = cfg.get("api_key", "")
    api_secret = cfg.get("api_secret", "")
    if not api_key or not api_secret:
        logging.warning("api_key/api_secret missing in program_config.json")
    return api_key, api_secret


TOKEN_FILE = "input/kite_access_token.txt"
CONFIG_FILE = "input/program_config.json"
STATE_FILE = "output/monitor/stock_positions_state.json"
JOURNAL_FILE = "output/monitor/trade_journal.csv"
INDEX_LOG_FILE = "output/logs/bull_index_trade_engine.log"
NIFTY50_LOG_FILE = "output/logs/bull_nifty50_scanner.log"
DAILY_LOG_FILE = "output/logs/bull_daily_scanner.log"
BEAR_INDEX_LOG_FILE = "output/logs/bear_index_trade_engine.log"
BEAR_NIFTY50_LOG_FILE = "output/logs/bear_nifty50_scanner.log"
BEAR_DAILY_LOG_FILE = "output/logs/bear_daily_scanner.log"

DASHBOARD_PORT = 5051
REFRESH_SECONDS = 5

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

PROGRAMS = {
    "index": {
        "name": "Index Trade Engine (Nifty & BankNifty)",
        "file": "live-trade/bull_index_engine.py",
        "desc": "Real-time index options intraday trading (3min) - Option Premium Scanning",
        "color": "#58a6ff",
        "log_file": INDEX_LOG_FILE,
        "config_fields": {
            "timeframe": {"label": "Entry Timeframe", "type": "select", "options": ["minute","3minute","5minute","10minute","15minute","30minute","60minute"], "default": "3minute"},
            "timeframe_anchor": {"label": "Anchor Timeframe", "type": "select", "options": ["10minute","15minute","30minute","60minute"], "default": "15minute"},
            "lookback_days": {"label": "Lookback Days", "type": "number", "default": 30},
            "scan_interval": {"label": "Scan Interval (s)", "type": "number", "default": 15},
            "risk_percent": {"label": "Risk %", "type": "number", "default": 1.0},
            "capital": {"label": "Capital", "type": "number", "default": 100000.0},
            "strike_range": {"label": "Strike Range (±)", "type": "number", "default": 3}
        }
    },
    "nifty50": {
        "name": "Nifty 50 Stock Scanner + Executor",
        "file": "live-trade/bull_nifty50_scanner.py",
        "desc": "Scans Nifty 50 stocks, picks best setup, executes (15min) - Option Premium Scanning",
        "color": "#3fb950",
        "log_file": NIFTY50_LOG_FILE,
        "config_fields": {
            "timeframe": {"label": "Entry Timeframe", "type": "select", "options": ["5minute","10minute","15minute","30minute","60minute"], "default": "15minute"},
            "timeframe_anchor": {"label": "Anchor Timeframe", "type": "select", "options": ["30minute","60minute","day"], "default": "60minute"},
            "lookback_days": {"label": "Lookback Days", "type": "number", "default": 30},
            "scan_interval": {"label": "Scan Interval (s)", "type": "number", "default": 300},
            "risk_percent": {"label": "Risk %", "type": "number", "default": 1.0},
            "capital": {"label": "Capital", "type": "number", "default": 100000.0}
        }
    },
    "daily": {
        "name": "Nifty 50 Daily Scanner (Export)",
        "file": "live-trade/bull_nifty50_daily_scanner.py",
        "desc": "Scans Nifty 50 on daily timeframe, exports to Excel",
        "color": "#d29922",
        "log_file": DAILY_LOG_FILE,
        "config_fields": {
            "lookback_days": {"label": "Lookback Days", "type": "number", "default": 120}
        }
    },
    "bear_index": {
        "name": "Bear Index Trade Engine (Nifty & BankNifty)",
        "file": "live-trade/bear_index_engine.py",
        "desc": "Bearish index options intraday trading (3min) - Option Premium Scanning",
        "color": "#f85149",
        "log_file": BEAR_INDEX_LOG_FILE,
        "config_fields": {
            "timeframe": {"label": "Entry Timeframe", "type": "select", "options": ["minute","3minute","5minute","10minute","15minute","30minute","60minute"], "default": "3minute"},
            "timeframe_anchor": {"label": "Anchor Timeframe", "type": "select", "options": ["10minute","15minute","30minute","60minute"], "default": "15minute"},
            "lookback_days": {"label": "Lookback Days", "type": "number", "default": 30},
            "scan_interval": {"label": "Scan Interval (s)", "type": "number", "default": 15},
            "risk_percent": {"label": "Risk %", "type": "number", "default": 1.0},
            "capital": {"label": "Capital", "type": "number", "default": 100000.0},
            "strike_range": {"label": "Strike Range (±)", "type": "number", "default": 3}
        }
    },
    "bear_nifty50": {
        "name": "Bear Nifty 50 Stock Scanner + Executor",
        "file": "live-trade/bear_nifty50_scanner.py",
        "desc": "Bearish Nifty 50 stock scanning & execution via PE (15min) - Option Premium Scanning",
        "color": "#da3633",
        "log_file": BEAR_NIFTY50_LOG_FILE,
        "config_fields": {
            "timeframe": {"label": "Entry Timeframe", "type": "select", "options": ["5minute","10minute","15minute","30minute","60minute"], "default": "15minute"},
            "timeframe_anchor": {"label": "Anchor Timeframe", "type": "select", "options": ["30minute","60minute","day"], "default": "60minute"},
            "lookback_days": {"label": "Lookback Days", "type": "number", "default": 30},
            "scan_interval": {"label": "Scan Interval (s)", "type": "number", "default": 300},
            "risk_percent": {"label": "Risk %", "type": "number", "default": 1.0},
            "capital": {"label": "Capital", "type": "number", "default": 100000.0}
        }
    },
    "bear_daily": {
        "name": "Bear Nifty 50 Daily Scanner (Export)",
        "file": "live-trade/bear_nifty50_daily_scanner.py",
        "desc": "Bearish daily scan on Nifty 50, exports to Excel",
        "color": "#ff7b72",
        "log_file": BEAR_DAILY_LOG_FILE,
        "config_fields": {
            "lookback_days": {"label": "Lookback Days", "type": "number", "default": 120}
        }
    }
}

processes = {}
process_lock = threading.Lock()

data_lock = threading.Lock()
cached_data = {
    "positions": {},
    "journal": [],
    "log_tail": {pid: [] for pid in PROGRAMS},
    "stats": {"total_trades": 0, "win_rate": 0, "active_positions": 0, "pnl": 0},
    "scans": {"index": [], "nifty50": [], "daily": [], "bear_index": [], "bear_nifty50": [], "bear_daily": []},
    "scan_summary": {"index": {"anchors": {}, "abc_matches": {}}, "nifty50": {"anchors": {}, "abc_matches": {}}, "daily": {"anchors": {}, "abc_matches": {}}, "bear_index": {"anchors": {}, "abc_matches": {}}, "bear_nifty50": {"anchors": {}, "abc_matches": {}}, "bear_daily": {"anchors": {}, "abc_matches": {}}},
    "all_trades": [],
    "kite_positions": [],
    "ltp": {},
    "anchor_status": {"running": False, "engine": None, "requested_at": None, "completed_at": None},
    "pending_trades": [],
    "backtest_results": {}
}
_ltp_last_fetch = 0
_kite_positions_last_fetch = 0
_kite_session = None

# ──────────────────────────────────────────────
#  PROCESS MANAGEMENT (Start/Stop Programs)
# ──────────────────────────────────────────────

def get_pid_for_program(prog_id):
    with process_lock:
        p = processes.get(prog_id)
        if p and p.poll() is None:
            return p.pid
    return None

def start_program(prog_id):
    token = check_token_valid()
    if not token["valid"]:
        print(f"Cannot start {prog_id}: {token['reason']}")
        return False
    with process_lock:
        if prog_id in processes and processes[prog_id].poll() is None:
            return False
        if is_pid_alive(prog_id):
            print(f"{prog_id} already running (PID file exists)")
            return False
        script_path = os.path.join(BASE_DIR, PROGRAMS[prog_id]["file"])
        try:
            p = subprocess.Popen(
                [sys.executable, script_path],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                cwd=BASE_DIR, creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
            )
            processes[prog_id] = p
            return True
        except Exception as e:
            print(f"Failed to start {prog_id}: {e}")
            return False

def stop_program(prog_id):
    with process_lock:
        p = processes.get(prog_id)
        if p and p.poll() is None:
            pid = p.pid
            if os.name == "nt":
                subprocess.run(["taskkill", "/PID", str(pid)], capture_output=True)
                time.sleep(2)
                if p.poll() is None:
                    subprocess.run(["taskkill", "/F", "/T", "/PID", str(pid)], capture_output=True)
            else:
                os.kill(pid, signal.SIGTERM)
                time.sleep(2)
                if p.poll() is None:
                    os.kill(pid, signal.SIGKILL)
            processes.pop(prog_id, None)
            remove_pid_file(prog_id)
            return True
    return False

# ──────────────────────────────────────────────
#  CONFIGURATION (program_config.json)
# ──────────────────────────────────────────────

def load_config():
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE) as f:
                return json.load(f)
    except:
        pass
    return {}

def save_config(prog_id, data):
    cfg = load_config()
    cleaned = {}
    for k, v in data.items():
        if isinstance(v, str):
            try:
                if "." in v:
                    cleaned[k] = float(v)
                else:
                    cleaned[k] = int(v)
            except (ValueError, TypeError):
                cleaned[k] = v
        else:
            cleaned[k] = v
    cfg[prog_id] = cleaned
    os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)
    return True

def get_backtest_mode():
    return load_config().get("_backtest", False)

def set_backtest_mode(enabled):
    cfg = load_config()
    cfg["_backtest"] = enabled
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

# ──────────────────────────────────────────────
#  KITE TOKEN MANAGEMENT
# ──────────────────────────────────────────────

def check_token_valid():
    if not os.path.exists(TOKEN_FILE):
        return {"valid": False, "reason": "Token file not found"}
    try:
        with open(TOKEN_FILE) as f:
            data = json.load(f)
        if not data.get("api_key") or not data.get("access_token"):
            return {"valid": False, "reason": "Invalid token file"}
        date_str = data.get("generated_at", "")
        if date_str:
            try:
                from datetime import datetime as dt2
                gen_date = dt2.strptime(date_str.split()[0], "%Y-%m-%d").date()
                today = dt.now().date()
                if gen_date < today:
                    return {"valid": False, "reason": f"Token expired (generated {date_str})"}
            except Exception:
                pass
        return {"valid": True, "reason": "Token valid"}
    except Exception as e:
        return {"valid": False, "reason": f"Token read error: {e}"}

def get_login_url():
    api_key, _ = get_kite_credentials()
    kite = KiteConnect(api_key=api_key)
    return f"https://kite.zerodha.com/connect/login?api_key={api_key}"

def exchange_request_token(request_token):
    """Exchange Kite request_token for access_token and save to file."""
    try:
        api_key, api_secret = get_kite_credentials()
        kite = KiteConnect(api_key=api_key)
        session = kite.generate_session(request_token, api_secret=api_secret)
        access_token = session["access_token"]
        token_data = {
            "api_key": api_key,
            "access_token": access_token,
            "generated_at": dt.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        os.makedirs(os.path.dirname(TOKEN_FILE), exist_ok=True)
        with open(TOKEN_FILE, "w") as f:
            json.dump(token_data, f, indent=4)
        return {"ok": True, "access_token": access_token[:8] + "..."}
    except Exception as e:
        return {"ok": False, "error": str(e)}

# ──────────────────────────────────────────────
#  DATA LOADING (trade_db, journal, logs)
# ──────────────────────────────────────────────

def load_positions():
    try:
        active = trade_db.get_active_trades()
        return {t["symbol"]: t for t in active}
    except:
        return {}

def load_journal():
    rows = []
    if os.path.exists(JOURNAL_FILE):
        try:
            with open(JOURNAL_FILE, newline="", encoding="utf-8") as f:
                sample = f.read(4096)
                f.seek(0)
                dialect = csv.Sniffer().sniff(sample, delimiters="\t,") if sample.strip() else None
                reader = csv.DictReader(f, delimiter=(dialect.delimiter if dialect else "\t"))
                for row in reader:
                    rows.append(row)
        except Exception:
            pass
    return rows[-200:]

def tail_log(filepath, n=200):
    if not os.path.exists(filepath):
        return []
    try:
        with open(filepath, encoding="utf-8") as f:
            lines = f.readlines()
        return lines[-n:]
    except:
        return []

def compute_stats(positions, journal):
    active = len(positions)
    total = len(journal)
    wins = sum(1 for j in journal if j.get("P&L %", "").replace("%", "").replace("-", "").strip()
               and j.get("Action", "").startswith("EXIT_"))
    win_rate = round((wins / total) * 100, 1) if total > 0 else 0
    pnl = 0.0
    for j in journal:
        try:
            pnl_str = j.get("P&L %", "").replace("%", "")
            if pnl_str and pnl_str != "-":
                pnl += float(pnl_str)
        except Exception:
            pass
    return {"total_trades": total, "win_rate": win_rate, "active_positions": active, "pnl": round(pnl, 2)}

SCAN_SYMBOLS = [
    "RELIANCE","TCS","HDFCBANK","ICICIBANK","INFY","ITC","SBIN","BHARTIARTL","LT","WIPRO",
    "ADANIENT","ADANIPORTS","APOLLOHOSP","ASIANPAINT","AXISBANK","BAJAJ-AUTO","BAJAJFINSV",
    "BAJFINANCE","BEL","CIPLA","COALINDIA","DRREDDY","EICHERMOT","GRASIM","HCLTECH",
    "HDFCLIFE","HEROMOTOCO","HINDALCO","HINDUNILVR","INDIGO","JIOFIN","JSWSTEEL",
    "KOTAKBANK","M&M","MARUTI","NESTLEIND","NTPC","ONGC","POWERGRID","SBILIFE",
    "SHRIRAMFIN","SUNPHARMA","TATACONSUM","TATASTEEL","TECHM","TITAN","TRENT","ULTRACEMCO",
    "NIFTY","BANKNIFTY"
]

# ──────────────────────────────────────────────
#  SCAN PARSING — Split anchor vs ABC per symbol
# ──────────────────────────────────────────────

def _extract_symbol(line):
    for sym in SCAN_SYMBOLS:
        if sym in line:
            return sym
    return None

def parse_scans_for_program(log_lines, prog_id):
    matches = []
    anchors = {}
    abc_matches = {}
    for line in log_lines:
        clean = line.strip()
        if "ANCHOR" in clean:
            matches.append(clean)
            sym = _extract_symbol(clean)
            if sym:
                anchors[sym] = clean
        elif "MATCH" in clean or "BEST TRADE" in clean or "Match" in clean:
            matches.append(clean)
            sym = _extract_symbol(clean)
            if sym:
                abc_matches[sym] = clean
    return matches, anchors, abc_matches

# ──────────────────────────────────────────────
#  BACKGROUND DATA REFRESH THREAD
# ──────────────────────────────────────────────

def refresh_data():
    global cached_data, _ltp_last_fetch, _kite_positions_last_fetch, _kite_session
    while True:
        with data_lock:
            pos = load_positions()
            journal = load_journal()
            cached_data["positions"] = pos
            cached_data["journal"] = journal
            cached_data["stats"] = compute_stats(pos, journal)
            try:
                all_t = trade_db.get_all_trades()
                cached_data["all_trades"] = all_t
            except Exception:
                cached_data["all_trades"] = []
            for pid in PROGRAMS:
                log_file = PROGRAMS[pid].get("log_file")
                log_lines = tail_log(log_file) if log_file else []
                cached_data["log_tail"][pid] = log_lines
                scan_lines, anchors, abc_matches = parse_scans_for_program(log_lines, pid)
                cached_data["scans"][pid] = scan_lines
                cached_data["scan_summary"][pid] = {"anchors": anchors, "abc_matches": abc_matches}

            # Build pending_trades from scan lines
            import re
            pending = []
            seen_keys = set()
            for pid in PROGRAMS:
                for line in cached_data["scans"].get(pid, []):
                    m = re.search(r'CYCLE MATCH staged:\s*(\S+)\s*\|\s*(\S+)\s*\|\s*(\S+)\s*@\s*(\d+)\s*\|\s*RR=([\d.]+)\s*\|\s*Entry:\s*([\d.]+)', line)
                    if not m:
                        m = re.search(r'CYCLE MATCH staged:\s*(\S+)\s*\|\s*(\S+)\s*\|\s*(\S+)\s*\|\s*Strike\s+(\d+)', line)
                    if not m:
                        continue
                    sym = m.group(1)
                    # Determine base symbol (strip contract suffix for options)
                    opt_match = re.match(r'^([A-Z]+)', sym)
                    base = opt_match.group(1) if opt_match else sym
                    key = f"{base}|{m.group(4) if m.lastindex >= 4 else ''}"
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)
                    ext = {}
                    if 'RR=' in line:
                        rr_m = re.search(r'RR=([\d.]+)', line)
                        if rr_m: ext['rr'] = float(rr_m.group(1))
                    if 'Entry:' in line:
                        e_m = re.search(r'Entry:\s*([\d.]+)', line)
                        if e_m: ext['entry'] = float(e_m.group(1))
                    if 'SL:' in line:
                        s_m = re.search(r'SL:\s*([\d.]+)', line)
                        if s_m: ext['sl'] = float(s_m.group(1))
                    for t in ['T1', 'T2', 'T3']:
                        tm = re.search(rf'{t}:\s*([\d.]+)', line)
                        if tm: ext[t.lower()] = float(tm.group(1))
                    strike_m = re.search(r'@\s*(\d+)', line)
                    if not strike_m:
                        strike_m = re.search(r'Strike\s+(\d+)', line)
                    date_m = re.search(r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})', line)
                    trade_date = date_m.group(1) if date_m else ''
                    exec_key = f"{base}|{m.group(2)}|{m.group(3)}|{strike_m.group(1) if strike_m else 0}"
                    executed = any(
                        trade_db.is_pattern_executed(e, exec_key)
                        for e in list(PROGRAMS.keys()) + ["manual"]
                    )
                    pending.append({
                        "symbol": base, "engine": pid,
                        "pattern": m.group(2), "side": m.group(3),
                        "strike": int(strike_m.group(1)) if strike_m else 0,
                        "rr": ext.get('rr', 0),
                        "entry": ext.get('entry', 0),
                        "sl": ext.get('sl', 0),
                        "t1": ext.get('t1', 0),
                        "t2": ext.get('t2', 0),
                        "t3": ext.get('t3', 0),
                        "executed": executed,
                        "date": trade_date,
                    })
            cached_data["pending_trades"] = pending
            # Load backtest results
            bt = {}
            for engine in ("index", "nifty50", "bear_index", "bear_nifty50"):
                p = f"output/monitor/backtest_results_{engine}.json"
                if os.path.exists(p):
                    try:
                        with open(p) as f:
                            bt[engine] = json.load(f)
                    except Exception:
                        pass
            cached_data["backtest_results"] = bt
            # Load scanner display data (staged trades + active positions)
            try:
                _dp = os.path.join("output", "monitor", "scan_display_data.json")
                if os.path.exists(_dp):
                    with open(_dp) as f:
                        cached_data["scan_display_data"] = json.load(f)
                else:
                    cached_data["scan_display_data"] = {}
            except Exception:
                cached_data["scan_display_data"] = {}
            cached_data["live_execution"] = {
                "nifty50": os.path.exists(os.path.join("input", "nifty50_live.flag"))
            }
            now = time.time()
            if now - _ltp_last_fetch > 30 and cached_data["all_trades"]:
                _ltp_last_fetch = now
                try:
                    active = trade_db.get_active_trades()
                    if active:
                        if not _kite_session:
                            tk = check_token_valid()
                            if tk["valid"]:
                                td = json.load(open(TOKEN_FILE))
                                api_key, _ = get_kite_credentials()
                                ks = KiteConnect(api_key=api_key)
                                ks.set_access_token(td["access_token"])
                                _kite_session = ks
                        if _kite_session:
                            syms = []
                            for t in active:
                                tok = t.get("option_token") or t.get("index_token")
                                if tok: syms.append(f"NFO:{tok}")
                            if syms:
                                quotes = _kite_session.quote(syms)
                                ltp = {}
                                for key, q in quotes.items():
                                    ltp[key.split(":")[-1]] = q.get("last_price", 0)
                                cached_data["ltp"] = ltp
                except Exception:
                    _kite_session = None
        if now - _kite_positions_last_fetch > 60:
            _kite_positions_last_fetch = now
            try:
                if not _kite_session:
                    tk = check_token_valid()
                    if tk["valid"]:
                        td = json.load(open(TOKEN_FILE))
                        api_key, _ = get_kite_credentials()
                        ks = KiteConnect(api_key=api_key)
                        ks.set_access_token(td["access_token"])
                        _kite_session = ks
                if _kite_session:
                    kite_positions = _kite_session.positions()
                    merged = []
                    seen_symbols = set()
                    for plist in [kite_positions.get("day", []), kite_positions.get("net", [])]:
                        for p in plist:
                            sym = p.get("tradingsymbol", "")
                            if not sym or sym in seen_symbols:
                                continue
                            seen_symbols.add(sym)
                            qty = abs(int(p.get("net_quantity", 0)))
                            if qty == 0:
                                continue
                            merged.append({
                                "contract": sym,
                                "quantity": qty,
                                "entry_price": float(p.get("net_price", 0)),
                                "pnl": float(p.get("pnl", 0)),
                                "exchange": p.get("exchange", ""),
                                "source": "kite"
                            })
                    cached_data["kite_positions"] = merged
            except Exception:
                pass
        if int(time.time()) % 3600 < REFRESH_SECONDS:
            auto_export_if_new_month()
        time.sleep(REFRESH_SECONDS)

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Trading -02-Underlying Stock bases</title>
    <script>
        // ── Filter State ──
        let journalFilter = 'all';
        let scanFilter = 'all';
        let logFilter = 'all';
        let positionFilter = 'active';

        // ── Tab Switching ──
        function switchLeftTab(tabId) {
            document.querySelectorAll('.left-tab-content').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('.left-tab-btn').forEach(b => b.classList.remove('active'));
            document.getElementById(tabId).classList.add('active');
            event.target.classList.add('active');
            if (tabId === 'backtest-tab') renderBacktest();
        }

        function switchTab(tabId) {
            document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
            document.getElementById(tabId).classList.add('active');
            event.target.classList.add('active');
        }

        // ── Filter Controls ──
        function setFilter(type, value) {
            if (type === 'journal') journalFilter = value;
            else if (type === 'scan') scanFilter = value;
            else if (type === 'log') logFilter = value;
            else if (type === 'position') { positionFilter = value; document.querySelectorAll('.pos-filter-btn').forEach(b => b.classList.remove('active')); event.target.classList.add('active'); }
            renderReport();
        }

        // ── Live Execution Toggle ──
        async function setLiveExec(engine, on) {
            try {
                const r = await fetch('/api/live-execution/' + engine, {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({enabled: on})
                });
                const j = await r.json();
                const lbl = document.getElementById(engine + '-live-label');
                if (lbl) {
                    lbl.textContent = j.enabled ? 'LIVE EXECUTION ON' : 'SCAN-ONLY';
                    lbl.classList.toggle('on', !!j.enabled);
                }
                if (!j.ok) console.log('live-exec toggle failed', j);
            } catch(e) { console.log(e); }
        }

        // ── Backtest Controls ──
        async function toggleBacktestMode(enabled) {
            try {
                await fetch('/api/backtest/mode', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({enabled: enabled})
                });
                refreshBacktestMode();
            } catch(e) { console.log(e); }
        }

        async function toggleNopa(disabled) {
            try {
                await fetch('/api/scanner/config', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({disable_nopa: disabled})
                });
            } catch(e) { console.log(e); }
        }

        // ── Program Start/Stop ──
        async function toggleProgram(progId, action) {
            if (action === 'stop') {
                try { await fetch('/api/anchor/stop', {method: 'POST'}); } catch(e) {}
            }
            const btn = document.querySelector(`.start-btn[onclick*="'${progId}'"]`);
            const fb = document.getElementById(`cfg-fb-${progId}`);
            if (action === 'start' && btn) btn.disabled = true;
            try {
                const r = await fetch(`/api/programs/${progId}/${action}`, {method: 'POST'});
                const d = await r.json();
                if (d.ok) {
                    if (fb) { fb.textContent = action === 'start' ? 'Running!' : 'Stopped'; fb.style.color = '#3fb950'; setTimeout(() => { fb.textContent = ''; }, 2000); }
                    setTimeout(refreshData, 500);
                }
                if (d.error) {
                    if (fb) { fb.textContent = d.error; fb.style.color = '#f85149'; setTimeout(() => { fb.textContent = ''; }, 4000); }
                    if (action === 'start' && btn) btn.disabled = false;
                }
            } catch(e) { console.log(e); if (action === 'start' && btn) btn.disabled = false; }
        }

        function toggleConfig(headerEl) {
            const body = headerEl.parentElement.querySelector('.config-body');
            const arrow = headerEl.querySelector('.config-arrow');
            if (body.style.display === 'block') {
                body.style.display = 'none';
                arrow.textContent = '\u25B6';
            } else {
                body.style.display = 'block';
                arrow.textContent = '\u25BC';
            }
        }

        async function saveConfig(progId) {
            const inputs = document.querySelectorAll(`.config-input[data-prog="${progId}"]`);
            const data = {};
            inputs.forEach(inp => {
                data[inp.getAttribute('data-field')] = inp.value;
            });
            try {
                const r = await fetch(`/api/config/${progId}`, {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(data)
                });
                const d = await r.json();
                const fb = document.getElementById(`cfg-fb-${progId}`);
                if (d.ok) {
                    fb.textContent = 'Saved!';
                    fb.style.color = '#3fb950';
                } else {
                    fb.textContent = 'Failed';
                    fb.style.color = '#f85149';
                }
                setTimeout(() => { fb.textContent = ''; }, 2000);
            } catch(e) { console.log(e); }
        }

        function sortTable(th, colIdx) {
            const table = th.closest('table');
            const tbody = table.querySelector('tbody');
            const rows = Array.from(tbody.querySelectorAll('tr'));
            const dir = th.getAttribute('data-dir') === 'asc' ? 'desc' : 'asc';
            th.setAttribute('data-dir', dir);
            rows.sort((a, b) => {
                let aVal = a.cells[colIdx].innerText.trim();
                let bVal = b.cells[colIdx].innerText.trim();
                let aNum = parseFloat(aVal.replace('%', '').replace(',', ''));
                let bNum = parseFloat(bVal.replace('%', '').replace(',', ''));
                if (!isNaN(aNum) && !isNaN(bNum)) {
                    return dir === 'asc' ? aNum - bNum : bNum - aNum;
                }
                return dir === 'asc' ? aVal.localeCompare(bVal) : bVal.localeCompare(aVal);
            });
            rows.forEach(r => tbody.appendChild(r));
        }

        // ── Main Dashboard Render ──
        function renderReport() {
            const d = window._lastData;
            if (!d) return;
            const stats = d.stats || {};
            const positions = d.positions || {};
            const kitePos = d.kite_positions || [];
            const journal = d.journal || [];

            const actPos = kitePos.length || Object.keys(positions).length;
            document.getElementById('stat-active').textContent = actPos;
            document.getElementById('stat-total').textContent = stats.total_trades || 0;

            let total = stats.total_trades || 0;
            let wins = 0;
            (journal || []).forEach(j => {
                const pnl = (j['P&L %'] || '').replace('%', '').replace('-', '').trim();
                if (pnl && (j.Action || '').startsWith('EXIT_')) wins++;
            });
            let wr = total > 0 ? ((wins/total)*100).toFixed(1) : 0;
            let wrEl = document.getElementById('stat-winrate');
            wrEl.textContent = wr + '%';
            wrEl.style.color = wr >= 50 ? '#3fb950' : '#f85149';

            let pnl = stats.pnl || 0;
            let pnlEl = document.getElementById('stat-pnl');
            pnlEl.textContent = pnl + '%';
            pnlEl.style.color = pnl >= 0 ? '#3fb950' : '#f85149';

            let posHtml = '';
            let allTrades = d.all_trades || [];
            let ltpData = d.ltp || {};
            let mergedPositions = [];
            let seenContracts = new Set();
            kitePos.forEach(kp => {
                seenContracts.add(kp.contract);
                mergedPositions.push({
                    symbol: kp.contract,
                    engine: kp.exchange === 'NFO' ? 'Index' : 'Nifty 50',
                    pattern: 'KITE_OPEN',
                    entry_spot: kp.entry_price,
                    quantity: kp.quantity,
                    pnl: kp.pnl,
                    status: 'ACTIVE',
                    source: 'kite'
                });
            });
            // In live mode show only Kite positions; in backtest mode include DB trades
            if (d.config && d.config._backtest === true) {
                allTrades.forEach(t => {
                    const contract = t.contract || t.symbol || '';
                    if (seenContracts.has(contract)) return;
                    const st = (t.status || '').toLowerCase();
                    if (positionFilter === 'active' && st !== 'active') return;
                    if (positionFilter === 'completed' && st !== 'sl_hit' && st !== 'target_hit') return;
                    if (positionFilter === 'sl_hit' && st !== 'sl_hit') return;
                    mergedPositions.push({
                        symbol: t.symbol || contract,
                        engine: (t.engine === 'index' || (t.symbol||'').includes('NIFTY') || (t.symbol||'').includes('BANK')) ? 'Index' : 'Nifty 50',
                        pattern: t.pattern || '',
                        entry_spot: t.entry_spot || '',
                        current_sl: t.current_sl || '',
                        t1: t.t1 || '',
                        t2: t.t2 || '',
                        t3: t.t3 || '',
                        status: t.status || 'ACTIVE',
                        created_at: t.created_at || '',
                        exit_time: t.exit_time || '',
                        pnl_percent: t.pnl_percent,
                        token: t.option_token || t.index_token || '',
                        source: 'db'
                    });
                });
            }
            if (mergedPositions.length) {
                posHtml = '<table><thead><tr><th onclick="sortTable(this,0)">Symbol</th><th onclick="sortTable(this,1)">Source</th><th onclick="sortTable(this,2)">Pattern</th><th onclick="sortTable(this,3)">Entry</th><th onclick="sortTable(this,4)">SL</th><th onclick="sortTable(this,5)">T1</th><th onclick="sortTable(this,6)">T2</th><th onclick="sortTable(this,7)">T3</th><th onclick="sortTable(this,8)">LTP</th><th onclick="sortTable(this,9)">Qty</th><th onclick="sortTable(this,10)">Status</th><th onclick="sortTable(this,11)">P&L</th><th>Close</th></tr></thead><tbody>';
                mergedPositions.forEach(t => {
                    const st = (t.status || '').toLowerCase();
                    let badge = 'badge-open';
                    let stLabel = t.status || 'ACTIVE';
                    if (st === 'sl_hit') { badge = 'badge-loss'; stLabel = 'SL HIT'; }
                    else if (st === 'target_hit') { badge = 'badge-profit'; stLabel = 'TARGET'; }
                    const pnl = t.pnl_percent !== undefined && t.pnl_percent !== null ? t.pnl_percent : (t.pnl || '');
                    const pnlBadge = pnl !== '' ? (pnl >= 0 ? 'badge-profit' : 'badge-loss') : '';
                    const ltpVal = t.token ? (ltpData[t.token] || '') : '';
                    const qty = t.quantity || '';
                    const slVal = t.current_sl || '';
                    const t1v = t.t1 || '';
                    const t2v = t.t2 || '';
                    const t3v = t.t3 || '';
                    const entryVal = t.entry_spot || '';
                    const isActive = st === 'active';
                    const closeBtn = isActive ? `<button class="trade-btn sell" onclick="closePosition('${t.symbol}','${t.token}','${t.engine}')">Close</button>` : '';
                    posHtml += `<tr><td><strong>${t.symbol}</strong></td><td>${t.source}</td><td><span class="badge badge-open">${t.pattern||''}</span></td><td>${entryVal}</td><td>${slVal}</td><td>${t1v}</td><td>${t2v}</td><td>${t3v}</td><td>${ltpVal}</td><td>${qty}</td><td><span class="badge ${badge}">${stLabel}</span></td><td>${pnl !== '' ? `<span class="badge ${pnlBadge}">${pnl}</span>` : '-'}</td><td>${closeBtn}</td></tr>`;
                });
                posHtml += '</tbody></table>';
            } else {
                posHtml = '<p class="empty-state">No positions match filter</p>';
            }
            document.getElementById('active-positions-body').innerHTML = posHtml;

            let jHtml = '';
            let filteredJournal = journal || [];
            if (journalFilter !== 'all') {
                filteredJournal = filteredJournal.filter(j => {
                    const sym = (j.Symbol || '').toUpperCase();
                    const isIndex = sym.includes('NIFTY') || sym.includes('BANK');
                    return journalFilter === 'index' ? isIndex : !isIndex;
                });
            }
            if (filteredJournal.length) {
                jHtml = '<table><thead><tr><th onclick="sortTable(this,0)">Timestamp</th><th onclick="sortTable(this,1)">Symbol</th><th onclick="sortTable(this,2)">Pattern</th><th onclick="sortTable(this,3)">Action</th><th onclick="sortTable(this,4)">Status</th><th onclick="sortTable(this,5)">Entry</th><th onclick="sortTable(this,6)">SL</th><th onclick="sortTable(this,7)">Target</th><th onclick="sortTable(this,8)">RR</th><th onclick="sortTable(this,9)">P&L %</th></tr></thead><tbody>';
                filteredJournal.forEach(j => {
                    const pnlv = j['P&L %'] || '-';
                    const ts = j.Timestamp ? j.Timestamp.trim() : '';
                    const time = ts ? (ts.includes(' ') ? ts.split(' ')[1] || '' : '') : '';
                    const date = ts ? (ts.includes(' ') ? ts.split(' ')[0] || '' : ts) : '';
                    const act = j.Action || '';
                    const st = j.Status || '';
                    let badge = 'badge-open';
                    if (act.startsWith('EXIT')) badge = 'badge-closed';
                    else if (st === 'FAILED') badge = 'badge-failed';
                    else if (st === 'MUTATED') badge = 'badge-mutated';
                    let pnlBadge = '';
                    if (pnlv !== '-') {
                        const pn = parseFloat(pnlv);
                        pnlBadge = pn >= 0 ? 'badge-profit' : 'badge-loss';
                    }
                    const entryVal = j['Entry'] || '';
                    const slVal = j['SL'] || '';
                    const targetVal = j['Target'] || '';
                    const rrVal = j['RR'] || '';
                    jHtml += `<tr><td style="font-size:11px">${date}<br><span style="color:#8b949e">${time}</span></td><td>${j.Symbol||''}</td><td>${j.Pattern||''}</td><td><span class="badge ${badge}">${act}</span></td><td>${st}</td><td>${entryVal}</td><td>${slVal}</td><td>${targetVal}</td><td>${rrVal}</td><td><span class="badge ${pnlBadge}">${pnlv}</span></td></tr>`;
                });
                jHtml += '</tbody></table>';
            } else {
                jHtml = '<p class="empty-state">No journal entries yet</p>';
            }
            document.getElementById('journal-body').innerHTML = jHtml;

            // ── Live Scanner Display (scan_display_data.json) ──
            const sdd = d.scan_display_data || {};
            const liveOn = (d.live_execution && d.live_execution.nifty50) || false;
            const tog = document.getElementById('nifty50-live-toggle');
            if (tog) tog.checked = liveOn;
            const sLbl = document.getElementById('nifty50-live-label');
            if (sLbl) {
                sLbl.textContent = liveOn ? 'LIVE EXECUTION ON' : 'SCAN-ONLY';
                sLbl.classList.toggle('on', liveOn);
            }

            // Daily reset: clear view if the display date is not today.
            // Use LOCAL date (engine writes local date) to avoid the UTC
            // mismatch that would blank the tab after ~18:30 IST.
            const sddDate = sdd.date || '';
            const todayStr = new Date().toLocaleDateString('en-CA'); // YYYY-MM-DD local
            let stagedList = sdd.staged_trades || [];
            const HIT = ['SL_HIT','TARGET_HIT','T1_HIT','T2_HIT','T3_HIT'];
            let activeList = (sdd.active_positions || []).filter(p => !HIT.includes(p.status));
            if (sddDate && sddDate !== todayStr) { stagedList = []; activeList = []; }

            function tradeRow(t) {
                const num = (v) => (v === null || v === undefined || v === 0) ? '-' : v;
                return '<tr>' +
                    '<td>' + (t.symbol || t.contract || '-') + '</td>' +
                    '<td>' + (t.contract || '-') + '</td>' +
                    '<td>' + (t.side || '-') + '</td>' +
                    '<td>' + num(t.entry) + '</td>' +
                    '<td>' + num(t.sl) + '</td>' +
                    '<td>' + num(t.t1) + '</td>' +
                    '<td>' + num(t.t2) + '</td>' +
                    '<td>' + num(t.t3) + '</td>' +
                    '<td>' + (t.entry_time || '-') + '</td>' +
                    '<td>' + (t.exit_time || '-') + '</td>' +
                    '<td>' + (t.status || 'STAGED') + '</td>' +
                    '<td>' + num(t.pnl_percent) + '</td>' +
                '</tr>';
            }
            const thead = '<table class="scan-table"><thead><tr>' +
                '<th>Symbol</th><th>Contract</th><th>Side</th><th>Entry</th><th>SL</th><th>T1</th><th>T2</th><th>T3</th><th>Entry Time</th><th>Exit Time</th><th>Result</th><th>P&amp;L%</th>' +
                '</tr></thead><tbody>';

            const sb = document.getElementById('scan-staged-body');
            if (sb) sb.innerHTML = stagedList.length
                ? thead + stagedList.map(tradeRow).join('') + '</tbody></table>'
                : '<p class="empty-state">No scan matches yet</p>';

            const ab = document.getElementById('scan-active-body');
            if (ab) ab.innerHTML = activeList.length
                ? thead + activeList.map(tradeRow).join('') + '</tbody></table>'
                : '<p class="empty-state">No active positions</p>';

            let logHtml = '';
            let allLogs = [];
            if (logFilter === 'all' || logFilter === 'index') {
                allLogs = allLogs.concat(d.programs?.index?.log_tail || []);
            }
            if (logFilter === 'all' || logFilter === 'bear_index') {
                allLogs = allLogs.concat(d.programs?.bear_index?.log_tail || []);
            }
            if (logFilter === 'all' || logFilter === 'nifty50') {
                allLogs = allLogs.concat(d.programs?.nifty50?.log_tail || []);
            }
            if (logFilter === 'all' || logFilter === 'bear_nifty50') {
                allLogs = allLogs.concat(d.programs?.bear_nifty50?.log_tail || []);
            }
            allLogs.forEach(l => { logHtml += '<div>'+l.replace(/</g,'&lt;').replace(/>/g,'&gt;')+'</div>'; });
            document.getElementById('log-body').innerHTML = logHtml;
            const logBox = document.getElementById('log-body');
            logBox.scrollTop = logBox.scrollHeight;

            document.getElementById('last-updated').textContent = 'Last refreshed: ' + new Date().toLocaleString();
            renderPendingTrades();
            renderBacktestSummary();
        }

        function renderBacktestSummary() {
            const d = window._lastData;
            if (!d) return;
            const results = d.backtest_results || {};
            let html = '';
            Object.entries(results).forEach(([engine, r]) => {
                if (!r.total_trades) return;
                const wr = r.win_rate || 0;
                html += `<div class="backtest-summary">
                    <strong>${r.engine || engine}</strong> &mdash; ${r.total_days}d scanned &middot; ${r.total_trades} trades &middot; <span class="pnl-positive">${r.wins}W</span> / <span class="pnl-negative">${r.losses}L</span> &middot; <strong>${wr}% WR</strong>`;
                Object.entries(r.by_symbol || {}).forEach(([sym, s]) => {
                    const swr = s.wins / (s.wins + s.losses) * 100 || 0;
                    html += `<br>&nbsp;&nbsp;${sym}: ${s.trades}t &middot; ${s.wins}W/${s.losses}L &middot; ${swr.toFixed(1)}% WR`;
                });
                html += '<br><span style="color:#8b949e;font-size:10px;">' + (r.start_date || '') + ' to ' + (r.end_date || '') + '</span></div>';
            });
            const el = document.getElementById('backtest-summary');
            if (el) {
                el.innerHTML = html;
                el.style.display = html ? 'block' : 'none';
            }
        }

        function renderBacktest() {
            const d = window._lastData;
            if (!d) return;
            const journal = d.journal || [];
            if (!journal.length) {
                document.getElementById('backtest-body').innerHTML = '<p class="empty-state">No journal data to analyze</p>';
                return;
            }
            const stats = {};
            journal.forEach(j => {
                const sym = j.Symbol || 'UNKNOWN';
                const act = j.Action || '';
                const pnlStr = j['P&L %'] || '-';
                const pnl = parseFloat(pnlStr.replace('%', '')) || 0;
                const entry = j.Entry || '';
                const sl = j.SL || '';
                const target = j.Target || '';
                const rr = j.RR || '';
                if (!stats[sym]) stats[sym] = { entries: 0, slHits: 0, targetHits: 0, totalPnl: 0, trades: 0, rrSum: 0, rrCount: 0 };
                if (act.startsWith('BACKTEST') || act.startsWith('BUY') || act.startsWith('ENTRY')) {
                    stats[sym].entries++;
                    stats[sym].trades++;
                    if (rr && !isNaN(parseFloat(rr))) { stats[sym].rrSum += parseFloat(rr); stats[sym].rrCount++; }
                }
                if (act === 'EXIT_SL') stats[sym].slHits++;
                if (act === 'EXIT_T3') stats[sym].targetHits++;
                if (pnl && pnlStr !== '-') stats[sym].totalPnl += pnl;
            });
            let html = '<table><thead><tr><th onclick="sortTable(this,0)">Symbol</th><th onclick="sortTable(this,1)">Entries</th><th onclick="sortTable(this,2)">SL Hits</th><th onclick="sortTable(this,3)">Target Hits</th><th onclick="sortTable(this,4)">Win Rate</th><th onclick="sortTable(this,5)">Total P&L</th><th onclick="sortTable(this,6)">Avg P&L</th><th onclick="sortTable(this,7)">Avg RR</th></tr></thead><tbody>';
            Object.entries(stats).sort((a, b) => b[1].entries - a[1].entries).forEach(([sym, s]) => {
                const exits = s.slHits + s.targetHits;
                const wr = exits > 0 ? ((s.targetHits / exits) * 100).toFixed(1) : '-';
                const avgPnl = s.trades > 0 ? (s.totalPnl / s.trades).toFixed(2) : '-';
                const avgRR = s.rrCount > 0 ? (s.rrSum / s.rrCount).toFixed(2) : '-';
                html += `<tr><td><strong>${sym}</strong></td><td>${s.entries}</td><td>${s.slHits}</td><td>${s.targetHits}</td><td>${wr}${wr !== '-' ? '%' : ''}</td><td class="${s.totalPnl >= 0 ? 'pnl-positive' : 'pnl-negative'}">${s.totalPnl.toFixed(2)}%</td><td>${avgPnl}%</td><td>${avgRR}</td></tr>`;
            });
            html += '</tbody></table>';
            document.getElementById('backtest-body').innerHTML = html;
        }

        function renderPendingTrades() {
            const d = window._lastData;
            if (!d) return;
            const trades = d.pending_trades || [];
            if (!trades.length) {
                document.getElementById('best-trades-body').innerHTML = '<p class="empty-state">No pending trades</p>';
                document.querySelector('#best-trades-tab .section-header span').textContent = 'Best Trade to Take';
                return;
            }
            const latestDate = trades.reduce((latest, t) => t.date > latest ? t.date : latest, trades[0].date || '');
            document.querySelector('#best-trades-tab .section-header span').textContent = 'Best Trade to Take' + (latestDate ? ' \u2014 ' + latestDate : '');
            let html = '<table><thead><tr><th>Date-Time</th><th>Symbol</th><th>Pattern</th><th>Side</th><th>Strike</th><th>Entry</th><th>SL</th><th>T1</th><th>T2</th><th>T3</th><th>RR</th><th>LTP</th><th>Action</th></tr></thead><tbody>';
            trades.forEach(t => {
                const ltp = (d.ltp || {})[t.symbol] || '-';
                const isExecuted = t.executed;
                let actionHtml;
                if (isExecuted) {
                    actionHtml = '<span class="badge badge-profit">Executed</span>';
                } else if (t.side === 'CE') {
                    actionHtml = `<button class="trade-btn buy" onclick="executeTrade('${t.symbol}','${t.side}',${t.strike},${t.entry},${t.sl},${t.t1},${t.t2},${t.t3},'${t.pattern}',${t.rr},'BUY')">Buy CE</button>`;
                } else {
                    actionHtml = `<button class="trade-btn sell" onclick="executeTrade('${t.symbol}','${t.side}',${t.strike},${t.entry},${t.sl},${t.t1},${t.t2},${t.t3},'${t.pattern}',${t.rr},'BUY')">Sell PE</button>`;
                }
                const dtParts = t.date ? t.date.split(' ') : ['-', ''];
                html += `<tr>
                    <td style="font-size:11px">${dtParts[0]}<br><span style="color:#8b949e">${dtParts[1] || ''}</span></td>
                    <td><strong>${t.symbol}</strong></td>
                    <td>${t.pattern}</td>
                    <td>${t.side}</td>
                    <td>${t.strike}</td>
                    <td>${t.entry || '-'}</td>
                    <td>${t.sl || '-'}</td>
                    <td>${t.t1 || '-'}</td>
                    <td>${t.t2 || '-'}</td>
                    <td>${t.t3 || '-'}</td>
                    <td>${t.rr || '-'}</td>
                    <td>${ltp}</td>
                    <td style="white-space:nowrap">${actionHtml}</td>
                </tr>`;
            });
            html += '</tbody></table>';
            document.getElementById('best-trades-body').innerHTML = html;
        }

        async function closePosition(symbol, token, engine) {
            if (!confirm(`Close position for ${symbol}?`)) return;
            try {
                const r = await fetch('/api/trade/close', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({symbol, token, engine})
                });
                const d = await r.json();
                if (d.ok) {
                    showAlert(`Closed ${symbol}: ${d.msg || 'Order placed'}`, 'success');
                    setTimeout(refreshData, 1000);
                } else {
                    showAlert(`Close failed: ${d.error}`, 'error');
                }
            } catch(e) {
                showAlert(`Close error: ${e.message}`, 'error');
            }
        }

        async function executeTrade(symbol, side, strike, entry, sl, t1, t2, t3, pattern, rr, orderType) {
            const btn = event?.target;
            if (btn) { btn.disabled = true; btn.textContent = 'Placing...'; }
            try {
                const r = await fetch('/api/trade/execute', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        symbol, side, strike, entry, sl, t1, t2, t3, pattern, rr,
                        order_type: orderType
                    })
                });
                const d = await r.json();
                if (d.ok) {
                    showAlert(`${orderType} ${symbol} order placed: ${d.order_id}`, 'success');
                    setTimeout(refreshData, 1000);
                } else {
                    showAlert(`Order failed: ${d.error}`, 'error');
                }
            } catch(e) {
                showAlert(`Order error: ${e.message}`, 'error');
            }
            if (btn) { btn.disabled = false; btn.textContent = orderType === 'BUY' ? 'Buy' : 'Sell'; }
        }

        function showAlert(msg, type) {
            const el = document.createElement('div');
            el.className = 'toast-alert ' + (type || 'info');
            el.textContent = msg;
            document.body.appendChild(el);
            setTimeout(() => el.remove(), 3000);
        }

        async function analyzeEntry() {
            const symbol = document.getElementById('analyze-symbol').value.trim().toUpperCase();
            if (!symbol) { document.getElementById('analyze-status').textContent = 'Enter a symbol'; return; }
            const tf = document.getElementById('analyze-tf').value;
            document.getElementById('analyze-status').textContent = 'Analyzing...';
            document.getElementById('analyze-body').innerHTML = '<p class="empty-state">Loading...</p>';
            try {
                const r = await fetch('/api/analyze/entry', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({symbol, timeframe: tf})
                });
                const d = await r.json();
                if (!d.ok) {
                    document.getElementById('analyze-status').textContent = 'Error: ' + (d.error||'');
                    document.getElementById('analyze-body').innerHTML = '<p class="empty-state">' + (d.error||'Error') + '</p>';
                    return;
                }
                document.getElementById('analyze-status').textContent = 'Done';
                let html = '';
                const label = (d.option_type || 'CE') === 'PE' ? 'Bearish' : 'Bullish';
                if (d.bcd) {
                    html += '<div class="match-highlight" style="border-left-color:#58a6ff"><strong>BCD ' + label + ' Reversal</strong><br>';
                    html += '<span style="font-size:11px">Pattern: ' + d.bcd.Pattern + '<br>';
                    html += 'Close(Entry): ' + d.bcd.Close.toFixed(2) + ' | SL: ' + d.bcd.SL.toFixed(2) + '<br>';
                    html += 'T1: ' + (d.bcd.T1||'-') + ' | T2: ' + (d.bcd.T2||'-') + ' | T3: ' + (d.bcd.T3||'-') + '<br>';
                    html += 'RR: ' + d.bcd.RR + '</span></div>';
                }
                if (d.pinbar) {
                    html += '<div class="match-highlight" style="border-left-color:#d29922"><strong>Pin Bar ' + label + '</strong><br>';
                    html += '<span style="font-size:11px">Close: ' + d.pinbar.Close.toFixed(2) + ' | SL: ' + d.pinbar.SL.toFixed(2) + '</span></div>';
                }
                if (d.swing) {
                    html += '<div class="match-highlight" style="border-left-color:#d29922"><strong>Swing ' + label + '</strong><br>';
                    html += '<span style="font-size:11px">Close: ' + d.swing.Close.toFixed(2) + ' | SL: ' + d.swing.SL.toFixed(2) + '</span></div>';
                }
                if (d.anchors && d.anchors.length) {
                    d.anchors.forEach(a => {
                        html += '<div class="match-highlight" style="border-left-color:#d29922"><strong>Anchor: ' + a.Pattern + '</strong><br>';
                        html += '<span style="font-size:11px">Close: ' + a.Close.toFixed(2) + ' | SL: ' + a.SL.toFixed(2) + '</span></div>';
                    });
                }
                if (!html) html = '<p class="empty-state">No patterns found for ' + (d.label||d.symbol||'') + ' on ' + tf + '</p>';
                document.getElementById('analyze-body').innerHTML = html;
            } catch(e) {
                document.getElementById('analyze-status').textContent = 'Error';
                document.getElementById('analyze-body').innerHTML = '<p class="empty-state">Request failed: ' + e.message + '</p>';
            }
        }

        async function refreshData() {
            try {
                const r = await fetch('/api/status');
                const d = await r.json();
                window._lastData = d;

                for (const [pid, prog] of Object.entries(d.programs || {})) {
                    const btn = document.querySelector(`.prog-card[data-prog="${pid}"]`);
                    if (!btn) continue;
                    const dot = btn.querySelector('.status-dot');
                    const label = btn.querySelector('.status-label');
                    const bar = btn.querySelector('.status-bar');
                    const startBtn = btn.querySelector('.start-btn');
                    const stopBtn = btn.querySelector('.stop-btn');
                    if (prog.running) {
                        dot.className = 'status-dot live';
                        label.textContent = 'Live';
                        label.className = 'status-label live';
                        bar.className = 'status-bar live';
                        startBtn.disabled = true;
                        startBtn.style.opacity = '0.4';
                        stopBtn.disabled = false;
                        stopBtn.style.opacity = '1';
                    } else {
                        dot.className = 'status-dot closed';
                        label.textContent = 'Closed';
                        label.className = 'status-label closed';
                        bar.className = 'status-bar closed';
                        startBtn.disabled = false;
                        startBtn.style.opacity = '1';
                        stopBtn.disabled = true;
                        stopBtn.style.opacity = '0.4';
                    }
                }

                const cfg = d.config || {};
                for (const [pid, progCfg] of Object.entries(cfg)) {
                    const inputs = document.querySelectorAll(`.config-input[data-prog="${pid}"]`);
                    inputs.forEach(inp => {
                        const field = inp.getAttribute('data-field');
                        if (progCfg[field] !== undefined) inp.value = progCfg[field];
                    });
                }

                for (const [pid, prog] of Object.entries(d.programs || {})) {
                    (prog.log_tail || []).forEach(line => {
                        if (!line.includes('[ERROR]')) return;
                        const key = pid + line.slice(0, 60);
                        if (seenAlerts.has(key)) return;
                        seenAlerts.add(key);
                        if (seenAlerts.size > 200) seenAlerts.clear();
                        showAlert(pid + ': ' + line.split('[ERROR]')[1] || line);
                    });
                }
                renderReport();
                refreshTokenStatus();
                refreshBacktestMode();
                refreshNopaConfig();
            } catch(e) { console.log('Refresh error:', e); }
        }

        async function refreshTokenStatus() {
            try {
                const r = await fetch('/api/token/check');
                const tk = await r.json();
                const banner = document.getElementById('token-banner');
                const text = document.getElementById('token-banner-text');
                const btn = document.getElementById('token-gen-btn');
                const panel = document.getElementById('token-panel');
                if (tk.valid) {
                    banner.className = 'token-banner token-valid';
                    text.innerHTML = '<strong>Token:</strong> Valid';
                    btn.style.display = 'none';
                    if (panel) panel.style.display = 'none';
                } else if (tk.reason && tk.reason.includes('expired')) {
                    banner.className = 'token-banner token-expired';
                    text.innerHTML = '<strong>Token:</strong> Expired <span style="font-weight:normal;font-size:11px;color:#8b949e;">(' + tk.reason + ')</span>';
                    btn.style.display = 'inline-block';
                } else {
                    banner.className = 'token-banner token-missing';
                    text.innerHTML = '<strong>Token:</strong> ' + (tk.reason || 'Missing');
                    btn.style.display = 'inline-block';
                }
            } catch(e) { console.log('Token check error:', e); }
        }

        async function refreshBacktestMode() {
            try {
                const r = await fetch('/api/backtest/mode');
                const d = await r.json();
                const btToggle = document.getElementById('backtest-toggle');
                if (btToggle) btToggle.checked = d.enabled === true;
                if (btToggle && btToggle.checked) renderBacktest();
            } catch(e) { console.log('Backtest mode error:', e); }
        }

        async function refreshNopaConfig() {
            try {
                const r = await fetch('/api/scanner/config');
                const d = await r.json();
                const nopaToggle = document.getElementById('nopa-toggle');
                if (nopaToggle) nopaToggle.checked = d.disable_nopa === true;
            } catch(e) { console.log('Nopa config error:', e); }
        }

        async function showTokenPanel() {
            document.getElementById('token-panel').style.display = 'block';
            document.getElementById('token-gen-btn').style.display = 'none';
            document.getElementById('token-feedback').textContent = '';
            try {
                const r = await fetch('/api/token/url');
                const d = await r.json();
                document.getElementById('token-url-text').textContent = d.url || 'Error loading URL';
            } catch(e) {
                document.getElementById('token-url-text').textContent = 'Failed to load login URL';
            }
        }

        function hideTokenPanel() {
            document.getElementById('token-panel').style.display = 'none';
            document.getElementById('token-gen-btn').style.display = 'inline-block';
        }

        function copyTokenUrl() {
            const url = document.getElementById('token-url-text').textContent;
            if (url && url.startsWith('http')) {
                navigator.clipboard.writeText(url).then(() => {
                    const hint = document.querySelector('.token-copy-hint');
                    hint.textContent = 'Copied!';
                    setTimeout(() => { hint.textContent = 'Click to copy'; }, 2000);
                }).catch(() => {});
            }
        }

        async function submitToken() {
            const input = document.getElementById('token-redirect-input');
            const fb = document.getElementById('token-feedback');
            const raw = input.value.trim();
            if (!raw) { fb.textContent = 'Please paste the redirect URL'; fb.style.color = '#f85149'; return; }
            let requestToken = raw;
            if (requestToken.includes('request_token=')) {
                requestToken = requestToken.split('request_token=')[1].split('&')[0];
            }
            fb.textContent = 'Exchanging token...';
            fb.style.color = '#8b949e';
            try {
                const r = await fetch('/api/token/exchange', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({request_token: requestToken})
                });
                const d = await r.json();
                if (d.ok) {
                    fb.textContent = 'Token saved successfully!';
                    fb.style.color = '#3fb950';
                    document.getElementById('token-gen-btn').style.display = 'none';
                    setTimeout(() => { hideTokenPanel(); refreshData(); }, 1500);
                } else {
                    fb.textContent = 'Failed: ' + (d.error || 'Unknown error');
                    fb.style.color = '#f85149';
                }
            } catch(e) {
                fb.textContent = 'Error: ' + e.message;
                fb.style.color = '#f85149';
            }
        }

        let seenAlerts = new Set();

        function showAlert(msg) {
            let el = document.getElementById('alert-toast');
            if (!el) {
                el = document.createElement('div');
                el.id = 'alert-toast';
                document.body.appendChild(el);
            }
            el.innerHTML = '<span class="alert-icon">&#9888;</span><span class="alert-msg">' + msg.replace(/</g,'&lt;').replace(/>/g,'&gt;') + '</span><button class="alert-close" onclick="this.parentElement.remove()">&times;</button>';
            el.className = 'alert-toast show';
            clearTimeout(el._hideTimer);
            el._hideTimer = setTimeout(() => { if (el) el.className = 'alert-toast'; }, 15000);
        }

        async function monthlyExport() {
            try {
                const r = await fetch('/api/export/monthly', {method: 'POST'});
                const d = await r.json();
                const fb = document.getElementById('cfg-fb-backtest');
                if (d.ok) {
                    fb.textContent = d.exported > 0 ? `Exported ${d.exported} trades to ${d.sheets.join(', ')}` : 'No completed trades to export';
                    fb.style.color = '#3fb950';
                } else {
                    fb.textContent = 'Export failed';
                    fb.style.color = '#f85149';
                }
                setTimeout(() => { fb.textContent = ''; }, 5000);
                if (d.ok) setTimeout(refreshData, 500);
            } catch(e) { console.log(e); }
        }

        async function clearLogs() {
            try {
                const r = await fetch('/api/logs/clear', {method: 'POST'});
                const d = await r.json();
                if (d.ok) setTimeout(refreshData, 300);
            } catch(e) { console.log(e); }
        }
        async function clearJournal() {
            try {
                const r = await fetch('/api/journal/clear', {method: 'POST'});
                const d = await r.json();
                if (d.ok) setTimeout(refreshData, 300);
            } catch(e) { console.log(e); }
        }

        setInterval(refreshData, {{ refresh * 1000 }});
        window.addEventListener('load', () => { refreshData(); });
    </script>
    <style>
        * { box-sizing: border-box; margin: 0; padding: 0; }
        body { font-family: 'Segoe UI', sans-serif; background: #0d1117; color: #c9d1d9; padding: 20px; }
        h1 { color: #58a6ff; margin-bottom: 16px; font-size: 22px; display: flex; align-items: center; gap: 10px; }
        h1 small { font-size: 12px; color: #8b949e; font-weight: normal; }
        h2 { color: #8b949e; font-size: 14px; margin-bottom: 10px; border-bottom: 1px solid #30363d; padding-bottom: 6px; text-transform: uppercase; letter-spacing: 0.5px; }

        .empty-state { color: #8b949e; padding: 20px; text-align: center; font-size: 13px; }

        .token-banner { border-radius: 8px; padding: 10px 14px; margin-bottom: 16px; font-size: 13px; transition: all 0.3s; }
        .token-banner.token-hidden { display: none; }
        .token-banner.token-valid { background: #3fb95022; border: 1px solid #3fb950; color: #3fb950; }
        .token-banner.token-expired { background: #d2992222; border: 1px solid #d29922; color: #d29922; }
        .token-banner.token-missing { background: #f8514922; border: 1px solid #f85149; color: #f85149; }
        .token-banner-row { display: flex; align-items: center; justify-content: space-between; gap: 12px; }
        .token-gen-btn { background: #238636; color: #fff; border: none; border-radius: 4px; padding: 5px 14px; font-size: 12px; cursor: pointer; font-weight: 600; }
        .token-gen-btn:hover { background: #2ea043; }
        .token-gen-btn:disabled { opacity: 0.5; cursor: not-allowed; }

        .token-panel { margin-top: 12px; padding: 14px; background: #0d1117; border-radius: 6px; border: 1px solid #30363d; }
        .token-step { display: flex; align-items: center; gap: 8px; font-size: 12px; color: #c9d1d9; margin-bottom: 8px; }
        .token-step-num { background: #21262d; color: #58a6ff; width: 20px; height: 20px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 11px; font-weight: bold; flex-shrink: 0; }
        .token-url-box { background: #161b22; border: 1px solid #30363d; border-radius: 4px; padding: 8px 10px; font-size: 11px; color: #58a6ff; word-break: break-all; cursor: pointer; margin: 6px 0 12px; display: flex; justify-content: space-between; align-items: center; gap: 8px; }
        .token-url-box:hover { border-color: #58a6ff; }
        .token-copy-hint { font-size: 10px; color: #8b949e; white-space: nowrap; }
        .token-input-row { display: flex; gap: 8px; margin: 8px 0; }
        .token-input { flex: 1; background: #161b22; color: #c9d1d9; border: 1px solid #30363d; border-radius: 4px; padding: 6px 10px; font-size: 12px; }
        .token-input:focus { outline: none; border-color: #58a6ff; }
        .token-submit-btn { background: #238636; color: #fff; border: none; border-radius: 4px; padding: 6px 14px; font-size: 12px; cursor: pointer; font-weight: 600; white-space: nowrap; }
        .token-submit-btn:hover { background: #2ea043; }
        .token-feedback { font-size: 12px; margin: 6px 0; }
        .token-close-btn { background: transparent; color: #8b949e; border: 1px solid #30363d; border-radius: 4px; padding: 3px 10px; font-size: 11px; cursor: pointer; margin-top: 6px; }
        .token-close-btn:hover { color: #c9d1d9; }

        .stats-grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin-bottom: 20px; }
        .stat-card { background: #161b22; border: 1px solid #30363d; border-radius: 8px; padding: 14px; text-align: center; }
        .stat-card .value { font-size: 26px; font-weight: bold; }
        .stat-card .label { font-size: 11px; color: #8b949e; margin-top: 2px; text-transform: uppercase; }

        .program-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin-bottom: 20px; }
        .prog-card { background: #161b22; border: 1px solid #30363d; border-radius: 8px; padding: 14px; position: relative; overflow: hidden; }

        .status-bar { position: absolute; top: 0; left: 0; width: 4px; height: 100%; transition: background 0.3s; }
        .status-bar.live { background: #3fb950; }
        .status-bar.closed { background: #f85149; }

        .prog-header { display: flex; align-items: center; gap: 10px; margin-bottom: 6px; }
        .prog-icon { width: 10px; height: 10px; border-radius: 50%; flex-shrink: 0; }
        .prog-name { font-weight: 600; font-size: 13px; flex: 1; }
        .prog-desc { font-size: 11px; color: #8b949e; margin-bottom: 10px; }
        .prog-footer { display: flex; align-items: center; justify-content: space-between; }

        .status-group { display: flex; align-items: center; gap: 5px; }
        .status-dot { display: inline-block; width: 8px; height: 8px; border-radius: 50%; }
        .status-dot.live { background: #3fb950; box-shadow: 0 0 8px #3fb950aa; animation: pulse 1.5s infinite; }
        .status-dot.closed { background: #f85149; }
        @keyframes pulse { 0%, 100% { opacity: 1; } 50% { opacity: 0.5; } }
        .status-label { font-size: 11px; font-weight: 600; }
        .status-label.live { color: #3fb950; }
        .status-label.closed { color: #f85149; }

        .prog-actions { display: flex; gap: 6px; }
        .prog-actions button { padding: 4px 14px; border-radius: 6px; cursor: pointer; font-size: 11px; font-weight: 600; border: 1px solid; transition: all 0.2s; }
        .prog-actions button:disabled { cursor: not-allowed; }
        .start-btn { background: #3fb95022; border-color: #3fb950; color: #3fb950; }
        .start-btn:hover:not(:disabled) { background: #3fb95044; }
        .stop-btn { background: #f8514922; border-color: #f85149; color: #f85149; }
        .stop-btn:hover:not(:disabled) { background: #f8514944; }
        .anchor-btn { background: #d2992222; border-color: #d29922; color: #d29922; }
        .anchor-btn:hover:not(:disabled) { background: #d2992244; }
        .anchor-btn:disabled { opacity: 0.5; cursor: not-allowed; }

        .prog-config { border-top: 1px solid #30363d; margin-top: 0; }
        .config-header { padding: 6px 14px; font-size: 11px; color: #8b949e; cursor: pointer; user-select: none; display: flex; align-items: center; gap: 6px; }
        .config-header:hover { color: #c9d1d9; background: #1c2128; }
        .config-arrow { font-size: 9px; }
        .config-body { display: none; padding: 6px 14px 10px; background: #0d1117; }
        .config-row { display: flex; align-items: center; justify-content: space-between; padding: 3px 0; gap: 8px; }
        .config-label { font-size: 10px; color: #8b949e; }
        .config-input { background: #161b22; color: #c9d1d9; border: 1px solid #30363d; border-radius: 3px; padding: 2px 6px; font-size: 10px; width: 120px; }
        .config-input:focus { outline: none; border-color: #58a6ff; }
        .config-save-btn { margin-top: 6px; padding: 3px 12px; background: #238636; color: #fff; border: none; border-radius: 4px; font-size: 10px; cursor: pointer; }
        .config-save-btn:hover { background: #2ea043; }
        .config-feedback { font-size: 10px; margin-left: 8px; }

        .section-panel { background: #161b22; border: 1px solid #30363d; border-radius: 8px; margin-bottom: 14px; overflow: hidden; }
        .section-panel .section-header { background: #1c2128; color: #c9d1d9; font-size: 12px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; padding: 8px 14px; border-bottom: 1px solid #30363d; display: flex; align-items: center; justify-content: space-between; gap: 10px; }
        .section-panel .empty-state { color: #8b949e; padding: 20px; text-align: center; font-size: 13px; }
        .filter-select { background: #0d1117; color: #c9d1d9; border: 1px solid #30363d; border-radius: 4px; padding: 3px 8px; font-size: 11px; cursor: pointer; }
        .filter-select:focus { outline: none; border-color: #58a6ff; }
        .clear-logs-btn { background: #21262d; color: #c9d1d9; border: 1px solid #30363d; border-radius: 4px; padding: 3px 10px; font-size: 10px; cursor: pointer; white-space: nowrap; }
        .clear-logs-btn:hover { background: #30363d; }
        .pos-filter-btn { background: #21262d; color: #8b949e; border: 1px solid #30363d; border-radius: 4px; padding: 2px 8px; font-size: 10px; cursor: pointer; white-space: nowrap; }
        .pos-filter-btn:hover { color: #c9d1d9; }
        .pos-filter-btn.active { background: #1f6feb33; color: #58a6ff; border-color: #58a6ff; }
        .export-btn { background: #21262d; color: #c9d1d9; border: 1px solid #30363d; border-radius: 4px; padding: 3px 10px; font-size: 10px; cursor: pointer; white-space: nowrap; }
        .export-btn:hover { background: #30363d; }

        .toggle-label { display: inline-flex; align-items: center; gap: 0; cursor: pointer; user-select: none; font-size: 10px; color: #8b949e; }
        .toggle-label input { display: none; }
        .toggle-slider { width: 28px; height: 14px; background: #30363d; border-radius: 10px; position: relative; transition: background 0.2s; }
        .toggle-slider::after { content: ''; position: absolute; top: 2px; left: 2px; width: 10px; height: 10px; background: #8b949e; border-radius: 50%; transition: all 0.2s; }
        .toggle-label input:checked + .toggle-slider { background: #238636; }
        .toggle-label input:checked + .toggle-slider::after { left: 16px; background: #fff; }

        .reports-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; align-items: start; }
        .reports-left .section-panel { margin-bottom: 0; border-top: none; border-radius: 0 0 8px 8px; }

        .pnl-positive { color: #3fb950 !important; font-weight: 600; }
        .pnl-negative { color: #f85149 !important; font-weight: 600; }
        .reports-right .section-panel { margin-bottom: 0; border-top: none; border-radius: 0 0 8px 8px; }
        .left-tab-bar { display: flex; gap: 0; background: #1c2128; border: 1px solid #30363d; border-bottom: none; border-radius: 8px 8px 0 0; overflow: hidden; }
        .left-tab-btn { background: transparent; border: none; color: #8b949e; padding: 8px 16px; cursor: pointer; font-size: 12px; font-weight: 600; border-bottom: 2px solid transparent; flex: 1; text-align: center; }
        .left-tab-btn:hover { color: #c9d1d9; background: #21262d; }
        .left-tab-btn.active { color: #58a6ff; border-bottom-color: #58a6ff; background: #161b22; }
        .left-tab-content { display: none; }
        .left-tab-content.active { display: block; }

        .tab-bar { display: flex; gap: 0; background: #1c2128; border: 1px solid #30363d; border-bottom: none; border-radius: 8px 8px 0 0; overflow: hidden; }
        .tab-btn { background: transparent; border: none; color: #8b949e; padding: 8px 16px; cursor: pointer; font-size: 12px; font-weight: 600; border-bottom: 2px solid transparent; flex: 1; text-align: center; }
        .tab-btn:hover { color: #c9d1d9; background: #21262d; }
        .tab-btn.active { color: #58a6ff; border-bottom-color: #58a6ff; background: #161b22; }
        .tab-content { display: none; }
        .tab-content.active { display: block; }

        table { width: 100%; border-collapse: collapse; margin-bottom: 16px; background: #161b22; border-radius: 8px; overflow: hidden; }
        th { background: #21262d; color: #8b949e; font-size: 11px; text-transform: uppercase; padding: 8px 10px; text-align: left; border-bottom: 1px solid #30363d; cursor: pointer; user-select: none; }
        th:hover { color: #58a6ff; }
        th::after { content: ' \\25B4\\25BE'; font-size: 8px; opacity: 0.3; }
        td { padding: 7px 10px; border-bottom: 1px solid #21262d; font-size: 12px; }
        tr:hover td { background: #1c2128; }

        .badge { display: inline-block; padding: 2px 8px; border-radius: 12px; font-size: 10px; font-weight: 600; }
        .badge-open { background: #1f6feb33; color: #58a6ff; }
        .badge-closed { background: #3fb95033; color: #3fb950; }
        .badge-failed { background: #f8514933; color: #f85149; }
        .badge-mutated { background: #d2992233; color: #d29922; }
        .badge-profit { background: #3fb95033; color: #3fb950; }
        .badge-loss { background: #f8514933; color: #f85149; }

        .log-box { background: #0d1117; border: 1px solid #30363d; border-radius: 8px; padding: 10px; font-family: 'Consolas', monospace; font-size: 11px; max-height: 400px; overflow-y: auto; line-height: 1.5; }
        .log-box div:nth-child(odd) { background: #161b22; }
        .match-highlight { background: #1f6feb11; border-left: 3px solid #58a6ff; padding: 2px 8px; margin: 2px 0; font-family: 'Consolas', monospace; font-size: 11px; }
        .backtest-summary { background: #161b22; border: 1px solid #30363d; border-radius: 8px; padding: 10px 14px; font-size: 12px; line-height: 1.7; margin-bottom: 8px; }

        .alert-toast { position: fixed; top: 12px; right: 12px; z-index: 9999; max-width: 420px; background: #f85149; color: #fff; padding: 10px 14px; border-radius: 8px; font-size: 12px; display: flex; align-items: flex-start; gap: 8px; box-shadow: 0 4px 20px rgba(0,0,0,0.5); transform: translateX(120%); opacity: 0; transition: all 0.4s; }
        .alert-toast.show { transform: translateX(0); opacity: 1; }
        .alert-icon { font-size: 16px; flex-shrink: 0; margin-top: 1px; }
        .alert-msg { flex: 1; word-break: break-word; }
        .alert-close { background: none; border: none; color: rgba(255,255,255,0.7); font-size: 18px; cursor: pointer; padding: 0 0 0 4px; line-height: 1; flex-shrink: 0; }
        .alert-close:hover { color: #fff; }
        .last-updated { color: #8b949e; font-size: 11px; text-align: right; margin-top: 8px; }
        .trade-btn { padding: 4px 12px; border: none; border-radius: 4px; cursor: pointer; font-size: 11px; font-weight: 600; margin: 0 2px; transition: opacity 0.2s; }
        .trade-btn:disabled { opacity: 0.5; cursor: not-allowed; }
        .trade-btn.buy { background: #3fb950; color: #0d1117; }
        .trade-btn.sell { background: #f85149; color: #fff; }
        .toast-alert { position: fixed; bottom: 20px; right: 20px; padding: 10px 18px; border-radius: 8px; font-size: 13px; font-weight: 500; z-index: 9999; box-shadow: 0 4px 20px rgba(0,0,0,0.5); animation: fadeInUp 0.3s ease; max-width: 400px; }
        .toast-alert.success { background: #3fb950; color: #0d1117; }
        .toast-alert.error { background: #f85149; color: #fff; }
        .toast-alert.info { background: #1f6feb; color: #fff; }
        @keyframes fadeInUp { from { opacity: 0; transform: translateY(20px); } to { opacity: 1; transform: translateY(0); } }

        @media (max-width: 1100px) {
            .reports-grid { grid-template-columns: 1fr; }
        }
        @media (max-width: 900px) {
            .stats-grid { grid-template-columns: repeat(2, 1fr); }
            .program-grid { grid-template-columns: 1fr; }
        }
    </style>
</head>
<body>
    <h1>Trading Control Center-02 <small>Steering & Dashboard-02</small></h1>

    <div id="token-banner" class="token-banner token-hidden">
        <div class="token-banner-row">
            <span id="token-banner-text"></span>
            <button id="token-gen-btn" class="token-gen-btn" onclick="showTokenPanel()" style="display:none">Generate Token</button>
        </div>
        <div id="token-panel" class="token-panel" style="display:none">
            <div class="token-step">
                <span class="token-step-num">1</span>
                <span>Open this URL and log in to Kite:</span>
            </div>
            <div class="token-url-box" id="token-url-box" onclick="copyTokenUrl()">
                <span id="token-url-text">Loading...</span>
                <span class="token-copy-hint">Click to copy</span>
            </div>
            <div class="token-step">
                <span class="token-step-num">2</span>
                <span>After login, paste the full redirect URL here:</span>
            </div>
            <div class="token-input-row">
                <input type="text" id="token-redirect-input" class="token-input" placeholder="https://... or ?request_token=...">
                <button class="token-submit-btn" onclick="submitToken()">Submit & Save</button>
            </div>
            <div id="token-feedback" class="token-feedback"></div>
            <button class="token-close-btn" onclick="hideTokenPanel()">Cancel</button>
        </div>
    </div>

    <div class="stats-grid">
        <div class="stat-card">
            <div class="value" id="stat-active" style="color:#58a6ff;">0</div>
            <div class="label">Active Positions</div>
        </div>
        <div class="stat-card">
            <div class="value" id="stat-total" style="color:#58a6ff;">0</div>
            <div class="label">Total Trades</div>
        </div>
        <div class="stat-card">
            <div class="value" id="stat-winrate" style="color:#8b949e;">0%</div>
            <div class="label">Win Rate</div>
        </div>
        <div class="stat-card">
            <div class="value" id="stat-pnl" style="color:#8b949e;">0%</div>
            <div class="label">P&L</div>
        </div>
    </div>

    <h2>Programs</h2>
    <div class="program-grid">
        {% for pid, prog in programs.items() %}
        <div class="prog-card" data-prog="{{ pid }}">
            <div class="status-bar closed"></div>
            <div class="prog-header">
                <span class="prog-icon" style="background:{{ prog.color }}"></span>
                <span class="prog-name">{{ prog.name }}</span>
                <div class="status-group">
                    <span class="status-dot closed"></span>
                    <span class="status-label closed">Closed</span>
                </div>
            </div>
            <div class="prog-desc">{{ prog.desc }}</div>
            <div class="prog-footer" style="margin-bottom:{% if prog.config_fields %}6px{% else %}0{% endif %};">
                <div class="prog-actions">
                    <button class="start-btn" onclick="event.stopPropagation();toggleProgram('{{ pid }}','start')">Start</button>
                    <button class="stop-btn" onclick="event.stopPropagation();toggleProgram('{{ pid }}','stop')" disabled>Stop</button>
                </div>
            </div>
            {% if prog.config_fields %}
            <div class="prog-config">
                <div class="config-header" onclick="event.stopPropagation();toggleConfig(this)">
                    <span class="config-arrow">&#9654;</span> Configuration
                </div>
                <div class="config-body">
                    {% for field_key, field in prog.config_fields.items() %}
                    <div class="config-row">
                        <span class="config-label">{{ field.label }}</span>
                        {% if field.type == "select" %}
                        <select class="config-input" data-prog="{{ pid }}" data-field="{{ field_key }}">
                            {% for opt in field.options %}
                            <option value="{{ opt }}">{{ opt }}</option>
                            {% endfor %}
                        </select>
                        {% else %}
                        <input type="number" class="config-input" data-prog="{{ pid }}" data-field="{{ field_key }}" value="{{ field.default }}" step="any">
                        {% endif %}
                    </div>
                    {% endfor %}
                    <button class="config-save-btn" onclick="event.stopPropagation();saveConfig('{{ pid }}')">Save Config</button>
                    <span class="config-feedback" id="cfg-fb-{{ pid }}"></span>
                </div>
            </div>
            {% endif %}
        </div>
        {% endfor %}
    </div>

    <h2>Live Reports</h2>

    <div class="reports-grid">
        <div class="reports-left">
            <div class="left-tab-bar">
                <button class="left-tab-btn active" onclick="switchLeftTab('active-pos-tab')">Live Positions</button>
                <button class="left-tab-btn" onclick="switchLeftTab('backtest-tab')">Backtest Summary</button>
                <div style="display:flex;align-items:center;gap:6px;margin-left:auto;padding:0 10px;">
                    <label class="toggle-label" title="Toggle backtest mode">
                        <span style="font-size:10px;font-weight:normal;text-transform:none;letter-spacing:0;margin-right:6px;color:#8b949e;">Off</span>
                        <input type="checkbox" id="backtest-toggle" onchange="toggleBacktestMode(this.checked)">
                        <span class="toggle-slider"></span>
                        <span style="font-size:10px;font-weight:normal;text-transform:none;letter-spacing:0;margin-left:6px;color:#8b949e;">On</span>
                    </label>
                    <label class="toggle-label" title="Disable NoPA left-side filter (engines reload config each scan cycle)">
                        <span style="font-size:10px;font-weight:normal;text-transform:none;letter-spacing:0;margin-right:6px;color:#8b949e;">NoPA</span>
                        <input type="checkbox" id="nopa-toggle" onchange="toggleNopa(this.checked)">
                        <span class="toggle-slider"></span>
                        <span style="font-size:10px;font-weight:normal;text-transform:none;letter-spacing:0;margin-left:6px;color:#8b949e;">NoPA</span>
                    </label>
                </div>
            </div>
            <div id="active-pos-tab" class="left-tab-content active">
                <div class="section-panel">
                    <div class="section-header">
                        <span>Positions</span>
                        <div style="display:flex;gap:4px;">
                            <button class="pos-filter-btn active" onclick="setFilter('position','active')">Active</button>
                            <button class="pos-filter-btn" onclick="setFilter('position','completed')">Completed</button>
                            <button class="pos-filter-btn" onclick="setFilter('position','sl_hit')">SL Hit</button>
                            <button class="pos-filter-btn" onclick="setFilter('position','all')">All</button>
                        </div>
                    </div>
                    <div id="active-positions-body"><p class="empty-state">No positions</p></div>
                </div>
            </div>
            <div id="backtest-tab" class="left-tab-content">
                <div class="section-panel">
                    <div class="section-header">
                        <span>Backtest Summary</span>
                        <div style="display:flex;align-items:center;gap:8px;">
                            <button class="export-btn" onclick="monthlyExport()">Export Month</button>
                        </div>
                    </div>
                    <div id="backtest-body"><p class="empty-state">No journal data to analyze</p></div>
                    <div style="padding:0 14px 8px;"><span id="cfg-fb-backtest" style="font-size:11px;"></span></div>
                </div>
            </div>
        </div>
        <div class="reports-right">
            <div class="tab-bar">
                <button class="tab-btn active" onclick="switchTab('log-tab')">Live Log</button>
                <button class="tab-btn" onclick="switchTab('journal-tab')">All Possible Trade</button>
                <button class="tab-btn" onclick="switchTab('scan-tab')">Scan Matches</button>
                <button class="tab-btn" onclick="switchTab('best-trades-tab')">Best Trades</button>
                <button class="tab-btn" onclick="switchTab('analyze-tab')">Analyze Entry</button>
            </div>
            <div id="log-tab" class="tab-content active">
                <div class="section-panel">
                    <div class="section-header"><span>Live Log</span>
                        <div style="display:flex;align-items:center;gap:6px;">
                            <select onchange="setFilter('log',this.value)" class="filter-select">
                                <option value="all">All</option>
                                <option value="index">Index</option>
                                <option value="bear_index">Bear Index</option>
                                <option value="nifty50">Nifty 50</option>
                                <option value="bear_nifty50">Bear Nifty50</option>
                            </select>
                            <button class="clear-logs-btn" onclick="clearLogs()">Clear Logs</button>
                        </div>
                    </div>
                    <div class="log-box" id="log-body"></div>
                </div>
            </div>
            <div id="journal-tab" class="tab-content">
                <div class="section-panel">
                    <div class="section-header"><span>All Possible Trade</span>
                        <div style="display:flex;align-items:center;gap:6px;">
                            <select onchange="setFilter('journal',this.value)" class="filter-select">
                                <option value="all">All</option>
                                <option value="index">Index</option>
                                <option value="bear_index">Bear Index</option>
                                <option value="nifty50">Nifty 50</option>
                                <option value="bear_nifty50">Bear Nifty50</option>
                            </select>
                            <button class="clear-logs-btn" onclick="clearJournal()">Clear</button>
                        </div>
                    </div>
                    <div id="backtest-summary"></div>
                    <div id="journal-body"><p class="empty-state">No journal entries yet</p></div>
                </div>
            </div>
            <div id="scan-tab" class="tab-content">
                <style>
                    #scan-tab .scan-table{border-collapse:collapse;width:100%;font-size:12px;color:#e6e6e6;margin-bottom:6px}
                    #scan-tab .scan-table th,#scan-tab .scan-table td{border:1px solid #333;padding:4px 8px;text-align:right;white-space:nowrap}
                    #scan-tab .scan-table th{background:#16213e;position:sticky;top:0}
                    #scan-tab .scan-table td:nth-child(1),#scan-tab .scan-table td:nth-child(2),#scan-tab .scan-table td:nth-child(3),#scan-tab .scan-table td:nth-child(11){text-align:left}
                    #scan-tab .scan-subhead{margin:14px 0 6px;font-weight:600;color:#58a6ff;border-bottom:1px solid #333;padding-bottom:4px}
                    #scan-tab .flag-label{font-size:11px;padding:2px 8px;border-radius:10px;background:#3a2a00;color:#d29922;border:1px solid #d29922;margin-left:8px}
                    #scan-tab .flag-label.on{background:#0a3a1a;color:#3fb950;border-color:#3fb950}
                    #scan-tab .scan-toggle{margin-left:auto;display:flex;align-items:center;gap:8px;font-size:12px}
                    #scan-tab .switch{position:relative;display:inline-block;width:42px;height:22px}
                    #scan-tab .switch input{opacity:0;width:0;height:0}
                    #scan-tab .slider{position:absolute;cursor:pointer;top:0;left:0;right:0;bottom:0;background:#444;border-radius:22px;transition:.2s}
                    #scan-tab .slider:before{position:absolute;content:"";height:16px;width:16px;left:3px;bottom:3px;background:#e6e6e6;border-radius:50%;transition:.2s}
                    #scan-tab .switch input:checked + .slider{background:#3fb950}
                    #scan-tab .switch input:checked + .slider:before{transform:translateX(20px)}
                </style>
                <div class="section-panel">
                    <div class="section-header"><span>Live Scanner &amp; Positions (Nifty 50)</span>
                        <span class="scan-toggle">
                            <span id="nifty50-live-label" class="flag-label">SCAN-ONLY</span>
                            <label class="switch" title="ON = place real orders. OFF = scan-only (display, no orders)">
                                <input type="checkbox" id="nifty50-live-toggle" onchange="setLiveExec('nifty50', this.checked)">
                                <span class="slider"></span>
                            </label>
                        </span>
                    </div>
                    <div class="scan-subhead">New Scan Results</div>
                    <div id="scan-staged-body"><p class="empty-state">No scan matches yet</p></div>
                    <div class="scan-subhead">Active Positions (carried forward)</div>
                    <div id="scan-active-body"><p class="empty-state">No active positions</p></div>
                </div>
            </div>
            <div id="best-trades-tab" class="tab-content">
                <div class="section-panel">
                    <div class="section-header"><span>Best Trade to Take</span></div>
                    <div id="best-trades-body"><p class="empty-state">No pending trades</p></div>
                </div>
            </div>
            <div id="analyze-tab" class="tab-content">
                <div class="section-panel">
                    <div class="section-header"><span>Analyze Entry</span></div>
                    <div style="padding:12px;display:flex;gap:8px;flex-wrap:wrap;align-items:center;">
                        <input id="analyze-symbol" placeholder="Symbol or contract (e.g. NIFTY or NESTLEIND26JUL1460CE)" style="width:300px;padding:6px 10px;border:1px solid #333;border-radius:4px;background:#1a1a2e;color:#e6e6e6;">
                        <select id="analyze-tf" style="padding:6px 10px;border:1px solid #333;border-radius:4px;background:#1a1a2e;color:#e6e6e6;">
                            <option value="3minute">3min</option>
                            <option value="5minute">5min</option>
                            <option value="15minute" selected>15min</option>
                            <option value="30minute">30min</option>
                        </select>
                        <button class="config-save-btn" onclick="analyzeEntry()">Analyze</button>
                        <span id="analyze-status" style="font-size:11px;color:#8b949e;"></span>
                    </div>
                    <div id="analyze-body"><p class="empty-state">Enter a symbol or contract and click Analyze</p></div>
                </div>
            </div>
        </div>
    </div>

    <div class="last-updated" id="last-updated">Loading...</div>
</body>
</html>
"""

# ──────────────────────────────────────────────
#  FLASK ROUTES — API Endpoints
# ──────────────────────────────────────────────

@app.route("/")
def dashboard():
    return render_template_string(HTML_TEMPLATE, refresh=REFRESH_SECONDS, programs=PROGRAMS)

@app.route("/api/status")
def api_status():
    with data_lock:
        prog_status = {}
        for pid in PROGRAMS:
            pid_running = get_pid_for_program(pid) is not None
            prog_status[pid] = {
                "running": pid_running,
                "scans": cached_data["scans"].get(pid, []),
                "log_tail": cached_data["log_tail"].get(pid, []),
                "scan_summary": cached_data["scan_summary"].get(pid, {"anchors": {}, "abc_matches": {}})
            }
        cfg = load_config()
        return jsonify({
            "programs": prog_status,
            "positions": cached_data["positions"],
            "all_trades": cached_data["all_trades"],
            "kite_positions": cached_data["kite_positions"],
            "ltp": cached_data["ltp"],
            "journal": cached_data["journal"],
            "stats": cached_data["stats"],
            "config": cfg,
            "pending_trades": cached_data.get("pending_trades", []),
            "backtest_results": cached_data.get("backtest_results", {}),
            "scan_display_data": cached_data.get("scan_display_data", {}),
            "live_execution": cached_data.get("live_execution", {"nifty50": False})
        })

@app.route("/api/token/check")
def api_token_check():
    return jsonify(check_token_valid())

@app.route("/api/token/url")
def api_token_url():
    return jsonify({"url": get_login_url()})

@app.route("/api/token/exchange", methods=["POST"])
def api_token_exchange():
    data = request.get_json(force=True, silent=True)
    if not data or not data.get("request_token"):
        return jsonify({"ok": False, "error": "No request_token provided"})
    result = exchange_request_token(data["request_token"].strip())
    return jsonify(result)

@app.route("/api/backtest/mode", methods=["GET", "POST"])
def api_backtest_mode():
    if request.method == "POST":
        data = request.get_json(force=True, silent=True)
        enabled = bool(data.get("enabled", False))
        set_backtest_mode(enabled)
    return jsonify({"enabled": get_backtest_mode()})

@app.route("/api/scanner/config", methods=["GET", "POST"])
def api_scanner_config():
    cfg_file = os.path.join("output", "monitor", "scanner_config.json")
    if request.method == "POST":
        data = request.get_json(force=True, silent=True) or {}
        os.makedirs(os.path.dirname(cfg_file), exist_ok=True)
        with open(cfg_file, "w") as f:
            json.dump({"disable_nopa": bool(data.get("disable_nopa", False))}, f)
        return jsonify({"ok": True})
    cfg = {}
    try:
        with open(cfg_file) as f:
            cfg = json.load(f)
    except Exception:
        cfg = {"disable_nopa": False}
    return jsonify(cfg)

@app.route("/api/live-execution/nifty50", methods=["GET", "POST"])
def api_live_execution_nifty50():
    flag = os.path.join("input", "nifty50_live.flag")
    if request.method == "POST":
        data = request.get_json(force=True, silent=True) or {}
        enabled = bool(data.get("enabled", False))
        try:
            if enabled:
                os.makedirs("input", exist_ok=True)
                open(flag, "w").close()
            elif os.path.exists(flag):
                os.remove(flag)
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)})
        return jsonify({"ok": True, "enabled": enabled})
    return jsonify({"enabled": os.path.exists(flag)})

@app.route("/api/config/<prog_id>", methods=["POST"])
def api_save_config(prog_id):
    if prog_id not in PROGRAMS:
        return jsonify({"ok": False, "error": "Unknown program"})
    data = request.get_json(force=True, silent=True)
    if not data:
        return jsonify({"ok": False, "error": "Invalid JSON"})
    save_config(prog_id, data)
    return jsonify({"ok": True})

@app.route("/api/programs/<prog_id>/start", methods=["POST"])
def api_start(prog_id):
    if prog_id not in PROGRAMS:
        return jsonify({"ok": False, "error": "Unknown program"})
    token = check_token_valid()
    if not token["valid"]:
        return jsonify({"ok": False, "error": token["reason"]})
    ok = start_program(prog_id)
    return jsonify({"ok": ok, "error": None if ok else "Start failed"})

@app.route("/api/programs/<prog_id>/stop", methods=["POST"])
def api_stop(prog_id):
    if prog_id not in PROGRAMS:
        return jsonify({"ok": False, "error": "Unknown program"})
    ok = stop_program(prog_id)
    return jsonify({"ok": ok})

ANCHOR_SCAN_REQUEST_FILE = os.path.join("output", "monitor", "anchor_scan_request.txt")
ANCHOR_SCAN_STOP_FILE = os.path.join("output", "monitor", "anchor_scan_stop.txt")

@app.route("/api/anchor/scan", methods=["POST"])
def api_anchor_scan():
    data = request.get_json(silent=True) or {}
    engine = data.get("engine", "index")
    try:
        with data_lock:
            cached_data["anchor_status"]["running"] = True
            cached_data["anchor_status"]["engine"] = engine
            cached_data["anchor_status"]["requested_at"] = time.time()
        if os.path.exists(ANCHOR_SCAN_STOP_FILE):
            os.remove(ANCHOR_SCAN_STOP_FILE)
        # Launch a dedicated --anchor-only subprocess. We intentionally do NOT
        # write ANCHOR_SCAN_REQUEST_FILE here: a running engine also polls that
        # file in its main loop, which would cause the anchor scan to run twice.
        script = PROGRAMS.get(engine, {}).get("file")
        if script:
            script_path = os.path.join(BASE_DIR, script)
            subprocess.Popen([sys.executable, script_path, "--anchor-only"],
                             cwd=BASE_DIR, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/anchor/stop", methods=["POST"])
def api_anchor_stop():
    try:
        os.makedirs(os.path.dirname(ANCHOR_SCAN_STOP_FILE), exist_ok=True)
        with open(ANCHOR_SCAN_STOP_FILE, "w") as f:
            f.write("stop")
        with data_lock:
            cached_data["anchor_status"]["running"] = False
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/anchor/status")
def api_anchor_status():
    with data_lock:
        st = dict(cached_data["anchor_status"])
    if not st.get("running"):
        still_running = os.path.exists(ANCHOR_SCAN_REQUEST_FILE) and not os.path.exists(ANCHOR_SCAN_STOP_FILE)
        if still_running:
            st["running"] = True
            if not st.get("engine"):
                try:
                    with open(ANCHOR_SCAN_REQUEST_FILE) as f:
                        st["engine"] = f.read().strip()
                except Exception:
                    pass
    return jsonify(st)

JOURNAL_FILE = "output/monitor/trade_journal.csv"
journal_lock = threading.Lock()

def _append_journal(symbol, pattern, action, status, details="", entry="", sl="", target="", rr=""):
    """Write a single row to trade_journal.csv (used by manual trade execute)."""
    file_exists = os.path.exists(JOURNAL_FILE)
    headers = ["Timestamp", "Symbol", "Pattern", "Timeframe", "Action", "Status", "Entry", "SL", "Target", "RR", "Details", "P&L %"]
    row = [dt.now().strftime("%Y-%m-%d %H:%M:%S"), symbol, pattern, "", action, status,
           f"{entry:.2f}" if isinstance(entry, (int, float)) and entry else str(entry) if entry else "",
           f"{sl:.2f}" if isinstance(sl, (int, float)) and sl else str(sl) if sl else "",
           f"{target:.2f}" if isinstance(target, (int, float)) and target else str(target) if target else "",
           f"{rr:.2f}" if isinstance(rr, (int, float)) and rr else str(rr) if rr else "",
           details, "-"]
    with journal_lock:
        try:
            with open(JOURNAL_FILE, mode="a", newline="", encoding="utf-8") as f:
                w = csv.writer(f, delimiter="\t")
                if not file_exists:
                    w.writerow(headers)
                w.writerow(row)
        except Exception as e:
            logging.error(f"Journal write error: {e}")

@app.route("/api/trade/close", methods=["POST"])
def api_trade_close():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"ok": False, "error": "No data"})
    symbol = data.get("symbol", "")
    token = data.get("token", "")
    engine = data.get("engine", "")
    if not symbol:
        return jsonify({"ok": False, "error": "symbol required"})
    try:
        active = trade_db.get_active_trades(engine)
        trade = None
        for t in active:
            if t.get("symbol") == symbol or t.get("contract", "").startswith(symbol):
                trade = t
                break
        if not trade and not _kite_session:
            return jsonify({"ok": False, "error": "No active trade found and Kite session unavailable"})
        if trade:
            contract = trade.get("contract", "")
            if contract and _kite_session:
                try:
                    nfo_key = f"NFO:{contract}"
                    oid = _kite_session.place_order(
                        variety=_kite_session.VARIETY_REGULAR,
                        tradingsymbol=contract,
                        exchange=_kite_session.EXCHANGE_NFO,
                        transaction_type=_kite_session.TRANSACTION_TYPE_SELL,
                        quantity=trade.get("position_size", 1) * trade.get("lot_size", 1),
                        order_type=_kite_session.ORDER_TYPE_MARKET,
                        product=_kite_session.PRODUCT_NRML
                    )
                    trade_db.update_trade(trade["id"], {"status": "CLOSED", "exit_time": dt.now().strftime("%Y-%m-%d %H:%M:%S")})
                    log_to_journal(symbol, trade.get("pattern", ""), "close", "EXIT_CLOSE", "SUCCESS",
                                   f"Manual close via dashboard, order: {oid}")
                except Exception as oe:
                    return jsonify({"ok": False, "error": f"Kite order failed: {oe}"})
            else:
                trade_db.update_trade(trade["id"], {"status": "CLOSED", "exit_time": dt.now().strftime("%Y-%m-%d %H:%M:%S")})
                log_to_journal(symbol, trade.get("pattern", ""), "close", "EXIT_CLOSE", "SUCCESS", "Manual close (backtest)")
        return jsonify({"ok": True, "msg": "Position closed"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/trade/execute", methods=["POST"])
def api_trade_execute():
    """Execute a manual trade from the Best Trade to Take tab — places Kite order + logs."""
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"ok": False, "error": "No data"})
    symbol = data.get("symbol", "")
    side = data.get("side", "CE")
    order_type = data.get("order_type", "BUY").upper()
    strike = data.get("strike", 0)
    entry = data.get("entry", 0)
    sl = data.get("sl", 0)
    t1 = data.get("t1", 0)
    t2 = data.get("t2", 0)
    t3 = data.get("t3", 0)
    pattern = data.get("pattern", "")
    rr = data.get("rr", 0)
    if not symbol or not pattern:
        return jsonify({"ok": False, "error": "symbol and pattern required"})
    try:
        if not _kite_session:
            return jsonify({"ok": False, "error": "Kite session not available. Ensure token is valid."})
        # Build a contract name roughly
        if strike:
            contract = f"{symbol}{dt.now().strftime('%y%b').upper()}{strike}{side}"
        else:
            contract = symbol
        tx_type = _kite_session.TRANSACTION_TYPE_BUY if order_type == "BUY" else _kite_session.TRANSACTION_TYPE_SELL
        try:
            oid = _kite_session.place_order(
                variety=_kite_session.VARIETY_REGULAR,
                tradingsymbol=contract,
                exchange=_kite_session.EXCHANGE_NFO if strike else _kite_session.EXCHANGE_NSE,
                transaction_type=tx_type,
                quantity=1,
                order_type=_kite_session.ORDER_TYPE_MARKET,
                product=_kite_session.PRODUCT_MIS
            )
        except Exception as oe:
            return jsonify({"ok": False, "error": f"Order failed: {oe}"})
        # Log to journal
        details = f"Manual {order_type} | Contract: {contract} | Strike: {strike} | Side: {side}"
        _append_journal(symbol, pattern, f"MANUAL_{order_type}", "SUCCESS", details,
                        entry=entry, sl=sl, target=t3 or t1, rr=rr)
        # Record in trade_db
        trade_data = {
            "symbol": symbol, "contract": contract, "pattern": pattern,
            "side": side, "strike": strike,
            "entry_spot": entry, "current_sl": sl, "t1": t1, "t2": t2, "t3": t3,
            "rr": rr, "trailing_stage": 0, "lot_size": 1, "position_size": 1,
            "timeframe": "manual",
            "entry_time": dt.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        tid = trade_db.create_trade(side.lower() if side else "manual", symbol, trade_data)
        engine_key = f"{symbol}|{pattern}|{side}|{strike}"
        trade_db.record_executed_pattern("manual", engine_key, {"contract": contract, "entry": entry})
        return jsonify({"ok": True, "order_id": oid if not isinstance(oid, Exception) else "", "trade_id": tid})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

# ──────────────────────────────────────────────
#  SINGLE ENTRY ANALYSIS
# ──────────────────────────────────────────────

def _resolve_token(symbol):
    """Resolve instrument token for a symbol using Kite LTP lookup."""
    if not _kite_session:
        return None
    nse_name = {"NIFTY":"NSE:NIFTY 50","BANKNIFTY":"NSE:NIFTY BANK","SENSEX":"NSE:SENSEX"}.get(symbol, f"NSE:{symbol}")
    try:
        q = _kite_session.ltp([nse_name])
        return int(list(q.keys())[0])
    except:
        try:
            q = _kite_session.ltp([f"NSE:{symbol}"])
            return int(list(q.keys())[0])
        except:
            return None

def _no_pa_left_simple(df, a_idx, ref_price, n=100, bullish=True):
    if a_idx < n: return True
    left = df.iloc[a_idx - n : a_idx]
    lr = (float(left['high'].max()) - float(left['low'].min())) / float(left['close'].mean())
    if lr >= 0.02: return False
    if bullish:
        return float(left['close'].max()) <= ref_price
    else:
        return float(left['close'].min()) >= ref_price

def _find_bcd_forward(df, a_idx, benchmark, pattern_type='bull'):
    n = len(df)
    fb = fc = fd = None
    if pattern_type == 'bull':
        for i in range(a_idx + 1, n):
            if float(df.iloc[i]['close']) > benchmark: fb = i; break
        if fb is None: return None
        for i in range(fb + 1, n):
            r = df.iloc[i]
            if float(r['close']) < float(r['open']) and float(r['low']) <= benchmark: fc = i; break
        if fc is None: return None
        for i in range(fc + 1, n):
            r = df.iloc[i]
            if float(r['close']) > float(r['open']) and float(r['close']) > benchmark: fd = i; break
        if fd is None: return None
    else:
        for i in range(a_idx + 1, n):
            if float(df.iloc[i]['close']) < benchmark: fb = i; break
        if fb is None: return None
        for i in range(fb + 1, n):
            r = df.iloc[i]
            if float(r['close']) > float(r['open']) and float(r['high']) >= benchmark: fc = i; break
        if fc is None: return None
        for i in range(fc + 1, n):
            r = df.iloc[i]
            if float(r['close']) < float(r['open']) and float(r['close']) < benchmark: fd = i; break
        if fd is None: return None
    return {'close': float(df.iloc[fd]['close']), 'd_idx': fd}

def _run_bcd_inline(df, a_idx, benchmark, sl, ref_price, pattern_type):
    a = df.iloc[a_idx]
    if not _no_pa_left_simple(df, a_idx, ref_price, 100, pattern_type == 'bull'):
        return None
    bcd = _find_bcd_forward(df, a_idx, benchmark, pattern_type)
    if bcd is None:
        return None
    cp = bcd['close']
    if pattern_type == 'bull':
        risk = cp - sl
        if risk <= 0 or risk < cp * 0.002: return None
        t1 = round(cp + risk * 2, 2)
    else:
        risk = sl - cp
        if risk <= 0 or risk < cp * 0.002: return None
        t1 = round(cp - risk * 2, 2)
    rr = round(2.0, 2)
    return {"Close": round(cp, 2), "SL": round(sl, 2), "T1": t1, "RR": rr}

def _find_swing_lows(df, window=5):
    indices = []
    for i in range(window, len(df) - window):
        seg = df.iloc[i-window:i+window+1]
        if float(df.iloc[i]['low']) == float(seg['low'].min()):
            if i > 0 and float(df.iloc[i]['low']) < float(df.iloc[i-1]['low']):
                indices.append(i)
    return indices

def _find_swing_highs(df, window=5):
    indices = []
    for i in range(window, len(df) - window):
        seg = df.iloc[i-window:i+window+1]
        if float(df.iloc[i]['high']) == float(seg['high'].max()):
            if i > 0 and float(df.iloc[i]['high']) > float(df.iloc[i-1]['high']):
                indices.append(i)
    return indices

def _find_pin_bars(df, pattern_type='bull'):
    results = []
    for i in range(1, len(df) - 1):
        row = df.iloc[i]
        body = abs(float(row['close']) - float(row['open']))
        lower = min(float(row['close']), float(row['open'])) - float(row['low'])
        upper = float(row['high']) - max(float(row['close']), float(row['open']))
        if body == 0: continue
        if pattern_type == 'bull':
            if float(row['close']) > float(row['open']) and lower >= 2 * body and upper < body:
                results.append(i)
        else:
            if float(row['close']) < float(row['open']) and upper >= 2 * body and lower < body:
                results.append(i)
    return results

def _find_double_bottom(df):
    lows = _find_swing_lows(df)
    if len(lows) < 2: return None
    for j in range(len(lows) - 1, 0, -1):
        l2 = lows[j]; l1 = lows[j - 1]
        if float(df.iloc[l2]['low']) >= float(df.iloc[l1]['low']): continue
        if l2 - l1 <= 2: continue
        btwn = df.iloc[l1 + 1 : l2]
        if btwn.empty or float(btwn['high'].max()) <= float(df.iloc[l1]['high']): continue
        if l2 >= len(df) - 1: continue
        if float(df.iloc[l2 + 1]['low']) < float(df.iloc[l2]['low']): continue
        return {'a_idx': l2, 'benchmark': float(df.iloc[l2]['high'])}
    return None

def _find_double_top(df):
    highs = _find_swing_highs(df)
    if len(highs) < 2: return None
    for j in range(len(highs) - 1, 0, -1):
        h2 = highs[j]; h1 = highs[j - 1]
        if float(df.iloc[h2]['high']) <= float(df.iloc[h1]['high']): continue
        if h2 - h1 <= 2: continue
        btwn = df.iloc[h1 + 1 : h2]
        if btwn.empty or float(btwn['low'].min()) >= float(df.iloc[h1]['low']): continue
        if h2 >= len(df) - 1: continue
        if float(df.iloc[h2 + 1]['high']) > float(df.iloc[h2]['high']): continue
        return {'a_idx': h2, 'benchmark': float(df.iloc[h2]['low'])}
    return None

def _scan_bcd_reversal(df):
    if len(df) < 10: return None
    swing_lows = _find_swing_lows(df)
    for a_idx in reversed(swing_lows):
        a = df.iloc[a_idx]
        res = _run_bcd_inline(df, a_idx, float(a['high']), float(a['low']) + 2, float(a['open']), 'bull')
        if res: return {**res, "Pattern": "BULL_BCD"}
    return None

def _scan_pinbar_reversal(df):
    if len(df) < 10: return None
    pins = _find_pin_bars(df, 'bull')
    for a_idx in reversed(pins):
        a = df.iloc[a_idx]
        res = _run_bcd_inline(df, a_idx, float(a['high']), float(a['low']) + 2, float(a['open']), 'bull')
        if res: return {**res, "Pattern": "BULL_PINBAR"}
    return None

def _scan_swing_reversal(df):
    if len(df) < 10: return None
    sw = _find_double_bottom(df)
    if sw is None: return None
    a = df.iloc[sw['a_idx']]
    res = _run_bcd_inline(df, sw['a_idx'], float(a['high']), float(a['low']) + 2, float(a['open']), 'bull')
    if res: return {**res, "Pattern": "BULL_SWING"}
    return None

def _scan_bcd_bearish_reversal(df):
    if len(df) < 10: return None
    swing_highs = _find_swing_highs(df)
    for a_idx in reversed(swing_highs):
        a = df.iloc[a_idx]
        res = _run_bcd_inline(df, a_idx, float(a['low']), float(a['high']) + 2, float(a['close']), 'bear')
        if res: return {**res, "Pattern": "BEAR_BCD"}
    return None

def _scan_pinbar_bearish_reversal(df):
    if len(df) < 10: return None
    pins = _find_pin_bars(df, 'bear')
    for a_idx in reversed(pins):
        a = df.iloc[a_idx]
        res = _run_bcd_inline(df, a_idx, float(a['low']), float(a['high']) + 2, float(a['close']), 'bear')
        if res: return {**res, "Pattern": "BEAR_PINBAR"}
    return None

def _scan_swing_bearish_reversal(df):
    if len(df) < 10: return None
    sw = _find_double_top(df)
    if sw is None: return None
    a = df.iloc[sw['a_idx']]
    res = _run_bcd_inline(df, sw['a_idx'], float(a['low']), float(a['high']) + 2, float(a['close']), 'bear')
    if res: return {**res, "Pattern": "BEAR_SWING"}
    return None

def _scan_anchors(df):
    results = []
    if len(df) < 5: return results
    # Engulfing
    bu, be = df.iloc[-4], df.iloc[-3]
    if float(bu['close']) > float(bu['open']) and float(be['close']) < float(be['open']) and float(be['open']) >= float(bu['close']) and float(be['close']) < float(bu['low']):
        results.append({"Pattern":"BEAR_A_Engulf","Close":round(float(be['close']),2),"SL":round(float(be['high'])+2,2)})
    if float(bu['close']) < float(bu['open']) and float(be['close']) > float(be['open']) and float(be['open']) <= float(bu['close']) and float(be['close']) > float(bu['high']):
        results.append({"Pattern":"BULL_A_Engulf","Close":round(float(be['close']),2),"SL":round(float(be['low'])-2,2)})
    # HH Sweep / LL Sweep
    if len(df) >= 30:
        lr = df.iloc[-29:-4]
        h1 = float(lr['high'].max()); l1 = float(lr['low'].min())
        sc, dc, c1, c2 = df.iloc[-4], df.iloc[-3], df.iloc[-2], df.iloc[-1]
        if float(sc['close']) > float(sc['open']) and (float(sc['high']) > h1 and float(sc['close']) < h1 or float(sc['close']) > h1 and float(dc['close']) < h1):
            if float(dc['close']) < float(sc['low']) and float(c1['close']) <= float(sc['high']) and float(c2['close']) <= float(sc['high']):
                results.append({"Pattern":"BEAR_A_HH_Sweep","Close":round(float(dc['close']),2),"SL":round(float(sc['high'])+2,2)})
        if float(sc['close']) < float(sc['open']) and (float(sc['low']) < l1 and float(sc['close']) > l1 or float(sc['close']) < l1 and float(dc['close']) > l1):
            if float(dc['close']) > float(sc['high']) and float(c1['close']) >= float(sc['low']) and float(c2['close']) >= float(sc['low']):
                results.append({"Pattern":"BULL_A_LL_Sweep","Close":round(float(dc['close']),2),"SL":round(float(sc['low'])-2,2)})
    # Shooting Star / Hammer
    mc = df.iloc[-5]; bc = df.iloc[-4]; pb1, pb2, pb3 = df.iloc[-3], df.iloc[-2], df.iloc[-1]
    if float(mc['close']) > float(mc['open']) and float(bc['close']) < float(bc['open']):
        body = float(bc['open'])-float(bc['close']); uw = float(bc['high'])-float(bc['open'])
        if uw > body and float(bc['high']) <= float(mc['close']) and float(bc['low']) >= float(mc['open']):
            if float(pb2['close']) <= float(bc['high']) and float(pb3['close']) <= float(bc['high']):
                results.append({"Pattern":"BEAR_A_ShootingStar","Close":round(float(bc['close']),2),"SL":round(float(bc['high'])+2,2)})
    if float(mc['close']) < float(mc['open']) and float(bc['close']) > float(bc['open']):
        body = float(bc['close'])-float(bc['open']); lw = float(bc['open'])-float(bc['low'])
        if lw > body and float(bc['high']) <= float(mc['open']) and float(bc['low']) >= float(mc['close']):
            if float(pb2['close']) >= float(bc['low']) and float(pb3['close']) >= float(bc['low']):
                results.append({"Pattern":"BULL_A_Hammer","Close":round(float(bc['close']),2),"SL":round(float(bc['low'])-2,2)})
    # Harami
    mom = df.iloc[-5]; child = df.iloc[-4]; ph1, ph2, ph3 = df.iloc[-3], df.iloc[-2], df.iloc[-1]
    if float(mom['close']) > float(mom['open']) and float(child['close']) < float(child['open']) and float(child['high']) <= float(mom['close']) and float(child['low']) >= float(mom['open']):
        if float(ph2['close']) <= float(child['high']) and float(ph3['close']) <= float(child['high']):
            results.append({"Pattern":"BEAR_A_Harami","Close":round(float(child['close']),2),"SL":round(float(child['high'])+2,2)})
    if float(mom['close']) < float(mom['open']) and float(child['close']) > float(child['open']) and float(child['high']) <= float(mom['open']) and float(child['low']) >= float(mom['close']):
        if float(ph2['close']) >= float(child['low']) and float(ph3['close']) >= float(child['low']):
            results.append({"Pattern":"BULL_A_Harami","Close":round(float(child['close']),2),"SL":round(float(child['low'])-2,2)})
    return results

@app.route("/api/analyze/entry", methods=["POST"])
def api_analyze_entry():
    data = request.get_json(silent=True)
    if not data:
        return jsonify({"ok": False, "error": "No data"})
    raw_symbol = data.get("symbol","").upper().strip()
    timeframe = data.get("timeframe","15minute")
    if not raw_symbol:
        return jsonify({"ok": False, "error": "symbol required"})
    if not _kite_session:
        return jsonify({"ok": False, "error": "Kite session unavailable"})
    try:
        token = None
        exchange = "NSE"
        symbol = raw_symbol
        opt_type = "CE"
        # Try NFO first (full contract like NESTLEIND26JUL1460CE)
        try:
            q = _kite_session.ltp([f"NFO:{raw_symbol}"])
            token = int(list(q.keys())[0])
            exchange = "NFO"
            if raw_symbol.endswith("PE"):
                opt_type = "PE"
        except Exception:
            pass
        # If NFO failed, try NSE — strip contract suffix if present
        if not token:
            base = raw_symbol
            m = re.match(r'^([A-Z]+?)(\d{2}[A-Z]{3})(\d+)(CE|PE)$', raw_symbol)
            if m:
                base = m.group(1)
                opt_type = m.group(4)
            try:
                q = _kite_session.ltp([f"NSE:{base}"])
                token = int(list(q.keys())[0])
                symbol = base
            except Exception:
                # Try common index mappings
                nse_map = {"NIFTY":"NSE:NIFTY 50","BANKNIFTY":"NSE:NIFTY BANK","SENSEX":"NSE:SENSEX"}
                mapped = nse_map.get(base)
                if mapped:
                    try:
                        q = _kite_session.ltp([mapped])
                        token = int(list(q.keys())[0])
                    except Exception:
                        pass
        if not token:
            return jsonify({"ok": False, "error": f"Cannot resolve token for {raw_symbol}"})

        to_dt = dt.now().strftime("%Y-%m-%d")
        limits = {"minute":60,"3minute":100,"5minute":100,"10minute":100,"15minute":200,"30minute":200,"60minute":400,"day":2000}
        maxd = limits.get(timeframe, 60)
        from_dt = (dt.now() - timedelta(days=maxd)).strftime("%Y-%m-%d")
        for attempt in range(3):
            try:
                raw = _kite_session.historical_data(token, from_dt, to_dt, timeframe)
                break
            except Exception as e:
                if "Too many requests" in str(e) and attempt < 2:
                    time.sleep(5)
                    continue
                raise
        df = pd.DataFrame(raw)
        if df.empty:
            return jsonify({"ok": False, "error": f"No historical data for {raw_symbol}"})
        # Run scanners based on option type direction
        if opt_type == "PE":
            bcd = _scan_bcd_bearish_reversal(df)
            pinbar = _scan_pinbar_bearish_reversal(df)
            swing = _scan_swing_bearish_reversal(df)
        else:
            bcd = _scan_bcd_reversal(df)
            pinbar = _scan_pinbar_reversal(df)
            swing = _scan_swing_reversal(df)
        anchors = _scan_anchors(df)
        label = f"{raw_symbol} ({exchange})"
        return jsonify({"ok": True, "label": label, "option_type": opt_type, "timeframe": timeframe,
                        "bcd": bcd, "pinbar": pinbar, "swing": swing, "anchors": anchors, "candles": len(df)})
    except Exception as e:
        logging.error(f"Analyze error: {e}")
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/journal/clear", methods=["POST"])
def api_journal_clear():
    try:
        if os.path.exists(JOURNAL_FILE):
            headers = ["Timestamp", "Symbol", "Pattern", "Timeframe", "Action", "Status", "Entry", "SL", "Target", "RR", "Details", "P&L %"]
            with open(JOURNAL_FILE, mode="w", newline="", encoding="utf-8") as f:
                csv.writer(f, delimiter="\t").writerow(headers)
        for f in glob("output/monitor/backtest_results_*.json"):
            try: os.remove(f)
            except: pass
        trade_db.clear_executed_patterns()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/logs/clear", methods=["POST"])
def api_logs_clear():
    log_files = [INDEX_LOG_FILE, NIFTY50_LOG_FILE, DAILY_LOG_FILE, BEAR_INDEX_LOG_FILE, BEAR_NIFTY50_LOG_FILE, BEAR_DAILY_LOG_FILE]
    for lf in log_files:
        try:
            if os.path.exists(lf):
                open(lf, "w").close()
        except Exception:
            pass
    return jsonify({"ok": True})

@app.route("/api/trades")
def api_trades():
    engine = request.args.get("engine")
    active_only = request.args.get("active", "false").lower() == "true"
    if active_only:
        return jsonify(trade_db.get_active_trades(engine))
    return jsonify(trade_db.get_all_trades(engine))

@app.route("/api/export/monthly", methods=["POST"])
def api_export_monthly():
    result = run_monthly_export()
    return jsonify({"ok": True, **result})

EXPORT_STATE_FILE = "output/monitor/export_state.json"

# ──────────────────────────────────────────────
#  MONTHLY EXPORT (trades to Excel archive)
# ──────────────────────────────────────────────

def run_monthly_export():
    import openpyxl
    from collections import defaultdict
    completed = trade_db.get_completed_trades()
    if not completed:
        return {"exported": 0, "sheets": []}
    groups = defaultdict(list)
    for t in completed:
        ts = t.get("exit_time") or t.get("updated_at") or t.get("created_at") or ""
        parts = ts.split(" ")[0].split("-") if " " in ts else ts.split("-")
        key = (parts[0], parts[1]) if len(parts) >= 2 else ("unknown", "00")
        groups[key].append(t)
    out_dir = "output/exports"
    os.makedirs(out_dir, exist_ok=True)
    xl_path = os.path.join(out_dir, "trade_archive.xlsx")
    sheet_names = []
    if os.path.exists(xl_path):
        wb = openpyxl.load_workbook(xl_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    headers = ["ID", "Engine", "Symbol", "Pattern", "Status", "Entry Spot", "SL", "T1", "T2", "T3",
               "Trailing Stage", "Lot Size", "Position Size", "Entry Date", "Exit Date", "P&L %", "Side", "Contract"]
    for (year, month), trades in sorted(groups.items()):
        month_names = ["", "January","February","March","April","May","June","July","August","September","October","November","December"]
        sheet_name = f"{month_names[int(month)]} {year}" if month.isdigit() and 1 <= int(month) <= 12 else f"{month} {year}"
        sheet_names.append(sheet_name)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for t in trades:
            entry_dt = (t.get("created_at") or "").split(" ")[0]
            exit_dt = (t.get("exit_time") or t.get("updated_at") or "").split(" ")[0]
            ws.append([
                t.get("id", ""), t.get("engine", ""), t.get("symbol", ""), t.get("pattern", ""),
                t.get("status", ""), t.get("entry_spot", ""), t.get("current_sl", ""),
                t.get("t1", ""), t.get("t2", ""), t.get("t3", ""),
                t.get("trailing_stage", ""), t.get("lot_size", ""), t.get("position_size", ""),
                entry_dt, exit_dt,
                t.get("pnl_percent", ""), t.get("side", ""), t.get("contract", "")
            ])
    wb.save(xl_path)
    exported_ids = [t["id"] for t in completed]
    trade_db.remove_trades(exported_ids)
    return {"exported": len(completed), "sheets": sheet_names}

def auto_export_if_new_month():
    now = dt.now()
    current_month = now.strftime("%Y-%m")
    try:
        if os.path.exists(EXPORT_STATE_FILE):
            with open(EXPORT_STATE_FILE) as f:
                state = json.load(f)
                last = state.get("last_export_month", "")
        else:
            last = ""
        if current_month > last:
            result = run_monthly_export()
            if result["exported"] > 0:
                print(f"Auto-export: {result['exported']} trades to {', '.join(result['sheets'])}")
            with open(EXPORT_STATE_FILE, "w") as f:
                json.dump({"last_export_month": current_month}, f)
    except Exception as e:
        print(f"Auto-export error: {e}")

class _ProcessRef:
    """Lightweight wrapper to track a recovered PID like a Popen object."""
    def __init__(self, pid):
        self.pid = pid
        self._returncode = None
    def poll(self):
        if self._returncode is not None:
            return self._returncode
        if os.name == "nt":
            import ctypes
            k32 = ctypes.windll.kernel32
            h = k32.OpenProcess(0x100000, False, self.pid)
            if h:
                k32.CloseHandle(h)
                return None
            self._returncode = 0
            return 0
        try:
            os.kill(self.pid, 0)
            return None
        except OSError:
            self._returncode = 0
            return 0

def main():
    os.makedirs("input", exist_ok=True)
    os.makedirs("output/logs", exist_ok=True)
    os.makedirs("output/monitor", exist_ok=True)
    os.makedirs("logs", exist_ok=True)
    auto_export_if_new_month()
    orphaned = get_running_engines()
    if orphaned:
        print(f"Recovered running engines from PID files: {orphaned}")
        for eid, pid in orphaned.items():
            if eid in PROGRAMS:
                processes[eid] = _ProcessRef(pid)
    worker = threading.Thread(target=refresh_data, daemon=True)
    worker.start()
    print(f"Trading Control Center starting on http://localhost:{DASHBOARD_PORT}")
    print(f"Refresh interval: {REFRESH_SECONDS}s")
    print("Available programs:")
    for pid, p in PROGRAMS.items():
        print(f"  [{pid}] {p['name']}")
    app.run(host="0.0.0.0", port=DASHBOARD_PORT, debug=False, use_reloader=False)

if __name__ == "__main__":
    main()
