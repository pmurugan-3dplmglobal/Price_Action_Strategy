import os, json, csv, time, threading, subprocess, sys, signal, logging
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
COMMON_DIR = os.path.join(BASE_DIR, "common")
for p in [BASE_DIR, COMMON_DIR]:
    if p not in sys.path:
        sys.path.insert(0, p)
import paths
from datetime import datetime as dt, time as datetime_time
from flask import Flask, render_template_string, jsonify, request, Response
from kiteconnect import KiteConnect
import trade_db
from dashboard_sl_overrides import write_sl_overrides
from trading_core import (
    lookup_scan_sl_target,
    derive_sl_targets_for_contract,
    load_kite_session,
    close_position as shared_close_position,
    close_stock_position as shared_close_stock_position,
    clear_executed_exit,
    log_to_journal
)
from ema_engine import (
    start_ema_engine, stop_ema_engine, get_ema_engine_status, get_ema_scan_data,
    EMA_DISPLAY_FILE_OPTION
)

def resolve_underlying(contract_or_symbol, engine="nifty50"):
    """Return the real underlying registry symbol for a contract string.

    Fixes the stale-ACTIVE anomaly where the DB `symbol` was stored as the full
    contract string (e.g. NIFTY2681124650PE) instead of the underlying (NIFTY).
    Falls back to the raw input when nothing matches.
    """
    from trading_core import INDEX_REGISTRY, STOCK_REGISTRY
    raw = str(contract_or_symbol or "").replace(" ", "").upper()
    if not raw:
        return contract_or_symbol or ""
    reg = INDEX_REGISTRY if engine == "index" else STOCK_REGISTRY
    for sym in sorted(reg.keys(), key=len, reverse=True):
        if sym.replace(" ", "").upper() in raw:
            return sym
    for sym in sorted(STOCK_REGISTRY.keys(), key=len, reverse=True):
        if sym.replace(" ", "").upper() in raw:
            return sym
    return contract_or_symbol or ""

app = Flask(__name__)

# ──────────────────────────────────────────────
#  FILE PATHS & DASHBOARD CONFIG
# ──────────────────────────────────────────────


def get_kite_credentials():
    """Read Kite API key/secret from environment, config, or token file."""
    api_key = os.environ.get("KITE_API_KEY", "")
    api_secret = os.environ.get("KITE_API_SECRET", "")
    cfg = load_config()
    if not api_key: api_key = cfg.get("api_key", "")
    if not api_secret: api_secret = cfg.get("api_secret", "")
    if (not api_key or not api_secret) and os.path.exists(TOKEN_FILE):
        try:
            with open(TOKEN_FILE) as f:
                td = json.load(f)
            if not api_key: api_key = td.get("api_key", "")
            if not api_secret: api_secret = td.get("api_secret", "")
        except Exception:
            pass
    if not api_secret:
        opt_cfg = os.path.join(BASE_DIR, "input", "program_config.json")
        if os.path.exists(opt_cfg):
            try:
                with open(opt_cfg) as f:
                    cdata = json.load(f)
                if not api_key: api_key = cdata.get("api_key", "")
                if not api_secret: api_secret = cdata.get("api_secret", "")
            except Exception:
                pass
    if not api_key or not api_secret:
        logging.warning("api_key/api_secret missing in program_config.json")
    return api_key, api_secret


BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
TOKEN_FILE = paths.TOKEN_FILE
CONFIG_FILE = paths.PROGRAM_CONFIG_FILE
SCAN_DISPLAY_FILE = paths.SCAN_DISPLAY_FILE
SCAN_DISPLAY_INDEX_FILE = paths.SCAN_DISPLAY_INDEX_FILE
POSITIONS_FILE = os.path.join(BASE_DIR, "output", "monitor", "positions.json")
JOURNAL_FILE = os.path.join(BASE_DIR, "output", "journal.json")
LOGS_DIR = os.path.join(BASE_DIR, "logs")

INDEX_LOG_FILE = paths.INDEX_LOG_FILE
NIFTY50_LOG_FILE = paths.NIFTY50_LOG_FILE
EMA_LOG_FILE = paths.EMA_LOG_FILE

LIVE_EXECUTION_FLAG = paths.NIFTY50_LIVE_FLAG
LIVE_EXECUTION_FLAG_INDEX = paths.INDEX_LIVE_FLAG
DASHBOARD_PORT = 5050

REFRESH_SECONDS = 1
ACTIVE_EDIT_LOCKS = set()

PROGRAMS = {
    "index": {
        "name": "Index Options Trade Engine",
        "file": "Trade_Option/index_options_trade_engine.py",
        "desc": "Real-time index options intraday trading (NIFTY, BANKNIFTY, SENSEX)",
        "color": "#58a6ff",
        "log_file": INDEX_LOG_FILE,
        "config_fields": {
            "timeframe_entry": {"label": "Entry Timeframe", "type": "select", "options": ["3minute","5minute","10minute","15minute","30minute","60minute","75min","day"], "default": "3minute"},
            "timeframe_anchor": {"label": "Anchor Timeframe", "type": "select", "options": ["3minute","5minute","10minute","15minute","30minute","60minute","75min","day"], "default": "15minute"},
            "lookback_days": {"label": "Lookback Days", "type": "number", "default": 30},
            "scan_interval": {"label": "Scan Interval (s)", "type": "number", "default": 15},
            "risk_percent": {"label": "Risk %", "type": "number", "default": 1.0},
            "capital": {"label": "Capital", "type": "number", "default": 100000.0},
            "strike_range": {"label": "Strike Range (±)", "type": "number", "default": 0}
        }
    },
    "nifty50": {
        "name": "Stock Options Trade Engine",
        "file": "Trade_Option/stock_options_trade_engine.py",
        "desc": "Scans Nifty 50 stock options, picks best setup & executes",
        "color": "#3fb950",
        "log_file": NIFTY50_LOG_FILE,
        "config_fields": {
            "timeframe_entry": {"label": "Entry Timeframe", "type": "select", "options": ["3minute","5minute","10minute","15minute","30minute","60minute","75min","day"], "default": "15minute"},
            "timeframe_anchor": {"label": "Anchor Timeframe", "type": "select", "options": ["3minute","5minute","10minute","15minute","30minute","60minute","75min","day"], "default": "30minute"},
            "lookback_days": {"label": "Lookback Days", "type": "number", "default": 30},
            "scan_interval": {"label": "Scan Interval (s)", "type": "number", "default": 300},
            "risk_percent": {"label": "Risk %", "type": "number", "default": 1.0},
            "capital": {"label": "Capital", "type": "number", "default": 100000.0},
            "strike_range": {"label": "Strike Range (±)", "type": "number", "default": 0}
        }
    },
    "ema_engine": {
        "name": "Stock EMA Engine",
        "file": "common/ema_engine.py",
        "desc": "Scans 13 EMA & 44 EMA crossovers on stock options",
        "color": "#a371f7",
        "log_file": EMA_LOG_FILE,
        "config_fields": {
            "timeframe": {"label": "Timeframe", "type": "select", "options": ["1d", "60minute", "30minute", "15minute", "5minute"], "default": "1d"},
            "target_universe": {"label": "Target Universe", "type": "select", "options": ["ALL", "NIFTY50", "NIFTY_NEXT_100", "NIFTY_MIDCAP_100", "NIFTY_SMALLCAP_250", "INDEX_OPTIONS"], "default": "ALL"},
            "scan_interval": {"label": "Scan Interval (s)", "type": "number", "default": 300}
        }
    }
}

processes = {}
process_lock = threading.Lock()

data_lock = threading.Lock()
cached_data = {
    "positions": {},
    "journal": [],
    "log_tail": {pid: [] for pid in PROGRAMS},
    "stats": {"total_trades": 0, "win_rate": 0, "active_positions": 0, "pnl": 0},
    "scans": {"index": [], "nifty50": []},
    "scan_summary": {"index": {"anchors": {}, "abc_matches": {}}, "nifty50": {"anchors": {}, "abc_matches": {}}},
    "all_trades": [],
    "kite_positions": [],
    "ltp": {},
    "anchor_status": {"running": False, "engine": None, "requested_at": None, "completed_at": None},
    "scan_display": {"date": "", "timestamp": "", "staged_trades": [], "active_positions": []},
    "live_execution": False,
    "live_execution_index": False,
    "executed_exits": {},
    "expired_contracts": []
}
_expired_cache_day = ""
_expired_cache_set = set()
_ltp_last_fetch = 0
_kite_positions_last_fetch = 0
_kite_session = None
_last_scan_reset = ""
_ltp_memory = {}
_pnl_memory = {}

# ──────────────────────────────────────────────
#  PROCESS MANAGEMENT (Start/Stop Programs)
# ──────────────────────────────────────────────

def get_pid_for_program(prog_id):
    with process_lock:
        p = processes.get(prog_id)
        if p and p.poll() is None:
            return p.pid
    return None

def start_program(prog_id):
    token = check_token_valid()
    if not token["valid"]:
        print(f"Cannot start {prog_id}: {token['reason']}")
        return False
    with process_lock:
        if prog_id in processes and processes[prog_id].poll() is None:
            return False
        rel_file = PROGRAMS[prog_id]["file"]
        script_path = os.path.join(BASE_DIR, rel_file)
        if not os.path.exists(script_path):
            script_path = os.path.join(os.path.dirname(__file__), os.path.basename(rel_file))
        try:
            p = subprocess.Popen(
                [sys.executable, "-u", script_path],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                cwd=BASE_DIR, creationflags=subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
            )
            processes[prog_id] = p
            return True
        except Exception as e:
            print(f"Failed to start {prog_id}: {e}")
            return False

def stop_program(prog_id):
    with process_lock:
        p = processes.get(prog_id)
        if p and p.poll() is None:
            if os.name == "nt":
                subprocess.run(["taskkill", "/F", "/T", "/PID", str(p.pid)], capture_output=True)
            else:
                os.kill(p.pid, signal.SIGTERM)
            processes.pop(prog_id, None)
    script_file = PROGRAMS.get(prog_id, {}).get("file")
    if script_file and os.name == "nt":
        try:
            ps_cmd = f"Get-CimInstance Win32_Process -Filter \"Name='python.exe'\" | Where-Object {{ $_.CommandLine -match '{script_file}' }} | ForEach-Object {{ Stop-Process -Id $_.ProcessId -Force }}"
            subprocess.run(["powershell.exe", "-Command", ps_cmd], capture_output=True)
        except Exception as e:
            logging.warning(f"System-wide kill for {prog_id} failed: {e}")
    return True

# ──────────────────────────────────────────────
#  CONFIGURATION (program_config.json)
# ──────────────────────────────────────────────

def load_config():
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE) as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def save_config(prog_id, data):
    cfg = load_config()
    cleaned = {}
    for k, v in data.items():
        if isinstance(v, str):
            try:
                if "." in v:
                    cleaned[k] = float(v)
                else:
                    cleaned[k] = int(v)
            except (ValueError, TypeError):
                cleaned[k] = v
        else:
            cleaned[k] = v
    cfg[prog_id] = cleaned
    os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)
    return True

def get_backtest_mode():
    return load_config().get("_backtest", False)

def set_backtest_mode(enabled):
    cfg = load_config()
    cfg["_backtest"] = enabled
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)

# ──────────────────────────────────────────────
#  KITE TOKEN MANAGEMENT
# ──────────────────────────────────────────────

def check_token_valid():
    if not os.path.exists(TOKEN_FILE):
        return {"valid": False, "reason": "Token file not found"}
    try:
        with open(TOKEN_FILE) as f:
            data = json.load(f)
        if not data.get("api_key") or not data.get("access_token"):
            return {"valid": False, "reason": "Invalid token file"}
        date_str = data.get("generated_at", "")
        if date_str:
            try:
                from datetime import datetime as dt2
                gen_dt = dt2.strptime(date_str, "%Y-%m-%d %H:%M:%S")
                now_dt = dt.now()
                gen_date = gen_dt.date()
                today = now_dt.date()
                if gen_date < today:
                    return {"valid": False, "reason": f"Token expired (generated {date_str})"}
                reset_cutoff = gen_dt.replace(hour=6, minute=0, second=0, microsecond=0)
                if gen_dt < reset_cutoff and now_dt >= reset_cutoff:
                    return {"valid": False, "reason": f"Token expired (generated {date_str} before 06:00 AM Zerodha reset)"}
            except Exception:
                try:
                    gen_date = dt.strptime(date_str.split()[0], "%Y-%m-%d").date()
                    if gen_date < dt.now().date():
                        return {"valid": False, "reason": f"Token expired (generated {date_str})"}
                except Exception:
                    pass
        return {"valid": True, "reason": "Token valid"}
    except Exception as e:
        return {"valid": False, "reason": f"Token read error: {e}"}

def get_login_url():
    api_key, _ = get_kite_credentials()
    if not api_key:
        return ""
    return f"https://kite.zerodha.com/connect/login?api_key={api_key}"

def exchange_request_token(request_token):
    """Exchange Kite request_token for access_token and save to file."""
    try:
        api_key, api_secret = get_kite_credentials()
        kite = KiteConnect(api_key=api_key)
        session = kite.generate_session(request_token, api_secret=api_secret)
        access_token = session["access_token"]
        token_data = {
            "api_key": api_key,
            "access_token": access_token,
            "generated_at": dt.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        os.makedirs(os.path.dirname(TOKEN_FILE), exist_ok=True)
        with open(TOKEN_FILE, "w") as f:
            json.dump(token_data, f, indent=4)
        return {"ok": True, "access_token": access_token[:8] + "..."}
    except Exception as e:
        return {"ok": False, "error": str(e)}

# ──────────────────────────────────────────────
#  DATA LOADING (trade_db, journal, logs)
# ──────────────────────────────────────────────

def load_positions():
    try:
        active = trade_db.get_active_trades()
        return {t["symbol"]: t for t in active}
    except Exception:
        return {}

def load_journal():
    try:
        from daily_trade_journal import load_journal_entries
        entries = load_journal_entries()
        if entries:
            return entries
    except Exception:
        pass
    rows = []
    if os.path.exists(JOURNAL_FILE):
        try:
            with open(JOURNAL_FILE, newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f, delimiter="\t")
                seen = set()
                for row in reader:
                    key = (row.get("Symbol", ""), row.get("Timestamp", ""))
                    if key not in seen:
                        seen.add(key)
                        rows.append(row)
        except Exception:
            pass
    return rows[-200:]


def tail_log(filepath, n=200):
    if not os.path.exists(filepath):
        return []
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            lines = [l.rstrip("\r\n") for l in f.readlines()]
        return lines[-n:]
    except Exception as e:
        logging.warning(f"tail_log error reading {filepath}: {e}")
        return []

def compute_stats(positions, journal):
    active = len(positions)
    total = len(journal)
    wins = sum(1 for j in (journal or []) if str(j.get("P&L %") or "").replace("%", "").replace("-", "").strip()
               and str(j.get("Action") or "").startswith("EXIT_"))
    win_rate = round((wins / total) * 100, 1) if total > 0 else 0
    pnl = 0.0
    for j in (journal or []):
        try:
            pnl_str = str(j.get("P&L %") or "").replace("%", "")
            if pnl_str and pnl_str != "-":
                pnl += float(pnl_str)
        except Exception:
            pass
    return {"total_trades": total, "win_rate": win_rate, "active_positions": active, "pnl": round(pnl, 2)}

SCAN_SYMBOLS = [
    "RELIANCE","TCS","HDFCBANK","ICICIBANK","INFY","ITC","SBIN","BHARTIARTL","LT","WIPRO",
    "ADANIENT","ADANIPORTS","APOLLOHOSP","ASIANPAINT","AXISBANK","BAJAJ-AUTO","BAJAJFINSV",
    "BAJFINANCE","BEL","CIPLA","COALINDIA","DRREDDY","EICHERMOT","ETERNAL","GRASIM","HCLTECH",
    "HDFCLIFE","HINDALCO","HINDUNILVR","INDIGO","JIOFIN","JSWSTEEL",
    "KOTAKBANK","M&M","MARUTI","MAXHEALTH","NESTLEIND","NTPC","ONGC","POWERGRID","SBILIFE",
    "SHRIRAMFIN","SUNPHARMA","TATACONSUM","TMPV","TATASTEEL","TECHM","TITAN","TRENT","ULTRACEMCO",
    "NIFTY","BANKNIFTY"
]

# ──────────────────────────────────────────────
#  SCAN PARSING — Split anchor vs ABC per symbol
# ──────────────────────────────────────────────

def _extract_symbol(line):
    for sym in SCAN_SYMBOLS:
        if sym in line:
            return sym
    return None

def parse_scans_for_program(log_lines, prog_id):
    matches = []
    anchors = {}
    abc_matches = {}
    for line in log_lines:
        clean = line.strip()
        if "ANCHOR" in clean:
            matches.append(clean)
            sym = _extract_symbol(clean)
            if sym:
                anchors[sym] = clean
        elif "MATCH" in clean or "BEST TRADE" in clean or "Match" in clean:
            matches.append(clean)
            sym = _extract_symbol(clean)
            if sym:
                abc_matches[sym] = clean
    return matches, anchors, abc_matches

_file_mtime_cache = {}
_parsed_json_cache = {}

def refresh_data(single_run=False):
    global cached_data, _ltp_last_fetch, _kite_positions_last_fetch, _kite_session, _last_scan_reset
    try:
        trade_db.run_db_housekeeping()
    except Exception:
        pass
    while True:
        with data_lock:
            pos = load_positions()
            journal = load_journal()
            cached_data["positions"] = pos
            cached_data["journal"] = journal
            cached_data["stats"] = compute_stats(pos, journal)
            try:
                all_t = trade_db.get_all_trades()
                cached_data["all_trades"] = all_t
            except Exception:
                cached_data["all_trades"] = []
            for pid in PROGRAMS:
                log_file = PROGRAMS[pid].get("log_file")
                log_lines = tail_log(log_file) if log_file else []
                cached_data["log_tail"][pid] = log_lines
                scan_lines, anchors, abc_matches = parse_scans_for_program(log_lines, pid)
                cached_data["scans"][pid] = scan_lines
                cached_data["scan_summary"][pid] = {"anchors": anchors, "abc_matches": abc_matches}
            now_ist = dt.now()
            today_str = now_ist.strftime("%Y-%m-%d")
            market_open = now_ist.replace(hour=9, minute=0, second=0, microsecond=0)
            if _last_scan_reset != today_str and now_ist >= market_open:
                _last_scan_reset = today_str
                for f in [SCAN_DISPLAY_FILE, SCAN_DISPLAY_INDEX_FILE]:
                    try:
                        if os.path.exists(f):
                            with open(f, "r") as fh:
                                existing = json.load(fh)
                            if existing.get("date") == today_str:
                                continue  # Keep today's scan display data intact
                        empty_scan = {"date": today_str, "timestamp": now_ist.strftime("%Y-%m-%d %H:%M:%S"), "staged_trades": [], "carry_forward": [], "active_live": []}
                        with open(f, "w") as fh:
                            json.dump(empty_scan, fh)
                    except Exception:
                        pass
            scan_display = {}
            try:
                if os.path.exists(SCAN_DISPLAY_FILE):
                    mtime = os.path.getmtime(SCAN_DISPLAY_FILE)
                    if _file_mtime_cache.get(SCAN_DISPLAY_FILE) == mtime and SCAN_DISPLAY_FILE in _parsed_json_cache:
                        scan_display["nifty50"] = _parsed_json_cache[SCAN_DISPLAY_FILE]
                    else:
                        with open(SCAN_DISPLAY_FILE, "r") as f:
                            data_obj = json.load(f)
                        _file_mtime_cache[SCAN_DISPLAY_FILE] = mtime
                        _parsed_json_cache[SCAN_DISPLAY_FILE] = data_obj
                        scan_display["nifty50"] = data_obj
            except Exception:
                pass
            for k in ["nifty50", "index"]:
                if k in scan_display and isinstance(scan_display[k], dict):
                    obj = scan_display[k]
                    if not obj.get("staged_trades") and obj.get("all_staged_today"):
                        obj["staged_trades"] = obj["all_staged_today"]
            cached_data["scan_display"] = scan_display
            cached_data["live_execution"] = os.path.exists(LIVE_EXECUTION_FLAG)
            cached_data["live_execution_index"] = os.path.exists(LIVE_EXECUTION_FLAG_INDEX)
            try:
                from trading_core import load_executed_exits, EXECUTED_EXITS
                load_executed_exits()
                cached_data["executed_exits"] = dict(EXECUTED_EXITS)
            except Exception:
                pass
            day_key = dt.now().strftime("%Y-%m-%d")
            global _expired_cache_day, _expired_cache_set
            if _expired_cache_day != day_key:
                try:
                    from trading_core import contract_is_expired
                    idx_set = set()
                    for _cat in ("staged_trades", "all_staged_today", "carry_forward", "active_live"):
                        for item in (cached_data["scan_display"].get("index") or {}).get(_cat) or []:
                            c = str(item.get("contract") or item.get("symbol") or "").replace(" ", "").upper()
                            if c:
                                idx_set.add(c)
                    for _cat in ("staged_trades", "all_staged_today", "carry_forward", "active_live"):
                        for item in (cached_data["scan_display"].get("nifty50") or {}).get(_cat) or []:
                            c = str(item.get("contract") or item.get("symbol") or "").replace(" ", "").upper()
                            if c:
                                idx_set.add(c)
                    for t in cached_data.get("all_trades", []):
                        c = str(t.get("contract") or t.get("symbol") or "").replace(" ", "").upper()
                        if c:
                            idx_set.add(c)
                    _expired_cache_set = {c for c in idx_set if contract_is_expired(c)}
                    _expired_cache_day = day_key
                except Exception:
                    pass
            cached_data["expired_contracts"] = sorted(_expired_cache_set)
            now = time.time()
            if now - _ltp_last_fetch > 3 and cached_data["all_trades"]:
                _ltp_last_fetch = now
                try:
                    active = trade_db.get_active_trades()
                    if active:
                        if not _kite_session:
                            tk = check_token_valid()
                            if tk["valid"]:
                                td = json.load(open(TOKEN_FILE))
                                api_key, _ = get_kite_credentials()
                                ks = KiteConnect(api_key=api_key)
                                ks.set_access_token(td["access_token"])
                                _kite_session = ks
                        if _kite_session:
                            syms = []
                            for t in active:
                                tok = t.get("option_token") or t.get("index_token")
                                if tok: syms.append(int(tok))
                            if syms:
                                quotes = _kite_session.quote(syms)
                                ltp = {}
                                for key, q in quotes.items():
                                    ltp[key.split(":")[-1]] = q.get("last_price", 0)
                                cached_data["ltp"] = ltp
                except Exception:
                    _kite_session = None
        if now - _kite_positions_last_fetch > 3:
            _kite_positions_last_fetch = now
            try:
                if not _kite_session:
                    tk = check_token_valid()
                    if tk["valid"]:
                        td = json.load(open(TOKEN_FILE))
                        api_key, _ = get_kite_credentials()
                        ks = KiteConnect(api_key=api_key)
                        ks.set_access_token(td["access_token"])
                        _kite_session = ks
                if _kite_session:
                    kite_positions = _kite_session.positions()
                    merged = []
                    net_pos = [p for p in kite_positions.get("net", []) if p.get("tradingsymbol") and int(p.get("quantity", 0)) > 0]
                    q_keys = [f"{p.get('exchange', 'NFO')}:{p.get('tradingsymbol')}" for p in net_pos]
                    quotes_bulk = {}
                    if q_keys:
                        try:
                            quotes_bulk = _kite_session.quote(q_keys)
                        except Exception:
                            pass
                    for p in net_pos:
                        sym = p.get("tradingsymbol", "")
                        qty = int(p.get("quantity", 0))
                        entry_pr = float(p.get("average_price", 0))
                        exch = p.get("exchange", "NFO")
                        q_key = f"{exch}:{sym}"
                        live_ltp = 0
                        if q_key in quotes_bulk:
                            live_ltp = float(quotes_bulk[q_key].get("last_price", 0))

                        tok_id = str(p.get("instrument_token", ""))
                        sym_str = str(sym)

                        if live_ltp > 0:
                            _ltp_memory[sym_str] = live_ltp
                            if tok_id:
                                _ltp_memory[tok_id] = live_ltp
                            cached_data["ltp"][sym_str] = live_ltp
                            if tok_id:
                                cached_data["ltp"][tok_id] = live_ltp
                        else:
                            live_ltp = _ltp_memory.get(sym_str) or _ltp_memory.get(tok_id) or 0
                            if live_ltp > 0:
                                cached_data["ltp"][sym_str] = live_ltp
                                if tok_id:
                                    cached_data["ltp"][tok_id] = live_ltp

                        if live_ltp > 0 and entry_pr > 0:
                            live_pnl = round((live_ltp - entry_pr) * qty, 2)
                            _pnl_memory[sym_str] = live_pnl
                        else:
                            live_pnl = _pnl_memory.get(sym_str, float(p.get("pnl", 0)))

                        # Fail-Safe Active Position Risk Monitor
                        try:
                            contract_name = p.get("tradingsymbol", sym)
                            engine_type = "index" if ("NIFTY" in contract_name or "BANK" in contract_name or "SENSEX" in contract_name) else "nifty50"
                            scan_sl = lookup_scan_sl_target(contract_name, contract_name, engine_type, _kite_session, entry_pr)
                            pos_item = {
                                "contract": contract_name,
                                "symbol": contract_name,
                                "quantity": qty,
                                "entry_price": entry_pr,
                                "entry_spot": entry_pr,
                                "ltp": live_ltp,
                                "pnl": live_pnl,
                                "exchange": exch,
                                "source": "kite"
                            }
                            if scan_sl:
                                pos_item["current_sl"] = scan_sl.get("current_sl", 0)
                                pos_item["t1"] = scan_sl.get("t1", 0)
                                pos_item["t2"] = scan_sl.get("t2", 0)
                                pos_item["t3"] = scan_sl.get("t3", 0)
                                pos_item["pattern"] = scan_sl.get("pattern", "SCAN_LINKED")
                            merged.append(pos_item)

                            if scan_sl:
                                def _safe_float(v):
                                    try:
                                        if v is None:
                                            return 0.0
                                        s = str(v).strip()
                                        return float(s) if s not in ("", "N/A", "None") else 0.0
                                    except (TypeError, ValueError):
                                        return 0.0

                                def _t1_early_buffer(v):
                                    if v <= 0:
                                        return 0.0
                                    if v <= 50:
                                        return max(0.50, round(v * 0.015, 2))
                                    elif v <= 200:
                                        return max(1.00, round(v * 0.015, 2))
                                    else:
                                        return max(2.00, round(v * 0.010, 2))

                                def _failsafe_exit_mark(action, status, details, exit_price):
                                    entry_s = entry_pr if entry_pr > 0 else float(scan_sl.get("entry_spot") or 0)
                                    pnl = ((exit_price - entry_s) / entry_s * 100) if entry_s else 0
                                    mark_tid = tid
                                    if not mark_tid:
                                        try:
                                            mark_tid = trade_db.find_active_trade_id(contract_name, engine_type or None) or trade_db.find_active_trade_id(contract_name)
                                        except Exception:
                                            mark_tid = None
                                    if mark_tid:
                                        trade_db.update_trade(mark_tid, {
                                            "status": status,
                                            "exit_time": dt.now().strftime("%Y-%m-%d %H:%M:%S"),
                                            "pnl_percent": round(pnl, 2),
                                            "details": details
                                        })
                                    try:
                                        log_to_journal(contract_name, scan_sl.get("pattern", "SCAN_LINKED"), "15minute",
                                                       action, "CLOSED", details, pnl,
                                                       entry=entry_s, sl=sl_val,
                                                       target=t1_val if t1_val else t3_val,
                                                       event_time=dt.now().strftime("%Y-%m-%d %H:%M:%S"))
                                    except Exception:
                                        pass

                                ltp_val = live_ltp
                                sl_val = _safe_float(scan_sl.get("current_sl"))
                                t1_val = _safe_float(scan_sl.get("t1"))
                                t2_val = _safe_float(scan_sl.get("t2"))
                                t3_val = _safe_float(scan_sl.get("t3"))
                                t_stage = int(scan_sl.get("trailing_stage") or 0)
                                tid = scan_sl.get("id")

                                clean_sym = str(contract_name).replace(" ", "").upper()
                                now_t = dt.now().time()
                                cfg_f = load_config()
                                fs_start_str = cfg_f.get("failsafe_start_time", "09:45")
                                try:
                                    f_h, f_m = map(int, fs_start_str.split(":"))
                                    fs_start_t = datetime_time(f_h, f_m)
                                except Exception:
                                    fs_start_t = datetime_time(9, 45)

                                # Buffer & Previous Candle Confirmation for SL Exit
                                sl_buffered = round(sl_val * 0.995, 2)
                                is_below_buffer = ltp_val <= sl_buffered
                                is_deep_break = ltp_val <= round(sl_val * 0.985, 2)

                                prev_closed_below = False
                                token_id = scan_sl.get("option_token") or scan_sl.get("index_token") or scan_sl.get("token")
                                if token_id and _kite_session:
                                    try:
                                        df_hist = fetch_and_resample_candles(_kite_session, token_id, (dt.now() - timedelta(days=2)).strftime("%Y-%m-%d"), dt.now().strftime("%Y-%m-%d"), "15minute")
                                        if len(df_hist) >= 2:
                                            prev_close_val = float(df_hist.iloc[-2]["close"])
                                            if prev_close_val > 0 and prev_close_val <= sl_val:
                                                prev_closed_below = True
                                    except Exception:
                                        pass

                                # TASK 0: Pause automated exit execution before 09:45 AM due to opening market volatility
                                if now_t < fs_start_t:
                                    logging.info(f"[FAILSAFE PAUSED BEFORE {fs_start_str} AM] {contract_name} automated exit paused until {fs_start_str} AM (Current time: {now_t.strftime('%H:%M:%S')}).")
                                # TASK 1: Pause automated exit execution if user is actively editing this symbol on the UI
                                elif clean_sym in ACTIVE_EDIT_LOCKS:
                                    logging.info(f"[FAILSAFE PAUSED] {contract_name} is currently being edited on UI. Automated exit execution paused.")
                                # TASK 2: Execute SL exit ONLY IF below 0.5% buffer AND (previous candle closed below SL OR emergency deep break)
                                elif ltp_val > 0 and sl_val > 0 and is_below_buffer and (prev_closed_below or is_deep_break):
                                    logging.warning(f"[FAILSAFE MONITOR EXIT SL CONFIRMED] {contract_name} LTP={ltp_val} <= Buffered SL={sl_buffered} (Prev Close Below: {prev_closed_below}, Deep Break: {is_deep_break})")
                                    pos_obj = {"contract": contract_name, "position_size": qty, "quantity": qty}
                                    shared_close_position(_kite_session, pos_obj, True, p.get("product"))
                                    _failsafe_exit_mark("EXIT_SL", "SL_HIT",
                                                        f"SL hit [{('CANDLE_CLOSE_SL' if prev_closed_below else 'EMERGENCY_HARD_SL')}] | LTP {ltp_val:.2f} | SL {sl_val:.2f}", ltp_val)
                                # 2. Check T3 Target Hit Exit
                                elif ltp_val > 0 and t3_val > 0 and ltp_val >= t3_val:
                                    logging.info(f"[FAILSAFE MONITOR EXIT T3] {contract_name} LTP={ltp_val} >= T3={t3_val}")
                                    pos_obj = {"contract": contract_name, "position_size": qty, "quantity": qty}
                                    shared_close_position(_kite_session, pos_obj, True, p.get("product"))
                                    _failsafe_exit_mark("EXIT_T3", "TARGET_HIT",
                                                        f"T3 exit ({ltp_val:.2f} >= T3 {t3_val:.2f})", ltp_val)
                                # 2b. Check T1 Target Exit (No T2/T3 -> Full exit 1-2 pts early on T1 touch)
                                elif ltp_val > 0 and t1_val > 0 and t2_val <= 0 and t3_val <= 0 and ltp_val >= (t1_val - _t1_early_buffer(t1_val)):
                                    logging.warning(f"[FAILSAFE MONITOR EXIT T1 (no T2/T3)] {contract_name} LTP={ltp_val} >= T1-buffer={t1_val - _t1_early_buffer(t1_val):.2f} (Target: {t1_val:.2f})")
                                    pos_obj = {"contract": contract_name, "position_size": qty, "quantity": qty}
                                    shared_close_position(_kite_session, pos_obj, True, p.get("product"))
                                    _failsafe_exit_mark("EXIT_T1", "TARGET_HIT",
                                                        f"T1 full exit ({ltp_val:.2f} >= {t1_val - _t1_early_buffer(t1_val):.2f}, no T2/T3)", ltp_val)
                                # Track highest price reached for position
                                prev_high = float(scan_sl.get("high_price") or 0)
                                pos_high = max(live_ltp, prev_high)
                                if tid and live_ltp > prev_high:
                                    trade_db.update_trade(tid, {"high_price": live_ltp})

                                effective_entry = entry_pr if entry_pr > 0 else float(scan_sl.get("entry_spot") or 0)

                                # 3. Trailing SL Stage 1 (T1 Hit -> Trail SL to Breakeven / Entry)
                                if t_stage == 0 and t1_val > effective_entry and pos_high >= t1_val and effective_entry > 0:
                                    logging.info(f"[FAILSAFE TRAIL 1] {contract_name} High={pos_high} >= T1={t1_val} -> Trailing SL to Breakeven ({effective_entry})")
                                    if tid: trade_db.update_trade(tid, {"current_sl": effective_entry, "trailing_stage": 1, "sl_set_time": dt.now().isoformat()})
                                # 4. Trailing SL Stage 2 (T2 Hit -> Trail SL to T1)
                                elif t_stage == 1 and t2_val > t1_val and pos_high >= t2_val and t1_val > 0:
                                    logging.info(f"[FAILSAFE TRAIL 2] {contract_name} High={pos_high} >= T2={t2_val} -> Trailing SL to T1 ({t1_val})")
                                    if tid: trade_db.update_trade(tid, {"current_sl": t1_val, "trailing_stage": 2, "sl_set_time": dt.now().isoformat()})
                        except Exception as fs_err:
                            logging.debug(f"Failsafe monitor error for {sym}: {fs_err}")

                    cached_data["kite_positions"] = merged
            except Exception:
                pass
        if single_run:
            break
        if int(time.time()) % 3600 < REFRESH_SECONDS:
            auto_export_if_new_month()
        time.sleep(REFRESH_SECONDS)

# Dashboard HTML/JS extracted to templates/index.html (keeps the Python file lean and reduces AI-context load)
with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates/index.html'), encoding="utf-8") as _template_f:
    HTML_TEMPLATE = _template_f.read()

# ──────────────────────────────────────────────
#  FLASK ROUTES — API Endpoints
# ──────────────────────────────────────────────

@app.route("/")
def dashboard():
    return render_template_string(HTML_TEMPLATE, refresh=REFRESH_SECONDS, programs=PROGRAMS)

@app.route("/api/status")
def api_status():
    with data_lock:
        prog_status = {}
        for pid in PROGRAMS:
            if pid == "ema_engine":
                pid_running = get_ema_engine_status(is_options_mode=True)
            else:
                pid_running = get_pid_for_program(pid) is not None
            log_lines = cached_data["log_tail"].get(pid, [])
            if not log_lines and PROGRAMS[pid].get("log_file"):
                log_lines = tail_log(PROGRAMS[pid].get("log_file"))
            print(f"[DEBUG_API] pid={pid} log_file={PROGRAMS[pid].get('log_file')} lines={len(log_lines)}")
            prog_status[pid] = {
                "running": pid_running,
                "scans": cached_data["scans"].get(pid, []),
                "log_tail": log_lines,
                "scan_summary": cached_data["scan_summary"].get(pid, {"anchors": {}, "abc_matches": {}})
            }
        cfg = load_config()
        return jsonify({
            "programs": prog_status,
            "positions": cached_data["positions"],
            "all_trades": cached_data["all_trades"],
            "kite_positions": cached_data["kite_positions"],
            "ltp": {str(k): v for k, v in cached_data["ltp"].items()},
            "journal": cached_data["journal"],
            "stats": cached_data["stats"],
            "config": cfg,
            "scan_display": cached_data["scan_display"],
            "ema_scan": get_ema_scan_data(is_options_mode=True),
            "live_execution": cached_data["live_execution"],
            "live_execution_index": cached_data["live_execution_index"],
            "executed_exits": cached_data.get("executed_exits", {}),
            "expired_contracts": cached_data.get("expired_contracts", [])
        })

@app.route("/api/token/check")
def api_token_check():
    return jsonify(check_token_valid())

@app.route("/api/token/url")
def api_token_url():
    url = get_login_url()
    if not url:
        return jsonify({"url": "", "error": "Kite API Key is missing. Please save your API Key & Secret below or configure input/program_config.json."})
    return jsonify({"url": url})

@app.route("/api/token/save-credentials", methods=["POST"])
def api_token_save_credentials():
    data = request.get_json(force=True, silent=True) or {}
    api_key = str(data.get("api_key", "")).strip()
    api_secret = str(data.get("api_secret", "")).strip()
    if not api_key or not api_secret:
        return jsonify({"ok": False, "error": "Both API Key and API Secret are required."})
    cfg = load_config()
    cfg["api_key"] = api_key
    cfg["api_secret"] = api_secret
    os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)
    return jsonify({"ok": True, "url": get_login_url()})

@app.route("/api/token/callback", methods=["GET", "POST"])
def api_token_callback():
    req_token = request.args.get("request_token") or request.form.get("request_token")
    if not req_token:
        data = request.get_json(force=True, silent=True) or {}
        req_token = data.get("request_token")
    if not req_token:
        return "<h3>Error: No request_token received from Zerodha</h3><p><a href='/'>Return to Dashboard</a></p>", 400
    
    res = exchange_request_token(req_token.strip())
    if res.get("ok"):
        return """
        <!DOCTYPE html>
        <html>
        <head><title>Kite Token Generated</title></head>
        <body style="font-family:sans-serif;background:#0d1117;color:#c9d1d9;text-align:center;padding:50px;">
            <div style="background:#161b22;border:1px solid #238636;border-radius:8px;padding:30px;max-width:500px;margin:auto;box-shadow:0 4px 12px rgba(0,0,0,0.5);">
                <h2 style="color:#3fb950;margin-top:0;">✅ Token Generated Successfully!</h2>
                <p style="color:#8b949e;">Zerodha access token has been generated and saved for today.</p>
                <a href="/" style="display:inline-block;background:#238636;color:#ffffff;padding:10px 20px;border-radius:6px;text-decoration:none;font-weight:bold;margin-top:15px;">Return to Dashboard</a>
            </div>
            <script>setTimeout(function(){ window.location.href = '/'; }, 1500);</script>
        </body>
        </html>
        """
    else:
        err = res.get("error", "Unknown error")
        return f"<h3>Token Exchange Failed</h3><p>Error: {err}</p><p><a href='/'>Return to Dashboard</a></p>", 500

@app.route("/api/postback", methods=["POST"])
def api_postback():
    try:
        data = request.get_json(force=True, silent=True) or request.form.to_dict()
        logging.info(f"[POSTBACK] Received Zerodha order update: {data}")
    except Exception as e:
        logging.warning(f"[POSTBACK] Error parsing postback: {e}")
    return jsonify({"status": "success"}), 200

@app.route("/api/token/exchange", methods=["POST"])
def api_token_exchange():
    data = request.get_json(force=True, silent=True)
    if not data or not data.get("request_token"):
        return jsonify({"ok": False, "error": "No request_token provided"})
    result = exchange_request_token(data["request_token"].strip())
    return jsonify(result)

@app.route("/api/backtest/mode", methods=["GET", "POST"])
def api_backtest_mode():
    if request.method == "POST":
        data = request.get_json(force=True, silent=True)
        enabled = bool(data.get("enabled", False))
        set_backtest_mode(enabled)
    return jsonify({"enabled": get_backtest_mode()})

@app.route("/api/config/<prog_id>", methods=["POST"])
def api_save_config(prog_id):
    if prog_id not in PROGRAMS:
        return jsonify({"ok": False, "error": "Unknown program"})
    data = request.get_json(force=True, silent=True)
    if not data:
        return jsonify({"ok": False, "error": "Invalid JSON"})
    save_config(prog_id, data)
    return jsonify({"ok": True})

@app.route("/api/scan/clear", methods=["POST"])
def api_scan_clear():
    now_str = dt.now().strftime("%Y-%m-%d %H:%M:%S")
    today_str = dt.now().strftime("%Y-%m-%d")
    for f in [paths.SCAN_DISPLAY_FILE, paths.SCAN_DISPLAY_INDEX_FILE, paths.SCAN_DISPLAY_STOCK_FILE, paths.SCAN_DISPLAY_BEAR_FILE]:
        try:
            empty_scan = {
                "date": today_str,
                "timestamp": now_str,
                "cleared_at": now_str,
                "staged_trades": [],
                "all_staged_today": [],
                "carry_forward": [],
                "active_live": []
            }
            os.makedirs(os.path.dirname(f), exist_ok=True)
            with open(f, "w", encoding="utf-8") as fh:
                json.dump(empty_scan, fh, indent=2)
        except Exception:
            pass
    try:
        if os.path.exists(paths.CYCLE_STORE_FILE):
            with open(paths.CYCLE_STORE_FILE, "w", encoding="utf-8") as fh:
                json.dump({}, fh)
    except Exception:
        pass
    with data_lock:
        for k in ["nifty50", "index", "daily", "bear_trade"]:
            cached_data["scan_display"][k] = {"staged_trades": [], "all_staged_today": [], "carry_forward": [], "active_live": [], "cleared_at": now_str}
    return jsonify({"ok": True})

@app.route("/api/scan/ema/clear", methods=["POST"])
def api_scan_ema_clear():
    now_str = dt.now().strftime("%Y-%m-%d %H:%M:%S")
    today_str = dt.now().strftime("%Y-%m-%d")
    for ema_file in [paths.SCAN_DISPLAY_EMA_FILE, paths.SCAN_DISPLAY_EMA_STOCK_FILE]:
        try:
            empty_scan = {
                "ema_engine": {"staged_trades": [], "all_staged_today": [], "carry_forward": [], "active_live": []},
                "last_updated": now_str,
                "cleared_at": now_str
            }
            os.makedirs(os.path.dirname(ema_file), exist_ok=True)
            with open(ema_file, "w", encoding="utf-8") as fh:
                json.dump(empty_scan, fh, indent=2)
        except Exception:
            pass
    with data_lock:
        cached_data["ema_scan"] = {
            "ema_engine": {"staged_trades": [], "all_staged_today": [], "carry_forward": [], "active_live": []},
            "last_updated": now_str,
            "cleared_at": now_str
        }
    return jsonify({"ok": True})

def _format_pattern_result(p):
    if not p: return '-'
    p_str = str(p)
    if 'Engulf' in p_str:
        return 'BEAR_ENG' if 'Bear' in p_str or 'BEAR' in p_str else 'BULL_ENG'
    elif 'Two_Higher' in p_str or 'Higher_Highs' in p_str:
        return 'BULL_2HH'
    elif 'Two_Lower' in p_str or 'Lower_Lows' in p_str:
        return 'BEAR_2LL'
    elif 'HH_Sweep' in p_str or 'HH_sweep' in p_str:
        return 'BEAR_HH'
    elif 'Sweep' in p_str or 'LL' in p_str:
        return 'BULL_LL'
    elif 'Star' in p_str or 'Shooting' in p_str:
        return 'BEAR_STAR'
    elif 'Baby' in p_str or 'Hammer' in p_str:
        return 'BULL_HAM'
    elif 'Harami' in p_str:
        return 'BEAR_HAR' if 'Bear' in p_str or 'BEAR' in p_str else 'BULL_HAR'
    elif 'Base' in p_str:
        return 'BULL_BASE'
    elif p_str == 'SCAN_READY':
        return 'BULL_ENG'
    return p_str

def _format_timestamp(ts):
    if not ts: return '-'
    try:
        s = str(ts).split('+')[0].replace('T', ' ')
        p = s.split(' ')
        dp = p[0].split('-') if p[0] else []
        tp = p[1].split(':') if len(p) > 1 and p[1] else []
        if len(dp) == 3 and len(tp) >= 2:
            h = int(tp[0])
            m = int(tp[1])
            # Clamp out-of-market night timestamps (e.g. 22:07 PM) to market open 09:15 AM
            if h > 15 or h < 9 or (h == 15 and m > 30):
                return f"{dp[2]}-{dp[1]}-{dp[0][-2:]} 09:15"
            return f"{dp[2]}-{dp[1]}-{dp[0][-2:]} {tp[0]}:{tp[1]}"
        return s
    except Exception:
        return str(ts)

def _format_float(val, dec=2):
    if val is None or val == '' or val == '-':
        return '-'
    try:
        return f"{float(val):.{dec}f}"
    except Exception:
        return str(val)

@app.route("/api/scan/export", methods=["POST"])
def api_scan_export():
    try:
        import io
        from spot_enricher import extract_underlying_symbol, evaluate_spot_trend_and_t1
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Symbol", "Contract", "Side", "Entry", "SL", "T1", "T2", "T3",
                         "AncherT", "EntryTime", "Result", "CF", "RR", "Engine", "Status",
                         "Spot_Trend", "Spot_T1_Target"])
        files = [("Nifty 50", SCAN_DISPLAY_FILE), ("Index", SCAN_DISPLAY_INDEX_FILE), ("Stock EMA", EMA_DISPLAY_FILE_OPTION)]
        spot_eval_cache = {}
        for label, path in files:
            full = path if os.path.isabs(path) else os.path.join(BASE_DIR, path)
            if not os.path.exists(full):
                continue
            with open(full) as f:
                data = json.load(f)
            if isinstance(data, dict) and "ema_engine" in data:
                data = data["ema_engine"]
            for section_name, status_tag in [("staged_trades", "Staged"), ("active_live", "Active"), ("carry_forward", "CarryFwd")]:
                for t in data.get(section_name, []):
                    side_val = t.get("side", "")
                    if not side_val:
                        cnt_str = str(t.get("contract") or t.get("symbol") or "").upper()
                        if "CE" in cnt_str:
                            side_val = "CE"
                        elif "PE" in cnt_str:
                            side_val = "PE"
                    writer.writerow([
                        t.get("symbol", ""),
                        t.get("contract", ""),
                        side_val,
                        _format_float(t.get("entry") or t.get("entry_spot")),
                        _format_float(t.get("sl") or t.get("current_sl")),
                        _format_float(t.get("t1")),
                        _format_float(t.get("t2")),
                        _format_float(t.get("t3")),
                        _format_timestamp(t.get("candle_a_time")),
                        _format_timestamp(t.get("entry_time")),
                        _format_pattern_result(t.get("pattern") or t.get("result")),
                        "Yes" if t.get("carry_forward") else "No",
                        _format_float(t.get("rr")),
                        label,
                        status_tag
                    ])
        csv_bytes = output.getvalue().encode("utf-8-sig")
        return Response(csv_bytes, mimetype="text/csv",
                        headers={"Content-Disposition": f"attachment; filename=scan_export_{dt.now().strftime('%d_%m_%y_%H%M')}.csv"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/export/ema", methods=["POST"])
def api_export_ema():
    try:
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Symbol", "Contract", "Side", "Entry", "SL", "T1", "T2", "T3",
                         "Spot", "RR", "Timeframe", "Pattern", "AncherT", "EntryTime", "Status"])
        full = EMA_DISPLAY_FILE_OPTION if os.path.isabs(EMA_DISPLAY_FILE_OPTION) else os.path.join(BASE_DIR, EMA_DISPLAY_FILE_OPTION)
        if os.path.exists(full):
            with open(full, encoding="utf-8") as f:
                data = json.load(f)
            ema_payload = data.get("ema_engine", data) if isinstance(data, dict) else {}
            for section_name, status_tag in [("staged_trades", "Staged"), ("active_live", "Active"), ("carry_forward", "CarryFwd")]:
                for t in ema_payload.get(section_name, []):
                    side_val = t.get("side", "")
                    if not side_val:
                        cnt_str = str(t.get("contract") or t.get("symbol") or "").upper()
                        if "CE" in cnt_str:
                            side_val = "CE"
                        elif "PE" in cnt_str:
                            side_val = "PE"
                    writer.writerow([
                        t.get("symbol", ""),
                        t.get("contract", ""),
                        side_val,
                        _format_float(t.get("entry")),
                        _format_float(t.get("sl") or t.get("current_sl")),
                        _format_float(t.get("t1")),
                        _format_float(t.get("t2")),
                        _format_float(t.get("t3")),
                        _format_float(t.get("entry_spot")),
                        _format_float(t.get("rr")),
                        t.get("timeframe", ""),
                        t.get("pattern", ""),
                        _format_timestamp(t.get("candle_a_time")),
                        _format_timestamp(t.get("entry_time")),
                        status_tag
                    ])
        csv_bytes = output.getvalue().encode("utf-8-sig")
        return Response(csv_bytes, mimetype="text/csv",
                        headers={"Content-Disposition": f"attachment; filename=ema_export_{dt.now().strftime('%d_%m_%y_%H%M')}.csv"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/journal/clear", methods=["POST"])
def api_journal_clear():
    try:
        if os.path.exists(JOURNAL_FILE):
            open(JOURNAL_FILE, "w").close()
    except Exception:
        pass
    with data_lock:
        cached_data["journal"] = []
        cached_data["stats"] = compute_stats(cached_data.get("positions", {}), [])
    return jsonify({"ok": True})

@app.route("/api/programs/<prog_id>/start", methods=["POST"])
def api_start(prog_id):
    if prog_id not in PROGRAMS:
        return jsonify({"ok": False, "error": "Unknown program"})
    token = check_token_valid()
    if not token["valid"]:
        return jsonify({"ok": False, "error": token["reason"]})
    if prog_id == "ema_engine":
        cfg = load_config()
        tf = cfg.get("ema_engine", {}).get("timeframe", "1d")
        tu = cfg.get("ema_engine", {}).get("target_universe", "ALL")
        interval = int(cfg.get("ema_engine", {}).get("scan_interval", 300))
        ok, msg = start_ema_engine(timeframe=tf, is_options_mode=True, scan_interval=interval, target_universe=tu)
        return jsonify({"ok": ok, "error": None if ok else msg})
    ok = start_program(prog_id)
    return jsonify({"ok": ok, "error": None if ok else "Start failed"})

@app.route("/api/programs/<prog_id>/stop", methods=["POST"])
def api_stop(prog_id):
    if prog_id not in PROGRAMS:
        return jsonify({"ok": False, "error": "Unknown program"})
    if prog_id == "ema_engine":
        ok, msg = stop_ema_engine(is_options_mode=True)
        return jsonify({"ok": ok, "error": None if ok else msg})
    ok = stop_program(prog_id)
    return jsonify({"ok": ok})

ANCHOR_SCAN_REQUEST_FILE = os.path.join("output", "monitor", "anchor_scan_request.txt")
ANCHOR_SCAN_STOP_FILE = os.path.join("output", "monitor", "anchor_scan_stop.txt")

@app.route("/api/anchor/scan", methods=["POST"])
def api_anchor_scan():
    data = request.get_json(silent=True) or {}
    engine = data.get("engine", "index")
    try:
        with data_lock:
            cached_data["anchor_status"]["running"] = True
            cached_data["anchor_status"]["engine"] = engine
            cached_data["anchor_status"]["requested_at"] = time.time()
        if os.path.exists(ANCHOR_SCAN_STOP_FILE):
            os.remove(ANCHOR_SCAN_STOP_FILE)
        # Launch a dedicated --anchor-only subprocess. We intentionally do NOT
        # write ANCHOR_SCAN_REQUEST_FILE here: a running engine also polls that
        # file in its main loop, which would cause the anchor scan to run twice.
        script = PROGRAMS.get(engine, {}).get("file")
        if script:
            script_path = os.path.join(BASE_DIR, script)
            subprocess.Popen([sys.executable, script_path, "--anchor-only"],
                             cwd=BASE_DIR, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/anchor/stop", methods=["POST"])
def api_anchor_stop():
    try:
        os.makedirs(os.path.dirname(ANCHOR_SCAN_STOP_FILE), exist_ok=True)
        with open(ANCHOR_SCAN_STOP_FILE, "w") as f:
            f.write("stop")
        with data_lock:
            cached_data["anchor_status"]["running"] = False
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/anchor/status")
def api_anchor_status():
    with data_lock:
        st = dict(cached_data["anchor_status"])
    if not st.get("running"):
        still_running = os.path.exists(ANCHOR_SCAN_REQUEST_FILE) and not os.path.exists(ANCHOR_SCAN_STOP_FILE)
        if still_running:
            st["running"] = True
            if not st.get("engine"):
                try:
                    with open(ANCHOR_SCAN_REQUEST_FILE) as f:
                        st["engine"] = f.read().strip()
                except Exception:
                    pass
    return jsonify(st)

@app.route("/api/logs/clear", methods=["POST"])
def api_logs_clear():
    log_files = [INDEX_LOG_FILE, NIFTY50_LOG_FILE, EMA_LOG_FILE]
    for lf in log_files:
        try:
            if os.path.exists(lf):
                open(lf, "w").close()
        except Exception:
            pass
    with data_lock:
        for pid in PROGRAMS:
            cached_data["log_tail"][pid] = []
            cached_data["scans"][pid] = []
            cached_data["scan_summary"][pid] = {"anchors": {}, "abc_matches": {}}
    return jsonify({"ok": True})

@app.route("/api/trades")
def api_trades():
    engine = request.args.get("engine")
    active_only = request.args.get("active", "false").lower() == "true"
    if active_only:
        return jsonify(trade_db.get_active_trades(engine))
    return jsonify(trade_db.get_all_trades(engine))

@app.route("/api/export/monthly", methods=["POST"])
def api_export_monthly():
    result = run_monthly_export()
    return jsonify({"ok": True, **result})

@app.route("/api/live-execution/nifty50", methods=["GET", "POST"])
def api_live_execution():
    if request.method == "POST":
        enabled = request.get_json(force=True, silent=True).get("enabled", False)
        flag_path = os.path.join(BASE_DIR, LIVE_EXECUTION_FLAG)
        if enabled:
            with open(flag_path, "w") as f:
                f.write("1")
        else:
            if os.path.exists(flag_path):
                os.remove(flag_path)
        with data_lock:
            cached_data["live_execution"] = enabled
        return jsonify({"ok": True, "enabled": enabled})
    return jsonify({"enabled": os.path.exists(os.path.join(BASE_DIR, LIVE_EXECUTION_FLAG))})

@app.route("/api/live-execution/index", methods=["GET", "POST"])
def api_live_execution_index():
    if request.method == "POST":
        enabled = request.get_json(force=True, silent=True).get("enabled", False)
        flag_path = os.path.join(BASE_DIR, LIVE_EXECUTION_FLAG_INDEX)
        if enabled:
            with open(flag_path, "w") as f:
                f.write("1")
        else:
            if os.path.exists(flag_path):
                os.remove(flag_path)
        with data_lock:
            cached_data["live_execution_index"] = enabled
        return jsonify({"ok": True, "enabled": enabled})
    return jsonify({"enabled": os.path.exists(os.path.join(BASE_DIR, LIVE_EXECUTION_FLAG_INDEX))})

@app.route("/api/edit-lock", methods=["POST"])
def api_edit_lock():
    data = request.json or {}
    sym = data.get("symbol")
    active = data.get("active", False)
    if sym:
        clean_s = str(sym).replace(" ", "").upper()
        if active:
            ACTIVE_EDIT_LOCKS.add(clean_s)
            logging.info(f"[EDIT LOCK ON] Automated exit execution paused for {clean_s}")
        else:
            ACTIVE_EDIT_LOCKS.discard(clean_s)
            logging.info(f"[EDIT LOCK OFF] Automated exit execution resumed for {clean_s}")
    return jsonify({"ok": True})

@app.route("/api/update-position", methods=["POST"])
def api_update_position():
    data = request.get_json(force=True, silent=True) or {}
    engine = data.get("engine", "nifty50")
    symbol = data.get("symbol", "")
    current_sl = data.get("current_sl")
    t1 = data.get("t1")
    t2 = data.get("t2")
    t3 = data.get("t3")
    entry_price = data.get("entry_price")
    if not symbol or (current_sl is None and t1 is None and t2 is None and t3 is None):
        return jsonify({"ok": False, "error": "symbol and at least one level required"}), 400
    vals = {}
    if current_sl is not None and str(current_sl).strip() != "": vals["current_sl"] = float(current_sl)
    if t1 is not None and str(t1).strip() != "": vals["t1"] = float(t1)
    if t2 is not None and str(t2).strip() != "": vals["t2"] = float(t2)
    if t3 is not None and str(t3).strip() != "": vals["t3"] = float(t3)
    if entry_price is not None and str(entry_price).strip() != "": vals["entry_spot"] = float(entry_price)
    vals["user_edited"] = True
    
    clean_target = str(symbol).replace(" ", "").upper()

    write_sl_overrides(engine, symbol, vals, (engine, "nifty50", "index"))

    clear_executed_exit(symbol)
    clear_executed_exit(clean_target)
    ACTIVE_EDIT_LOCKS.discard(clean_target)

    matched = False
    with data_lock:
        update_keys = list(vals.keys())

        # 1. Update in-memory all_trades
        for t in cached_data.get("all_trades", []):
            t_sym = str(t.get("symbol") or "").replace(" ", "").upper()
            t_cnt = str(t.get("contract") or "").replace(" ", "").upper()
            if clean_target in (t_sym, t_cnt) or t_sym in clean_target or t_cnt in clean_target:
                matched = True
                for k in update_keys: t[k] = vals[k]
                tid = t.get("id")
                if tid:
                    trade_db.update_trade(tid, vals)

        # 2. Update in-memory positions
        for pos_key, pos in (cached_data.get("positions", {}).items() if isinstance(cached_data.get("positions"), dict) else enumerate(cached_data.get("positions", []))):
            if isinstance(pos, dict):
                p_sym = str(pos.get("symbol") or "").replace(" ", "").upper()
                p_cnt = str(pos.get("contract") or "").replace(" ", "").upper()
                if clean_target in (p_sym, p_cnt) or p_sym in clean_target or p_cnt in clean_target:
                    matched = True
                    for k in update_keys: pos[k] = vals[k]
                    tid = pos.get("id")
                    if tid:
                        trade_db.update_trade(tid, vals)

        # 3. Update in-memory kite_positions so UI refreshes immediately
        for kp in cached_data.get("kite_positions", []):
            k_sym = str(kp.get("symbol") or "").replace(" ", "").upper()
            k_cnt = str(kp.get("contract") or "").replace(" ", "").upper()
            if clean_target in (k_sym, k_cnt) or k_sym in clean_target or k_cnt in clean_target:
                for k in update_keys: kp[k] = vals[k]

        if not matched:
            contract = symbol
            exchange = "NSE"
            for kp in cached_data.get("kite_positions", []):
                k_sym = str(kp.get("symbol") or "").replace(" ", "").upper()
                k_cnt = str(kp.get("contract") or "").replace(" ", "").upper()
                if clean_target in (k_sym, k_cnt) or k_sym in clean_target or k_cnt in clean_target:
                    contract = kp.get("contract", symbol)
                    exchange = kp.get("exchange", "NSE")
                    break
            is_stock = exchange == "NSE"
            trade_data = {"contract": contract, "entry_spot": vals.get("entry_spot", 0), "position_type": "stock" if is_stock else "option"}
            trade_data.update(vals)
            db_symbol = resolve_underlying(symbol or contract, engine)
            tid, _created = trade_db.create_trade(engine, db_symbol, trade_data)
            entry = {"symbol": db_symbol, "contract": contract, "id": tid, "engine": engine, "status": "ACTIVE", "position_type": "stock" if is_stock else "option"}
            entry.update(vals)
            cached_data["all_trades"].append(entry)
            cached_data["positions"][symbol] = entry
            logging.info(f"[OVERRIDE] Created new DB trade for {engine}/{symbol}")

        # Synchronize scan_display in memory and on disk so 1s polling preserves edit immediately
        disp_file = SCAN_DISPLAY_FILE if engine == "nifty50" else SCAN_DISPLAY_INDEX_FILE
        eng_disp = cached_data.get("scan_display", {}).get(engine, {})
        clean_sym = str(symbol).replace(" ", "").upper()
        if isinstance(eng_disp, dict):
            for cat in ["staged_trades", "active_live", "carry_forward"]:
                for item in eng_disp.get(cat, []):
                    if isinstance(item, dict):
                        i_sym = str(item.get("symbol") or "").replace(" ", "").upper()
                        i_cnt = str(item.get("contract") or "").replace(" ", "").upper()
                        if clean_sym in (i_sym, i_cnt) or i_sym in clean_sym or i_cnt in clean_sym:
                            for k in update_keys:
                                item[k] = vals[k]
                            if "current_sl" in vals and "entry_spot" in item and item.get("entry_spot"):
                                item["rr"] = round(calc_rr(item.get("entry_spot"), vals["current_sl"], vals.get("t1", item.get("t1")), vals.get("t2", item.get("t2"))), 2)
            if os.path.exists(disp_file):
                try:
                    with open(disp_file, "w") as fh:
                        json.dump(eng_disp, fh, indent=2)
                except Exception as fe:
                    logging.warning(f"Failed to update scan display file: {fe}")

    logging.info(f"Position override queued: {engine}/{symbol} {vals}")
    return jsonify({"ok": True})

# ──────────────────────────────────────────────
#  1-CLICK BUY SCANNED TRADE API
# ──────────────────────────────────────────────
@app.route("/api/buy-scanned-trade", methods=["POST"])
def api_buy_scanned_trade():
    try:
        data = request.json or {}
        symbol = data.get("symbol")
        contract = data.get("contract") or symbol
        side = data.get("side", "CE")
        entry_spot = float(data.get("entry_spot") or 0)
        current_sl = float(data.get("current_sl") or data.get("sl") or 0)
        t1 = float(data.get("t1") or 0)
        t2 = float(data.get("t2") or 0)
        t3 = float(data.get("t3") or 0)
        engine = data.get("engine", "nifty50")
        
        if not symbol:
            return jsonify({"ok": False, "error": "symbol is required"}), 400

        c_str = str(contract).upper()
        if "SENSEX" in c_str or "BSE" in c_str:
            exch = "BFO"
        elif "CE" in c_str or "PE" in c_str or "NIFTY" in c_str or "BANK" in c_str:
            exch = "NFO"
        else:
            exch = "NSE"

        global _kite_session
        order_id = None
        if not _kite_session:
            try:
                from trading_core import load_kite_session
                api_k, acc_t = load_kite_session()
                if api_k and acc_t:
                    from kiteconnect import KiteConnect
                    _kite_session = KiteConnect(api_key=api_k)
                    _kite_session.set_access_token(acc_t)
            except Exception as init_err:
                logging.warning(f"1-Click Buy auto-init kite session failed: {init_err}")

        if _kite_session:
            try:
                q_key = f"{exch}:{contract}"
                q = _kite_session.quote([q_key])
                ltp = float(q.get(q_key, {}).get("last_price", 0))
                ask = 0
                depth = q.get(q_key, {}).get("depth", {}).get("sell", [])
                if depth and len(depth) > 0:
                    ask = float(depth[0].get("price", 0))
                price = round((ask if ask > 0 else ltp) * 1.005, 1)
                if price <= 0:
                    price = round(entry_spot * 1.005, 1)
                
                from trading_core import INDEX_REGISTRY, STOCK_REGISTRY, get_option_lot_size
                registry = INDEX_REGISTRY if engine == "index" else STOCK_REGISTRY
                lot_size = get_option_lot_size(contract) or registry.get(symbol, {}).get("lot_size", 1)
                prod = _kite_session.PRODUCT_CNC if exch == "NSE" else _kite_session.PRODUCT_NRML
                
                order_id = _kite_session.place_order(
                    variety=_kite_session.VARIETY_REGULAR,
                    tradingsymbol=contract,
                    exchange=exch,
                    transaction_type=_kite_session.TRANSACTION_TYPE_BUY,
                    quantity=lot_size,
                    order_type=_kite_session.ORDER_TYPE_LIMIT,
                    price=price,
                    product=prod
                )
                logging.info(f"[1-CLICK BUY] Placed buy order for {contract} on {exch} (Order ID: {order_id})")
            except Exception as k_err:
                logging.warning(f"[1-CLICK BUY KITE ORDER WARNING] {contract}: {k_err}")
                return jsonify({"ok": False, "error": f"Kite Order Placement Failed: {k_err}"}), 400

        tf_param = data.get("timeframe") or data.get("timeframe_entry")
        if not tf_param:
            cfg = load_config()
            tf_param = cfg.get(engine, {}).get("timeframe_entry") or "15minute"

        candle_a_time = data.get("candle_a_time") or data.get("CandleATime")
        benchmark = data.get("benchmark")
        anchor_floor = data.get("anchor_floor")
        direction = data.get("direction")

        if not candle_a_time:
            try:
                import os as _os
                disp_backfill = SCAN_DISPLAY_FILE if engine == "nifty50" else SCAN_DISPLAY_INDEX_FILE
                if _os.path.exists(disp_backfill):
                    with open(disp_backfill, "r", encoding="utf-8") as fh:
                        _sd = json.load(fh)
                    c_target = str(contract).replace(" ", "").upper()
                    for _cat in ["staged_trades", "all_staged_today", "carry_forward", "active_live"]:
                        for item in _sd.get(_cat) or []:
                            i_cnt = str(item.get("contract") or item.get("symbol") or "").replace(" ", "").upper()
                            if i_cnt == c_target:
                                candle_a_time = candle_a_time or item.get("candle_a_time") or item.get("CandleATime")
                                if benchmark is None:
                                    benchmark = item.get("benchmark")
                                if anchor_floor is None:
                                    anchor_floor = item.get("anchor_floor")
                                if not direction:
                                    direction = item.get("direction")
                                break
                        if candle_a_time:
                            break
            except Exception as bf_err:
                logging.warning(f"1-Click Buy display backfill skipped: {bf_err}")

        trade_data = {
            "contract": contract,
            "entry_spot": entry_spot,
            "current_sl": current_sl,
            "t1": t1,
            "t2": t2,
            "t3": t3,
            "side": side,
            "pattern": "1CLICK_BUY",
            "timeframe": tf_param,
            "position_type": "stock" if exch == "NSE" else "option",
            "user_edited": True,
            "entry_time": dt.now().isoformat()
        }
        if candle_a_time:
            trade_data["candle_a_time"] = candle_a_time
            trade_data["CandleATime"] = candle_a_time
        if benchmark is not None:
            trade_data["benchmark"] = benchmark
        if anchor_floor is not None:
            trade_data["anchor_floor"] = anchor_floor
        if direction:
            trade_data["direction"] = direction
        try:
            from trading_core import contract_is_expired
            if contract_is_expired(contract):
                return jsonify({"ok": False, "error": f"Contract {contract} is expired. Cannot place 1-Click Buy."}), 400
        except Exception as exp_check_err:
            logging.warning(f"1-Click Buy expiry check skipped: {exp_check_err}")
        symbol = resolve_underlying(symbol or contract, engine)
        tid, _created = trade_db.create_trade(engine, symbol, trade_data)
        clear_executed_exit(contract)

        return jsonify({
            "ok": True,
            "message": f"Successfully placed 1-Click BUY for {contract}" + (f" (Order ID: {order_id})" if order_id else ""),
            "trade_id": tid
        })
    except Exception as e:
        logging.error(f"1-Click Buy API failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

# ──────────────────────────────────────────────
#  MANUAL EXIT POSITION API (SINGLE & ALL)
# ──────────────────────────────────────────────
@app.route("/api/exit-position", methods=["POST"])
def api_exit_position():
    try:
        data = request.json or {}
        symbol = data.get("symbol", "")
        contract = data.get("contract") or symbol
        engine = data.get("engine", "nifty50")
        
        if not symbol and not contract:
            return jsonify({"ok": False, "error": "Symbol or contract name required"}), 400

        target_str = str(contract or symbol).replace(" ", "").upper()
        now_str = dt.now().strftime("%Y-%m-%d %H:%M:%S")

        all_t = trade_db.get_all_trades()
        exited_ids = []
        for t in all_t:
            t_sym = str(t.get("symbol") or "").replace(" ", "").upper()
            t_cnt = str(t.get("contract") or "").replace(" ", "").upper()
            if t.get("status") == "ACTIVE" and (target_str in (t_sym, t_cnt) or t_sym in target_str or t_cnt in target_str):
                trade_db.update_trade(t["id"], {
                    "status": "USER_EXIT",
                    "exit_time": now_str,
                    "result": "USER_EXIT",
                    "updated_at": now_str
                })
                exited_ids.append(t["id"])

        global _kite_session
        if _kite_session:
            try:
                c_str = target_str
                if "SENSEX" in c_str or "BSE" in c_str:
                    exch = "BFO"
                elif "CE" in c_str or "PE" in c_str or "NIFTY" in c_str or "BANK" in c_str:
                    exch = "NFO"
                else:
                    exch = "NSE"
                
                pos_obj = {
                    "contract": contract,
                    "symbol": symbol,
                    "exchange": exch,
                    "quantity": data.get("quantity", 0)
                }
                from trading_core import close_position as shared_close
                shared_close(_kite_session, pos_obj, True)
            except Exception as k_err:
                logging.warning(f"Live exit execution warning for {contract}: {k_err}")

        with data_lock:
            if isinstance(cached_data.get("positions"), dict):
                cached_data["positions"].pop(symbol, None)
                cached_data["positions"].pop(contract, None)

        for disp_path in [SCAN_DISPLAY_FILE, SCAN_DISPLAY_INDEX_FILE]:
            if os.path.exists(disp_path):
                try:
                    with open(disp_path, "r", encoding="utf-8") as f:
                        sd = json.load(f)
                    sd["active_positions"] = [p for p in sd.get("active_positions", []) if str(p.get("contract") or p.get("symbol")).replace(" ", "").upper() != target_str]
                    sd["active_live"] = [p for p in sd.get("active_live", []) if str(p.get("contract") or p.get("symbol")).replace(" ", "").upper() != target_str]
                    with open(disp_path, "w", encoding="utf-8") as f:
                        json.dump(sd, f, indent=2)
                except Exception as e:
                    pass

        return jsonify({"ok": True, "message": f"Manual EXIT executed for {contract or symbol}"})
    except Exception as e:
        logging.error(f"Manual Exit API failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/exit-all-positions", methods=["POST"])
def api_exit_all_positions():
    try:
        now_str = dt.now().strftime("%Y-%m-%d %H:%M:%S")
        all_t = trade_db.get_all_trades()
        exited_count = 0
        for t in all_t:
            if t.get("status") == "ACTIVE":
                trade_db.update_trade(t["id"], {
                    "status": "USER_EXIT",
                    "exit_time": now_str,
                    "result": "USER_EXIT",
                    "updated_at": now_str
                })
                exited_count += 1

        global _kite_session
        if _kite_session:
            try:
                kp = _kite_session.positions()
                for p in (kp.get("day", []) + kp.get("net", [])):
                    if abs(int(p.get("quantity", 0))) > 0:
                        contract = p.get("tradingsymbol")
                        pos_obj = {"contract": contract, "quantity": abs(int(p.get("quantity", 0)))}
                        from trading_core import close_position as shared_close
                        shared_close(_kite_session, pos_obj, True)
            except Exception as k_err:
                logging.warning(f"Exit all live positions warning: {k_err}")

        for disp_path in [SCAN_DISPLAY_FILE, SCAN_DISPLAY_INDEX_FILE]:
            if os.path.exists(disp_path):
                try:
                    with open(disp_path, "r", encoding="utf-8") as f:
                        sd = json.load(f)
                    sd["active_positions"] = []
                    sd["active_live"] = []
                    with open(disp_path, "w", encoding="utf-8") as f:
                        json.dump(sd, f, indent=2)
                except Exception as e:
                    pass

        return jsonify({"ok": True, "message": f"Manual EXIT ALL executed for {exited_count} position(s)"})
    except Exception as e:
        logging.error(f"Manual Exit All API failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

# ──────────────────────────────────────────────
#  INTERACTIVE NEGATION ANALYZER API
# ──────────────────────────────────────────────
@app.route("/api/analyze-trade", methods=["POST"])
def api_analyze_trade():
    try:
        data = request.json or {}
        symbol = str(data.get("symbol", "")).strip().upper()
        entry_price = float(data.get("entry_price", 0)) if data.get("entry_price") else 0.0
        timeframe = str(data.get("timeframe", "75min")).strip()
        engine = str(data.get("engine", "nifty50")).strip()
        
        if not symbol:
            return jsonify({"ok": False, "error": "Valid Symbol or Contract Name required"}), 400

        kite = None
        try:
            api_k, acc_t = load_kite_session()
            kite = KiteConnect(api_key=api_k, access_token=acc_t)
        except Exception:
            kite = None

        timeframe_entry = timeframe
        timeframe_anchor = timeframe

        analysis = derive_sl_targets_for_contract(kite, symbol, entry_price, timeframe_entry, timeframe_anchor)
        if not analysis:
            sl_val = round(entry_price * 0.90, 2) if entry_price > 0 else 0.0
            analysis = {
                "entry_price": entry_price,
                "current_sl": sl_val,
                "t1": None, "t2": None, "t3": None,
                "pattern": "NEGATION_ESTIMATED"
            }

        resolved_entry = float(analysis.get("entry_price") or entry_price or 0.0)
        sl_val = analysis.get("current_sl", round(resolved_entry * 0.90, 2) if resolved_entry > 0 else 0.0)
        t1_val = analysis.get("t1")
        t2_val = analysis.get("t2")
        t3_val = analysis.get("t3")
        
        risk = (resolved_entry - sl_val) if (resolved_entry > 0 and sl_val < resolved_entry) else 0
        rr = round((t1_val - resolved_entry) / risk, 2) if (t1_val and risk > 0) else 0.0

        return jsonify({
            "ok": True,
            "symbol": symbol,
            "contract": symbol,
            "entry_price": resolved_entry,
            "current_sl": sl_val,
            "t1": t1_val if t1_val else "N/A",
            "t2": t2_val if t2_val else "N/A",
            "t3": t3_val if t3_val else "N/A",
            "rr": rr,
            "pattern": analysis.get("pattern", "NEGATION_DERIVED"),
            "engine": engine
        })
    except Exception as e:
        logging.error(f"Analyze Trade API failed: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

# ──────────────────────────────────────────────
#  DAILY SELF-LEARNING TRADE JOURNAL API
# ──────────────────────────────────────────────
@app.route("/api/journal/get", methods=["GET"])
def api_journal_get():
    try:
        from daily_trade_journal import load_journal_entries
        return jsonify(load_journal_entries())
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/journal/sync", methods=["POST"])
def api_journal_sync():
    try:
        from daily_trade_journal import generate_daily_journal
        req = request.json or {}
        dt_str = req.get("date")
        entries = generate_daily_journal(dt_str, kite=_kite_session)
        return jsonify({"ok": True, "count": len(entries), "entries": entries})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/journal/update", methods=["POST"])
def api_journal_update():
    try:
        from daily_trade_journal import load_journal_entries, save_journal_entries
        data = request.json or {}
        symbol = data.get("symbol")
        date_str = data.get("date")
        trade_id = data.get("trade_id")
        remarks = data.get("remarks")
        lesson = data.get("lesson")
        if trade_id and lesson is not None:
            trade_db.update_self_learning_lesson(trade_id, lesson)
        if not symbol or not date_str:
            return jsonify({"ok": True, "message": "Updated trade_db lesson"})
        entries = load_journal_entries()
        updated = False
        for e in entries:
            if e.get("Date") == date_str and (e.get("Symbol") == symbol or symbol in e.get("Symbol", "")):
                if remarks is not None: e["Analysis_Remarks"] = remarks
                if lesson is not None: e["Self_Learning_Lesson"] = lesson
                updated = True
        if updated:
            save_journal_entries(entries)
            return jsonify({"ok": True})
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/journal/export", methods=["GET", "POST"])
def api_journal_export():
    try:
        import io
        from daily_trade_journal import load_journal_entries
        entries = load_journal_entries()
        headers = [
            "Date", "Engine", "Symbol", "Side", "Timeframe", "Pattern",
            "Entry_Time", "Entry_Price", "Exit_Time", "Exit_Price",
            "SL", "T1", "T2", "T3", "Quantity", "Lot_Size",
            "PnL_Rs", "PnL_Pct", "Outcome", "Analysis_Remarks", "Self_Learning_Lesson"
        ]
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for e in entries:
            writer.writerow(e)
        csv_data = output.getvalue()
        fname = f"trade_journal_export_{dt.now().strftime('%Y_%m_%d_%H%M')}.csv"
        return Response(
            csv_data,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={fname}"}
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/journal/export/excel", methods=["GET", "POST"])
def api_journal_export_excel():
    try:
        import openpyxl, io
        from openpyxl.styles import Font, PatternFill, Alignment
        from daily_trade_journal import load_journal_entries

        entries = load_journal_entries()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trade Journal"

        headers = [
            "Date", "Engine", "Symbol", "Side", "Timeframe", "Pattern",
            "Entry Time", "Entry Price", "Exit Time", "Exit Price",
            "SL", "T1", "T2", "T3", "Quantity", "Lot Size",
            "PnL (₹)", "PnL (%)", "Outcome", "Analysis Remarks", "Self-Learning Lesson"
        ]
        
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for e in entries:
            ws.append([
                e.get("Date", ""),
                e.get("Engine", ""),
                e.get("Symbol", ""),
                e.get("Side", ""),
                e.get("Timeframe", ""),
                e.get("Pattern", ""),
                e.get("Entry_Time", ""),
                e.get("Entry_Price", ""),
                e.get("Exit_Time", ""),
                e.get("Exit_Price", ""),
                e.get("SL", ""),
                e.get("T1", ""),
                e.get("T2", ""),
                e.get("T3", ""),
                e.get("Quantity", ""),
                e.get("Lot_Size", ""),
                e.get("PnL_Rs", 0),
                e.get("PnL_Pct", ""),
                e.get("Outcome", ""),
                e.get("Analysis_Remarks", ""),
                e.get("Self_Learning_Lesson", "")
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        ws2 = wb.create_sheet(title="Performance Summary")
        ws2.append(["Metric", "Value"])
        ws2.cell(row=1, column=1).font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=2).font = header_font
        ws2.cell(row=1, column=2).fill = header_fill

        total_trades = len(entries)
        wins = sum(1 for e in entries if float(e.get("PnL_Rs") or 0) > 0)
        losses = sum(1 for e in entries if float(e.get("PnL_Rs") or 0) < 0)
        win_rate = (wins / total_trades * 100) if total_trades > 0 else 0
        total_pnl = sum(float(e.get("PnL_Rs") or 0) for e in entries)

        ws2.append(["Total Trades", total_trades])
        ws2.append(["Winning Trades", wins])
        ws2.append(["Losing Trades", losses])
        ws2.append(["Win Rate (%)", f"{win_rate:.1f}%"])
        ws2.append(["Total Net PnL (₹)", f"₹{total_pnl:.2f}"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        fname = f"trade_journal_export_{dt.now().strftime('%Y_%m_%d_%H%M')}.xlsx"
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname}"}
        )
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500




@app.route("/api/get-chart-data", methods=["GET"])
def api_get_chart_data():
    try:
        contract = str(request.args.get("symbol", "")).strip().upper()
        chart_type = str(request.args.get("type", "option")).strip().lower()
        tf = str(request.args.get("timeframe", "30minute")).strip()

        if not contract:
            return jsonify({"ok": False, "error": "Symbol is required"}), 400

        api_k, acc_t = load_kite_session()
        kite = KiteConnect(api_key=api_k, access_token=acc_t)

        token = None
        exchange = "NFO"
        spot_symbol = contract
        spot_token = None

        nfo_cache_path = os.path.join(BASE_DIR, "output", "monitor", "nfo_instruments_cache.csv")
        if os.path.exists(nfo_cache_path):
            try:
                import pandas as pd
                nfo_df = pd.read_csv(nfo_cache_path)
                opt_rows = nfo_df[nfo_df["tradingsymbol"] == contract]
                if not opt_rows.empty:
                    token = int(opt_rows.iloc[0]["instrument_token"])
                    spot_symbol = str(opt_rows.iloc[0]["name"]).strip().upper()
            except Exception as e:
                logging.warning(f"NFO cache lookup failed: {e}")

        from trading_core import STOCK_REGISTRY
        if spot_symbol in STOCK_REGISTRY:
            spot_token = STOCK_REGISTRY[spot_symbol]["token"]
        elif spot_symbol in ["NIFTY", "NIFTY 50", "NIFTY50"]:
            spot_symbol = "NIFTY"
            spot_token = 256265
        elif spot_symbol in ["BANKNIFTY", "NIFTY BANK"]:
            spot_symbol = "BANKNIFTY"
            spot_token = 260105
        elif spot_symbol in ["SENSEX", "BSESN"]:
            spot_symbol = "SENSEX"
            spot_token = 265
        elif not token and contract in STOCK_REGISTRY:
            spot_token = STOCK_REGISTRY[contract]["token"]
            spot_symbol = contract

        if chart_type == "spot":
            try:
                ltp_res = kite.ltp([f"NSE:{spot_symbol}"])
                if ltp_res and f"NSE:{spot_symbol}" in ltp_res:
                    spot_token = ltp_res[f"NSE:{spot_symbol}"]["instrument_token"]
            except Exception:
                pass
            target_token = spot_token or token
            target_symbol = spot_symbol
            target_exchange = "NSE" if (spot_symbol in STOCK_REGISTRY or spot_symbol in ["NIFTY", "BANKNIFTY"]) else "BSE"
        else:
            target_token = token or spot_token
            target_symbol = contract if token else spot_symbol
            target_exchange = "NFO" if token else "NSE"

        if not target_token:
            return jsonify({"ok": False, "error": f"Instrument token not found for {contract}"}), 400

        from datetime import datetime as dt, timedelta
        from trading_core import fetch_and_resample_candles

        ref_now = dt.now()
        from_date = (ref_now - timedelta(days=60)).strftime("%Y-%m-%d")
        to_date = ref_now.strftime("%Y-%m-%d")

        df_candles = fetch_and_resample_candles(kite, target_token, from_date, to_date, tf)
        if df_candles is None or df_candles.empty:
            return jsonify({"ok": False, "error": f"No candle data available for {target_symbol}"}), 400

        import pandas as pd
        candles = []
        for _, r in df_candles.iterrows():
            c_dt = pd.to_datetime(r['date'])
            ts = int(c_dt.timestamp())
            candles.append({
                "time": ts,
                "open": round(float(r['open']), 2),
                "high": round(float(r['high']), 2),
                "low": round(float(r['low']), 2),
                "close": round(float(r['close']), 2),
                "volume": int(r['volume']) if 'volume' in r else 0
            })

        kite_url = f"https://kite.zerodha.com/chart/ext/tvc/{target_exchange}/{target_symbol}/{target_token}"

        return jsonify({
            "ok": True,
            "symbol": target_symbol,
            "contract": contract,
            "spot_symbol": spot_symbol,
            "chart_type": chart_type,
            "exchange": target_exchange,
            "token": target_token,
            "timeframe": tf,
            "candles": candles,
            "kite_url": kite_url
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

EXPORT_STATE_FILE = "output/monitor/export_state.json"

# ──────────────────────────────────────────────
#  MONTHLY EXPORT (trades to Excel archive)
# ──────────────────────────────────────────────

def run_monthly_export():
    import openpyxl
    from collections import defaultdict
    completed = trade_db.get_completed_trades()
    if not completed:
        return {"exported": 0, "sheets": []}
    groups = defaultdict(list)
    for t in completed:
        ts = t.get("exit_time") or t.get("updated_at") or t.get("created_at") or ""
        parts = ts.split(" ")[0].split("-") if " " in ts else ts.split("-")
        key = (parts[0], parts[1]) if len(parts) >= 2 else ("unknown", "00")
        groups[key].append(t)
    out_dir = "output/exports"
    os.makedirs(out_dir, exist_ok=True)
    xl_path = os.path.join(out_dir, "trade_archive.xlsx")
    sheet_names = []
    if os.path.exists(xl_path):
        wb = openpyxl.load_workbook(xl_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    headers = ["ID", "Engine", "Symbol", "Pattern", "Status", "Entry Spot", "SL", "T1", "T2", "T3",
               "Trailing Stage", "Lot Size", "Position Size", "Entry Date", "Exit Date", "P&L %", "Side", "Contract"]
    for (year, month), trades in sorted(groups.items()):
        month_names = ["", "January","February","March","April","May","June","July","August","September","October","November","December"]
        sheet_name = f"{month_names[int(month)]} {year}" if month.isdigit() and 1 <= int(month) <= 12 else f"{month} {year}"
        sheet_names.append(sheet_name)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for t in trades:
            entry_dt = (t.get("created_at") or "").split(" ")[0]
            exit_dt = (t.get("exit_time") or t.get("updated_at") or "").split(" ")[0]
            ws.append([
                t.get("id", ""), t.get("engine", ""), t.get("symbol", ""), t.get("pattern", ""),
                t.get("status", ""), t.get("entry_spot", ""), t.get("current_sl", ""),
                t.get("t1", ""), t.get("t2", ""), t.get("t3", ""),
                t.get("trailing_stage", ""), t.get("lot_size", ""), t.get("position_size", ""),
                entry_dt, exit_dt,
                t.get("pnl_percent", ""), t.get("side", ""), t.get("contract", "")
            ])
    wb.save(xl_path)
    exported_ids = [t["id"] for t in completed]
    trade_db.remove_trades(exported_ids)
    return {"exported": len(completed), "sheets": sheet_names}

def auto_export_if_new_month():
    now = dt.now()
    current_month = now.strftime("%Y-%m")
    try:
        if os.path.exists(EXPORT_STATE_FILE):
            with open(EXPORT_STATE_FILE) as f:
                state = json.load(f)
                last = state.get("last_export_month", "")
        else:
            last = ""
        if current_month > last:
            result = run_monthly_export()
            if result["exported"] > 0:
                print(f"Auto-export: {result['exported']} trades to {', '.join(result['sheets'])}")
            with open(EXPORT_STATE_FILE, "w") as f:
                json.dump({"last_export_month": current_month}, f)
    except Exception as e:
        print(f"Auto-export error: {e}")

_last_eod_journal_triggered_date = None

def auto_eod_journal_scheduler():
    global _last_eod_journal_triggered_date
    while True:
        try:
            now = dt.now()
            today_str = now.strftime("%Y-%m-%d")
            # Trigger once daily at/after 15:35 IST on weekdays (Mon-Fri)
            if now.weekday() < 5 and (now.hour > 15 or (now.hour == 15 and now.minute >= 35)):
                if _last_eod_journal_triggered_date != today_str:
                    _last_eod_journal_triggered_date = today_str
                    logging.info(f"[AUTO EOD JOURNAL] Market closed. Auto-generating EOD trade journal for {today_str}...")
                    from daily_trade_journal import generate_daily_journal
                    generate_daily_journal(target_date=today_str, kite=_kite_session)
                    logging.info(f"[AUTO EOD JOURNAL] Successfully completed EOD trade journal sync for {today_str}.")
        except Exception as e:
            logging.warning(f"[AUTO EOD JOURNAL] Error in scheduler: {e}")
        time.sleep(60)

def main():
    os.makedirs("input", exist_ok=True)
    os.makedirs("output/logs", exist_ok=True)
    os.makedirs("output/monitor", exist_ok=True)
    os.makedirs("logs", exist_ok=True)
    auto_export_if_new_month()
    try:
        refresh_data(single_run=True)
    except Exception as e:
        logging.warning(f"Initial position pre-fetch warning: {e}")
    worker = threading.Thread(target=refresh_data, daemon=True)
    worker.start()
    eod_worker = threading.Thread(target=auto_eod_journal_scheduler, daemon=True)
    eod_worker.start()
    print(f"Trading Control Center starting on http://localhost:{DASHBOARD_PORT}")
    print(f"Refresh interval: {REFRESH_SECONDS}s")
    print("Available programs:")
    for pid, p in PROGRAMS.items():
        print(f"  [{pid}] {p['name']}")
    app.run(host="0.0.0.0", port=DASHBOARD_PORT, debug=False, use_reloader=False)

if __name__ == "__main__":
    main()
